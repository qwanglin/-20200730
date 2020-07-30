# -*- coding: utf-8 -*- 

"""
# @Time : 2019/4/4 
# @Author : xucaimin
"""


import csv
import xlwt
import pandas as pd
import os
from openpyxl import Workbook

#excle转换为csv
def xlsx_to_csv(xlsxpath,csvpath):
    data_xls = pd.read_excel(xlsxpath, index_col=0)
    data_xls.to_csv(csvpath, encoding='utf-8')

#csv转换为excel(这样可以避免实验xlwt写入execl遇到写入限制的尴尬）
def csv_to_xlsx(csvpath,xlsxpath):

    if (os.path.exists(xlsxpath) == True):
        os.remove(xlsxpath)
    with open(csvpath, 'r', encoding='utf-8') as f:
        read = csv.reader(f)
        wb = Workbook()
        sheet = wb.active  # 激活sheet
        sheet.title="data"  #对sheet进行命名
        l = 1
        for line in read:
            # print(line)
            r = 1
            for i in line:
                # sheet.write(l, r, i)  # 一个一个将单元格数据写入
                sheet.cell(l, r).value=i# 一个一个将单元格数据写入
                r = r + 1
            l = l + 1
        wb.save(xlsxpath)  # 保存Excel


#csv转换为excel（当行数过多的时候，会报错）
def csv_to_xlsx2(csvpath,xlsxpath):
    # 判断一下文件夹下的文件是否存在，存在就删除
    if (os.path.exists(xlsxpath) == True):
        os.remove(xlsxpath)
    with open(csvpath, 'r', encoding='utf-8') as f:
        read = csv.reader(f)
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('data')  # 创建一个sheet表格
        l = 0
        for line in read:
            # print(line)
            r = 0
            for i in line:
                sheet.write(l, r, i)  # 一个一个将单元格数据写入
                r = r + 1
            l = l + 1
        workbook.save(xlsxpath)  # 保存Excel



if __name__ == '__main__':
    csvpath="E:\\PythonWorkspace\\ZentaoTest/result\\20190416 100208/bugs.csv"
    excelpath="E:\\PythonWorkspace\\ZentaoTest/result\\20190416 100208/bugs.xls"
    csv_to_xlsx2(csvpath,excelpath)
