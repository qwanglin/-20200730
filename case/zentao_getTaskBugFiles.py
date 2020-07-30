# -*- coding: utf-8 -*- 

"""
# @Time : 2019/3/29 
# @Author : xucaimin
"""
import sys
sys.path.append("../")
from common.configPar import *
from selenium import webdriver
import time
from common.WriteToTxt import *
from common.zentaoInterface import _login,_getiterate,_getProduct,_getTasks,_getDepartment,_getIteraBugId,_getAllClosediterate,\
    _getBugActivityLastTime,_getVersions,_getTestCases,getHoliday,getManyResponseTask
from selenium.webdriver.common.action_chains import ActionChains
import os
from datetime import datetime
from common.zentao_login import *
import csv
from common.getfileName import *
import pandas as pd
import glob
from common.csv__xlsx import *
from xlwt import *
import xlrd
from xlutils.copy import copy
from common.commonUtil import *
import openpyxl
import  datetime as dt
from bs4 import BeautifulSoup
from common.seleniumUtil import *
url=getConfig("DEFAULT","url")

#得到bug文件
def getBugFiles(driver):
    computerName = os.getlogin()
    downloadPath = "C:\\Users\\" + computerName + "\\Downloads"
    path = os.path.dirname(os.getcwd()) + '/file'
    Folder_Path = new_file(path)  # 得到下载任务文件最新的文件夹

    driver.find_element_by_link_text("测试").click()
    logWriteToTxt("点击测试")
    driver.find_element_by_link_text("Bug").click()
    logWriteToTxt("点击Bug")
    time.sleep(2)
    currentItemText = driver.find_element_by_id("currentItem").text
    driver.find_element_by_id("currentItem").click()
    logWriteToTxt("点击当前产品:" + currentItemText)
    time.sleep(2)
    eles = driver.find_elements_by_xpath("//*[@id='dropMenu']/div[2]/div/div[1]/div[1]/a")
    elelen = len(eles)
    print("多少:",elelen)
    for m in range(1,(elelen+1)):
        if(m!=1):
            driver.refresh()  # 页面刷新
            time.sleep(2)
            currentItemText = driver.find_element_by_id("currentItem").text
            driver.find_element_by_id("currentItem").click()
            logWriteToTxt("点击当前产品:" + currentItemText)
        time.sleep(8)
        driver.find_element_by_xpath("//*[@id='dropMenu']/div[2]/div/div[1]/div[1]/a["+str(m)+"]").click()
        currentItemText = driver.find_element_by_id("currentItem").text
        logWriteToTxt('已定位到元素，并且点击该产品:' + currentItemText)
        # print(m, currentItemText)
        time.sleep(3)
        driver.find_element_by_xpath("//*[@id='mainMenu']/div[2]/a[1]").click()
        logWriteToTxt("点击所有")
        time.sleep(5)
        bugelement = driver.find_element_by_xpath("//*[@id='mainMenu']/div[3]/div/button")
        # ActionChains(driver).move_to_element(bugelement).perform()
        bugelement.click()
        logWriteToTxt("点击导出按钮")
        time.sleep(2)
        while 1:
            try:
                driver.find_element_by_link_text("导出数据").click()
                logWriteToTxt("点击导出数据按钮")
                time.sleep(4)
                break
            except:
                logWriteToTxt("还未定位到导出数据元素!")
        driver.switch_to.frame(driver.find_element_by_id("iframe-triggerModal"))
        time.sleep(1)
        while 1:
            try:
                driver.find_element_by_id("submit").click()
                logWriteToTxt("点击导出按钮")
                time.sleep(15)
                break
            except:
                logWriteToTxt("还未定位到导出按钮!")
                time.sleep(1)
        time.sleep(10)
        copy__file(get_str_new_file(downloadPath,".csv"), Folder_Path)
    driver.close()

#得到一件关闭的迭代任务
def getClosedTaskFiles(driver,session,isHalfYear):

    closedTimes=_getAllClosediterate(session)
    computerName = os.getlogin()
    downloadPath = "C:\\Users\\" + computerName + "\\Downloads"
    path = os.path.dirname(os.getcwd()) + '/file'
    Folder_Path = new_file(path)  # 得到下载任务文件最新的文件夹

    today = time.strftime("%Y/%m/%d")  # 今天
    day_num = int(today.split('/')[2])  # 几号
    month_num = int(today.split('/')[1])  # 几月
    year_num = int(today.split('/')[0])  # 几年
    flag = is_leap_year(year_num)
    if (flag == True):  # 该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    if (int(isHalfYear) == 1):
        starttime = time.strftime(str(year_num) + "-" + str(1) + "-" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num) + "-" + str(6) + "-" + str(days[5]))  # 结束日期
        print("前半年任务starttime,endtime:", starttime, endtime)
    elif (int(isHalfYear) == 2):
        starttime = time.strftime(str(year_num) + "-" + str(7) + "-" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num) + "-" + str(12) + "-" + str(days[11]))  # 结束日期
        print("后半年任务starttime,endtime:", starttime, endtime)
    strftime = datetime.datetime.strptime(starttime, "%Y-%m-%d")
    strftime2 = datetime.datetime.strptime(endtime, "%Y-%m-%d")
    for i in range(0,len(closedTimes)):
        closetime=closedTimes[i]
        if(closetime!="无"):
            # print("i,closetime:", i, closetime)
            strftime3 = datetime.datetime.strptime(closetime, "%Y-%m-%d")
            if ((strftime <= strftime3) & (strftime2 >= strftime3)):
                # driver.refresh()  # 页面刷新
                driver.get(url+"/zentao/project-all-closed-51-order_desc-0-"+str(len(closedTimes))+"-2000-1.html")
                # time.sleep(3)
                # driver.find_element_by_link_text("迭代").click()
                # logWriteToTxt("点击迭代")
                # driver.find_element_by_link_text("任务").click()
                # logWriteToTxt("点击任务")
                # driver.find_element_by_xpath("//*[@id='pageNav']/div[1]/div/button").click()
                # logWriteToTxt("点击左上角迭代主页")
                # time.sleep(2)
                # driver.find_element_by_xpath("//*[@id='pageNav']/div[1]/div/ul/li[2]/a").click()
                # logWriteToTxt("点击所有迭代")
                # time.sleep(2)
                # driver.find_element_by_xpath("//*[@id='closedTab']/span").click()
                # logWriteToTxt("点击已关闭tab")
                # time.sleep(2)
                # if(i>9):##这时候就要翻页了
                #     count=int(i/10)+1 #第几页
                #     for j in range(1,count):
                #         driver.find_element_by_class_name("icon-angle-right").click()
                #         logWriteToTxt("点击下一页")
                #         time.sleep(2)
                # driver.find_element_by_xpath("//*[@id='projectTableList']/tr[" + str(i%10 + 1) + "]/td[2]/a").click()
                from selenium.webdriver.common.keys import Keys
                driver.find_element_by_css_selector(
                    # "#projectTableList > tr:nth-child(" + str(i % 10 + 1) + ") > td:nth-child(2) > a").send_keys(
                    "#projectTableList > tr:nth-child(" + str(i+1) + ") > td:nth-child(2) > a").send_keys(
                    Keys.ENTER)
                print("选择一个任务 "+str(i))
                logWriteToTxt("选择一个任务")
                time.sleep(2)
                # driver.find_elements_by_class_name("dropdown-toggle")[1].click()
                # logWriteToTxt("点击每页几项弹出框")
                # time.sleep(3)
                # driver.find_element_by_xpath("//*[@id='projectsForm']/div/ul/li[2]/div/ul/li[15]").click()
                # logWriteToTxt("选择每页2000项")
                # time.sleep(3)
                # driver.find_element_by_xpath("//*[@id='projectTableList']/tr["+str(i+1)+"]/td[2]/a").click()
                # logWriteToTxt("选择一个任务")
                # time.sleep(2)
                driver.find_element_by_link_text("任务").click()
                logWriteToTxt("点击任务")
                time.sleep(1)
                driver.find_element_by_id("all").click()
                logWriteToTxt("点击所有")
                time.sleep(5)
                taskelement = driver.find_element_by_xpath("//*[@id='mainMenu']/div[3]/div[1]/button")
                # 悬停在导出-元素上//*[@id="mainMenu"]/div[3]/div[1]/button
                ActionChains(driver).move_to_element(taskelement).perform()
                logWriteToTxt("鼠标悬停在导出按钮上")
                time.sleep(4)
                while 1:
                    try:
                        driver.find_element_by_link_text("导出数据").click()
                        logWriteToTxt("点击导出数据按钮")
                        time.sleep(4)
                        break
                    except:
                        taskelement = driver.find_element_by_xpath("//*[@id='mainMenu']/div[3]/div[1]/button")
                        # 悬停在导出-元素上//*[@id="mainMenu"]/div[3]/div[1]/button
                        ActionChains(driver).move_to_element(taskelement).perform()
                        logWriteToTxt("鼠标悬停在导出按钮上")
                driver.switch_to.frame(driver.find_element_by_id("iframe-triggerModal"))
                time.sleep(1)
                while 1:
                    try:
                        driver.find_element_by_id("submit").click()
                        logWriteToTxt("点击导出按钮")
                        time.sleep(20)
                        break
                    except:
                        logWriteToTxt("还未定位到导出按钮!")
                        time.sleep(1)
                time.sleep(14)
                copy__file(get_str_new_file(downloadPath,".csv"), Folder_Path)

        else:
            # print("i,closetime:", i, closetime)
            # driver.refresh()  # 页面刷新
            driver.get(url + "/zentao/project-all-closed-51-order_desc-0-" + str(len(closedTimes)) + "-2000-1.html")
            time.sleep(3)
            # driver.find_element_by_link_text("迭代").click()
            # logWriteToTxt("点击迭代")
            # driver.find_element_by_link_text("任务").click()
            # logWriteToTxt("点击任务")
            # driver.find_element_by_xpath("//*[@id='pageNav']/div[1]/div/button").click()
            # logWriteToTxt("点击左上角迭代主页")
            # time.sleep(2)
            # driver.find_element_by_xpath("//*[@id='pageNav']/div[1]/div/ul/li[2]/a").click()
            # logWriteToTxt("点击所有迭代")
            # time.sleep(2)
            # driver.find_element_by_xpath("//*[@id='closedTab']/span").click()
            # logWriteToTxt("点击已关闭tab")
            # time.sleep(2)
            # if (i > 9):  ##这时候就要翻页了
            #     count = int(i / 10) + 1  # 第几页
            #     for j in range(1, count):
            #         driver.find_element_by_class_name("icon-angle-right").click()
            #         logWriteToTxt("点击下一页")
            #         time.sleep(2)
            # driver.find_element_by_xpath("//*[@id='projectTableList']/tr[" + str(i % 10 + 1) + "]/td[2]/a").click()
            from selenium.webdriver.common.keys import Keys
            driver.find_element_by_css_selector(
                # "#projectTableList > tr:nth-child(" + str(i % 10 + 1) + ") > td:nth-child(2) > a").send_keys(
                "#projectTableList > tr:nth-child(" + str(i + 1) + ") > td:nth-child(2) > a").send_keys(
                Keys.ENTER)
            print("选择一个任务 " + str(i))
            logWriteToTxt("选择一个任务")
            time.sleep(2)
            # driver.find_elements_by_class_name("dropdown-toggle")[1].click()
            # logWriteToTxt("点击每页几项弹出框")
            # time.sleep(3)
            # driver.find_element_by_xpath("//*[@id='projectsForm']/div/ul/li[2]/div/ul/li[15]").click()
            # logWriteToTxt("选择每页2000项")
            # time.sleep(3)
            # driver.find_element_by_xpath("//*[@id='projectTableList']/tr[" + str(i+1) + "]/td[2]/a").click()
            # logWriteToTxt("选择一个任务")
            # time.sleep(2)
            driver.find_element_by_link_text("任务").click()
            logWriteToTxt("点击任务")
            time.sleep(1)
            driver.find_element_by_id("all").click()
            logWriteToTxt("点击所有")
            time.sleep(5)
            taskelement = driver.find_element_by_xpath("//*[@id='mainMenu']/div[3]/div[1]/button")
            # 悬停在导出-元素上
            ActionChains(driver).move_to_element(taskelement).perform()
            logWriteToTxt("鼠标悬停在导出按钮上")
            time.sleep(4)
            while 1:
                try:
                    driver.find_element_by_link_text("导出数据").click()
                    logWriteToTxt("点击导出数据按钮")
                    time.sleep(4)
                    break
                except:
                    taskelement = driver.find_element_by_xpath("//*[@id='mainMenu']/div[3]/div[1]/button")
                    # 悬停在导出-元素上
                    ActionChains(driver).move_to_element(taskelement).perform()
                    logWriteToTxt("鼠标悬停在导出按钮上")
            driver.switch_to.frame(driver.find_element_by_id("iframe-triggerModal"))
            time.sleep(1)
            while 1:
                try:
                    driver.find_element_by_id("submit").click()
                    logWriteToTxt("点击导出按钮")
                    time.sleep(20)
                    break
                except:
                    logWriteToTxt("还未定位到导出按钮!")
                    time.sleep(1)
            time.sleep(14)
            copy__file(get_str_new_file(downloadPath,".csv"), Folder_Path)
    driver.close()


#得到一件关闭的迭代任务
def getClosedTaskFiles2(driver,session,isHalfYear):

    closedTimes=_getAllClosediterate(session)
    computerName = os.getlogin()
    downloadPath = "C:\\Users\\" + computerName + "\\Downloads"
    path = os.path.dirname(os.getcwd()) + '/file'
    Folder_Path = new_file(path)  # 得到下载任务文件最新的文件夹

    today = time.strftime("%Y/%m/%d")  # 今天
    day_num = int(today.split('/')[2])  # 几号
    month_num = int(today.split('/')[1])  # 几月
    year_num = int(today.split('/')[0])  # 几年
    flag = is_leap_year(year_num)
    if (flag == True):  # 该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    if (int(isHalfYear) == 1):
        starttime = time.strftime(str(year_num) + "-" + str(1) + "-" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num) + "-" + str(6) + "-" + str(days[5]))  # 结束日期
        print("前半年任务starttime,endtime:", starttime, endtime)
    elif (int(isHalfYear) == 2):
        starttime = time.strftime(str(year_num) + "-" + str(7) + "-" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num) + "-" + str(12) + "-" + str(days[11]))  # 结束日期
        print("后半年任务starttime,endtime:", starttime, endtime)
    strftime = datetime.datetime.strptime(starttime, "%Y-%m-%d")
    strftime2 = datetime.datetime.strptime(endtime, "%Y-%m-%d")
    for i in range(0,len(closedTimes)):
        closetime=closedTimes[i]
        if(closetime!="无"):
            # print("i,closetime:", i, closetime)
            strftime3 = datetime.datetime.strptime(closetime, "%Y-%m-%d")
            if ((strftime <= strftime3) & (strftime2 >= strftime3)):
                driver.refresh()  # 页面刷新
                time.sleep(3)
                driver.find_element_by_link_text("迭代").click()
                logWriteToTxt("点击迭代")
                driver.find_element_by_link_text("任务").click()
                logWriteToTxt("点击任务")
                driver.find_element_by_xpath("//*[@id='pageNav']/div[1]/div/button").click()
                logWriteToTxt("点击左上角迭代主页")
                time.sleep(2)
                driver.find_element_by_xpath("//*[@id='pageNav']/div[1]/div/ul/li[2]/a").click()
                logWriteToTxt("点击所有迭代")
                time.sleep(2)
                driver.find_element_by_xpath("//*[@id='closedTab']/span").click()
                logWriteToTxt("点击已关闭tab")
                time.sleep(2)
                if(i>9):##这时候就要翻页了
                    count=int(i/10)+1 #第几页
                    for j in range(1,count):
                        driver.find_element_by_class_name("icon-angle-right").click()
                        logWriteToTxt("点击下一页")
                        time.sleep(2)
                # driver.find_element_by_xpath("//*[@id='projectTableList']/tr[" + str(i%10 + 1) + "]/td[2]/a").click()
                from selenium.webdriver.common.keys import Keys
                driver.find_element_by_css_selector(
                    "#projectTableList > tr:nth-child(" + str(i % 10 + 1) + ") > td:nth-child(2) > a").send_keys(
                    Keys.ENTER)
                print("选择一个任务")
                logWriteToTxt("选择一个任务")
                time.sleep(2)
                # driver.find_elements_by_class_name("dropdown-toggle")[1].click()
                # logWriteToTxt("点击每页几项弹出框")
                # time.sleep(3)
                # driver.find_element_by_xpath("//*[@id='projectsForm']/div/ul/li[2]/div/ul/li[15]").click()
                # logWriteToTxt("选择每页2000项")
                # time.sleep(3)
                # driver.find_element_by_xpath("//*[@id='projectTableList']/tr["+str(i+1)+"]/td[2]/a").click()
                # logWriteToTxt("选择一个任务")
                # time.sleep(2)
                driver.find_element_by_link_text("任务").click()
                logWriteToTxt("点击任务")
                time.sleep(1)
                driver.find_element_by_id("all").click()
                logWriteToTxt("点击所有")
                time.sleep(5)
                taskelement = driver.find_element_by_xpath("//*[@id='mainMenu']/div[3]/div[1]/button")
                # 悬停在导出-元素上//*[@id="mainMenu"]/div[3]/div[1]/button
                ActionChains(driver).move_to_element(taskelement).perform()
                logWriteToTxt("鼠标悬停在导出按钮上")
                time.sleep(4)
                while 1:
                    try:
                        driver.find_element_by_link_text("导出数据").click()
                        logWriteToTxt("点击导出数据按钮")
                        time.sleep(4)
                        break
                    except:
                        taskelement = driver.find_element_by_xpath("//*[@id='mainMenu']/div[3]/div[1]/button")
                        # 悬停在导出-元素上//*[@id="mainMenu"]/div[3]/div[1]/button
                        ActionChains(driver).move_to_element(taskelement).perform()
                        logWriteToTxt("鼠标悬停在导出按钮上")
                driver.switch_to.frame(driver.find_element_by_id("iframe-triggerModal"))
                time.sleep(1)
                while 1:
                    try:
                        driver.find_element_by_id("submit").click()
                        logWriteToTxt("点击导出按钮")
                        time.sleep(20)
                        break
                    except:
                        logWriteToTxt("还未定位到导出按钮!")
                        time.sleep(1)
                time.sleep(14)
                copy__file(get_str_new_file(downloadPath,".csv"), Folder_Path)

        else:
            # print("i,closetime:", i, closetime)
            driver.refresh()  # 页面刷新
            time.sleep(3)
            driver.find_element_by_link_text("迭代").click()
            logWriteToTxt("点击迭代")
            driver.find_element_by_link_text("任务").click()
            logWriteToTxt("点击任务")
            driver.find_element_by_xpath("//*[@id='pageNav']/div[1]/div/button").click()
            logWriteToTxt("点击左上角迭代主页")
            time.sleep(2)
            driver.find_element_by_xpath("//*[@id='pageNav']/div[1]/div/ul/li[2]/a").click()
            logWriteToTxt("点击所有迭代")
            time.sleep(2)
            driver.find_element_by_xpath("//*[@id='closedTab']/span").click()
            logWriteToTxt("点击已关闭tab")
            time.sleep(2)
            if (i > 9):  ##这时候就要翻页了
                count = int(i / 10) + 1  # 第几页
                for j in range(1, count):
                    driver.find_element_by_class_name("icon-angle-right").click()
                    logWriteToTxt("点击下一页")
                    time.sleep(2)
            # driver.find_element_by_xpath("//*[@id='projectTableList']/tr[" + str(i % 10 + 1) + "]/td[2]/a").click()
            from selenium.webdriver.common.keys import Keys
            driver.find_element_by_css_selector(
                "#projectTableList > tr:nth-child(" + str(i % 10 + 1) + ") > td:nth-child(2) > a").send_keys(
                Keys.ENTER)
            print("选择一个任务")
            logWriteToTxt("选择一个任务")
            time.sleep(2)

            # driver.find_elements_by_class_name("dropdown-toggle")[1].click()
            # logWriteToTxt("点击每页几项弹出框")
            # time.sleep(3)
            # driver.find_element_by_xpath("//*[@id='projectsForm']/div/ul/li[2]/div/ul/li[15]").click()
            # logWriteToTxt("选择每页2000项")
            # time.sleep(3)
            # driver.find_element_by_xpath("//*[@id='projectTableList']/tr[" + str(i+1) + "]/td[2]/a").click()
            # logWriteToTxt("选择一个任务")
            # time.sleep(2)
            driver.find_element_by_link_text("任务").click()
            logWriteToTxt("点击任务")
            time.sleep(1)
            driver.find_element_by_id("all").click()
            logWriteToTxt("点击所有")
            time.sleep(5)
            taskelement = driver.find_element_by_xpath("//*[@id='mainMenu']/div[3]/div[1]/button")
            # 悬停在导出-元素上
            ActionChains(driver).move_to_element(taskelement).perform()
            logWriteToTxt("鼠标悬停在导出按钮上")
            time.sleep(4)
            while 1:
                try:
                    driver.find_element_by_link_text("导出数据").click()
                    logWriteToTxt("点击导出数据按钮")
                    time.sleep(4)
                    break
                except:
                    taskelement = driver.find_element_by_xpath("//*[@id='mainMenu']/div[3]/div[1]/button")
                    # 悬停在导出-元素上
                    ActionChains(driver).move_to_element(taskelement).perform()
                    logWriteToTxt("鼠标悬停在导出按钮上")
            driver.switch_to.frame(driver.find_element_by_id("iframe-triggerModal"))
            time.sleep(1)
            while 1:
                try:
                    driver.find_element_by_id("submit").click()
                    logWriteToTxt("点击导出按钮")
                    time.sleep(20)
                    break
                except:
                    logWriteToTxt("还未定位到导出按钮!")
                    time.sleep(1)
            time.sleep(14)
            copy__file(get_str_new_file(downloadPath,".csv"), Folder_Path)
    driver.close()


#得到未关闭任务文件,并且得到产品线
def getTaskFiles(driver,isHalfYear):
    newdriver=NewDriver(driver)
    computerName=os.getlogin()
    downloadPath = "C:\\Users\\" + computerName + "\\Downloads"
    path = os.path.dirname(os.getcwd()) + '/file'
    Folder_Path = new_file(path)  # 得到下载任务文件最新的文件夹
    driver.find_element_by_link_text("迭代").click()
    logWriteToTxt("点击迭代")
    driver.find_element_by_link_text("任务").click()
    logWriteToTxt("点击任务")
    session =_login()
    hrefs = _getiterate(session)
    a = hrefs.split(',')
    print("多少：",len(a)-1)
    m=0
    for i in a:
        m=m+1
        # print("i",i)
        # print("m",m)
        if(m==len(a)):
            # print("拜拜")
            break
        if(m!=1):
            driver.refresh()#页面刷新
        # all_handles = driver.window_handles #到底有几个页面
        # print(m,all_handles)
        currentItemText = driver.find_element_by_id("currentItem").text
        # print(currentItemText)
        ac=newdriver.findElement(By.ID,"currentItem")
        ActionChains(driver).move_to_element(ac).click(ac).perform()
        time.sleep(2)
        logWriteToTxt("点击当前迭代:" + currentItemText)
        while 1:
            try:
                # ac = newdriver.findElement(By.XPATH, "//div[@id='dropMenu']//div[@class='list-group']//a[@title='"+i+"']")
                ac = newdriver.findElement(By.XPATH,
                                           "//*[@id='dropMenu']/div[2]/div/div[1]/div[1]/a["+str(m)+"]")
                ActionChains(driver).move_to_element(ac).click(ac).perform()
                time.sleep(2)
                currentItemText = newdriver.findElement(By.ID, "currentItem").text
                logWriteToTxt('已定位到元素，并且点击该迭代:' + currentItemText)
                newdriver.findElement(By.XPATH, "//div[@id='mainMenu']//a[@id='all']").click()
                logWriteToTxt("点击所有")
                time.sleep(2)
                taskelement = newdriver.findElement(By.XPATH, "//div[@id='mainMenu']//button/span[text()='导出']/..")
                # 悬停在导出-元素上
                ActionChains(driver).move_to_element(taskelement).perform()
                logWriteToTxt("鼠标悬停在导出按钮上")
                newdriver.findElement(By.XPATH, "//ul[@id='exportActionMenu']//a[text()='导出数据']").click()
                logWriteToTxt("点击导出数据按钮")
                time.sleep(5)
                driver.switch_to.frame(newdriver.findElement(By.ID, "iframe-triggerModal"))
                newdriver.findElement(By.XPATH, "//div[@id='mainContent']//button[@id='submit']").click()
                logWriteToTxt("点击导出按钮")
                time.sleep(20)
                break
            except Exception as e:
                logWriteToTxt(e)
                logWriteToTxt("点击当前迭代:" + currentItemText + "执行失败")

        # while 1:
        #     start = time.clock()
        #     try:
        #         driver.find_element_by_xpath("//*[@id='dropMenu']/div[2]/div/div[1]/div[1]/a["+str(m)+"]").click()
        #         currentItemText = driver.find_element_by_id("currentItem").text
        #         logWriteToTxt('已定位到元素，并且点击该迭代:'+currentItemText)
        #         print(m, currentItemText)
        #         end = time.clock()
        #         # print(m, currentItemText)
        #         time.sleep(3)
        #         driver.find_element_by_id("all").click()
        #         logWriteToTxt("点击所有" )
        #         time.sleep(5)
        #         taskelement = driver.find_element_by_xpath("//*[@id='mainMenu']/div[3]/div[1]/button")
        #         # 悬停在导出-元素上
        #         ActionChains(driver).move_to_element(taskelement).perform()
        #         logWriteToTxt("鼠标悬停在导出按钮上")
        #         time.sleep(4)
        #         while 1:
        #             try:
        #                 driver.find_element_by_link_text("导出数据").click()
        #                 logWriteToTxt("点击导出数据按钮")
        #                 time.sleep(4)
        #                 break
        #             except:
        #                 taskelement = driver.find_element_by_xpath("//*[@id='mainMenu']/div[3]/div[1]/button")
        #                 # 悬停在导出-元素上
        #                 ActionChains(driver).move_to_element(taskelement).perform()
        #                 logWriteToTxt("鼠标悬停在导出按钮上")
        #         driver.switch_to.frame(driver.find_element_by_id("iframe-triggerModal"))
        #         time.sleep(1)
        #         while 1:
        #             try:
        #                 driver.find_element_by_id("submit").click()
        #                 logWriteToTxt("点击导出按钮")
        #                 time.sleep(20)
        #                 break
        #             except:
        #                 logWriteToTxt("还未定位到导出按钮!")
        #                 time.sleep(1)
        #         time.sleep(10)
        #         break
        #     except:
        #         logWriteToTxt("还未定位到元素!")
        copy__file(get_str_new_file(downloadPath,".csv"), Folder_Path)
        #         # print('定位耗费时间：' + str(end - start))
    time.sleep(10)
    if(int(isHalfYear)==0):
        driver.close()
    return session



#得到未关闭需求文件,并且得到产品线
def getRequestFiles(driver,session):
    computerName=os.getlogin()
    downloadPath = "C:\\Users\\" + computerName + "\\Downloads"
    path = os.path.dirname(os.getcwd()) + '/file'
    Folder_Path = new_file(path)  # 得到下载任务文件最新的文件夹
    driver.find_element_by_link_text("迭代").click()
    logWriteToTxt("点击迭代")
    driver.find_element_by_link_text("需求").click()
    logWriteToTxt("点击需求")
    hrefs = _getiterate(session)
    a = hrefs.split(',')
    print("多少：",len(a)-1)
    m=0
    for i in a:
        m=m+1
        # print("i",i)
        # print("m",m)
        if(m==len(a)):
            # print("拜拜")
            break
        if(m!=1):
            driver.refresh()#页面刷新
        # all_handles = driver.window_handles #到底有几个页面
        # print(m,all_handles)
        currentItemText = driver.find_element_by_id("currentItem").text
        # print(currentItemText)
        driver.find_element_by_id("currentItem").click()
        logWriteToTxt("点击当前迭代:" + currentItemText)
        while 1:
            start = time.clock()
            try:
                driver.find_element_by_xpath("//*[@id='dropMenu']/div[2]/div/div[1]/div[1]/a["+str(m)+"]").click()
                currentItemText = driver.find_element_by_id("currentItem").text
                logWriteToTxt('已定位到元素，并且点击该迭代:'+currentItemText)
                print(m, currentItemText)
                end = time.clock()
                # print(m, currentItemText)
                time.sleep(3)
                # // *[ @ id = "mainMenu"] / div[1] / a[1]
                driver.find_element_by_xpath("//*[@id='mainMenu']/div[1]/a[1]").click()
                logWriteToTxt("点击所有需求" )
                time.sleep(2)
                driver.find_element_by_xpath("//*[@id='mainMenu']/div[2]/a").click()
                logWriteToTxt("点击导出数据按钮")
                time.sleep(4)
                driver.switch_to.frame(driver.find_element_by_id("iframe-triggerModal"))
                time.sleep(1)
                while 1:
                    try:
                        driver.find_element_by_id("submit").click()
                        logWriteToTxt("点击导出按钮")
                        time.sleep(20)
                        break
                    except:
                        logWriteToTxt("还未定位到导出按钮!")
                        time.sleep(1)
                time.sleep(10)
                break
            except:
                logWriteToTxt("还未定位到元素!")
        copy__file(get_str_new_file(downloadPath,".csv"), Folder_Path)
        #         # print('定位耗费时间：' + str(end - start))
    time.sleep(10)
    driver.close()




#合并文件--任务和bug
def  merge_csv(title,session,isHalfYear,push):

    if ((title == "tasks")&(int(isHalfYear)==0)):
        resultpath = os.path.dirname(os.getcwd()) + "/result/" + datetime.datetime.now().strftime('%Y%m%d %H%M%S')
        try:
            os.makedirs(resultpath)
        except:
            pass
    else:
        path = os.path.dirname(os.getcwd()) + '/result'
        resultpath = new_file(path)  # 得到结果文件最新的文件夹
    # filepath=resultpath+"/"+datetime.datetime.now().strftime('%Y%m%d%H%M%S') + ".csv"
    filepath = resultpath + "/"+title+".csv"

    #判断一下文件夹下的文件是否存在，存在就删除
    if(os.path.exists(filepath)==True):
        os.remove(filepath)

    path = os.path.dirname(os.getcwd()) + '/file'
    Folder_Path = new_file(path)  # 得到下载文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    # Folder_Path="E:\\PythonWorkspace\\ZentaoTest\\file\\20190412 155807"
    file_list = getfileName(Folder_Path)
    m=-1
    for i in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[i], encoding='UTF-8')
        df = pd.read_csv(f)  # 编码默认UTF-8，若乱码自行更改
        if(push==True &(len(df) != 0)):
            m = i
            break
        elif (("ZEDNE SprintTodo-所有任务" not in file_list[i])&(len(df)!=0)&push==False):
            m=i
            break

    # 读取第一个不为空的CSV文件并包含表头
    print("第一个文件:",Folder_Path + '\\' + file_list[m])
    f = open(Folder_Path + '\\' + file_list[m], encoding='UTF-8')
    df = pd.read_csv(f)  # 编码默认UTF-8，若乱码自行更改
    # data = pd.read_csv(Folder_Path + '\\' + file_list[0], nrows=1)  # 取出第一行
    # print(data)
    # 将读取的第一个CSV文件写入合并后的文件保存
    df.to_csv(filepath, encoding="utf_8_sig", mode='a+', index_label=False)

    # 循环遍历列表中各个CSV文件名，并追加到合并后的文件
    for i in range(m+1, len(file_list)):
        if (push == False):#不是微信推送
            if ("ZEDNE SprintTodo-所有任务" not in file_list[i]):
                print("后面文件:", Folder_Path + '\\' + file_list[i])
                f = open(Folder_Path + '\\' + file_list[i], encoding='UTF-8')
                df = pd.read_csv(f)
                if((len(df) != 0)):
                    df.to_csv(filepath, encoding="utf_8_sig", header=False, mode='a+', index_label=False)

        else:#微信推送可以要ZEDNE SprintTodo-所有任务
            print("后面文件:", Folder_Path + '\\' + file_list[i])
            f = open(Folder_Path + '\\' + file_list[i], encoding='UTF-8')
            df = pd.read_csv(f)
            if ((len(df) != 0)):
                df.to_csv(filepath, encoding="utf_8_sig", header=False, mode='a+', index_label=False)
    #把csv文件转换为xls文件
    # xlsxpath=resultpath+"/" + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + "tasks.xls"
    xlsxpath = resultpath + "/" + title + ".xls"
    if (os.path.exists(xlsxpath) == True):
        os.remove(xlsxpath)
    csv_to_xlsx(filepath, xlsxpath)
    if(title=="tasks"): #因为有任务是总任务的，所以要把责任人分开出来
        # 得到部门下的名称和人名字
        departNamelist = []
        departNamelist = _getDepartment(session)
        departs = departNamelist[0]  # 部门
        names = departNamelist[1]  # 人

        rb = xlrd.open_workbook(xlsxpath)  # 打开tasks.xls文件，修改其的预计开始时间
        ro = rb.sheets()[0]  # 读取表单
        wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
        ws = wb.get_sheet(0)  # 获取表单

        bk = xlrd.open_workbook(xlsxpath)  # 打开该文件，复制到all.xls文件
        try:
            sh = bk.sheet_by_name("data")
        except:
            print("代码出错")
        nrows = sh.nrows  # 获取行数
        ncols = sh.ncols  # 获取列数
        idcol=0  #编号列
        endcol = 11  # 截止时间列
        statecol=12       #任务状态列
        departcol = 33  # 责任部门列
        namecol = 34  # 责任人列
        firstexceptcol = 13  # 最初预计列
        alllosecol = 14  # 总消耗列
        leftcol=15  #剩余列
        appointcol=22 #由谁完成列
        accappointcol=23 #实际完成时间列
        finishTimecol=35#完成及时性

        mi=-1
        for i in range(0, nrows):
            if(i==0):
                mi=mi+1
                for j in range(0, ncols):
                    ws.write(mi, j, sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
            else:
                idcolresult=sh.cell_value(i,idcol)
                endcolresult=sh.cell_value(i,endcol)
                statecolresult=sh.cell_value(i,statecol)
                firstexceptcolresult = sh.cell_value(i, firstexceptcol)
                alllosecolresult = sh.cell_value(i, alllosecol)
                leftcolresult = sh.cell_value(i, leftcol)
                if (len(firstexceptcolresult.split('\n')) >1):
                    # print("多任务任务任务：",idcolresult)
                    for sp in range(0, len(firstexceptcolresult.split('\n')) - 1):
                        mi = mi + 1
                        departcolresult="" #责任部门
                        namecolresult=""#责任人
                        # print("222:",firstexceptcolresult,alllosecolresult,leftcolresult)
                        namecolresult = firstexceptcolresult.split('\n')[sp].split(':')[0]
                        ws.write(mi, firstexceptcol, firstexceptcolresult.split('\n')[sp].split(':')[1])  # 向第i行第j列写入获取到的值
                        ws.write(mi, alllosecol, alllosecolresult.split('\n')[sp].split(':')[1])  # 向第i行第j列写入获取到的值
                        ws.write(mi, leftcol, leftcolresult.split('\n')[sp].split(':')[1])  # 向第i行第j列写入获取到的值
                        ind = -1
                        for na in names:
                            ind = ind + 1
                            for name in na:
                                if (namecolresult == name):
                                    departcolresult = departs[ind]
                        ws.write(mi, namecol, namecolresult)  # 向第i行第j列写入获取到的值
                        ws.write(mi, departcol, departcolresult)  # 向第i行第j列写入获取到的值
                        if (statecolresult == "进行中"):  # 当在多人任务中，任务状态为进行中的时候，就要判断是否有人已经完成任务了
                            NameTimes=getManyResponseTask(session,idcolresult)
                            Names=NameTimes[0]
                            Times=NameTimes[1]
                            flag=False
                            for jj in range(0,len(Names)):
                                if(namecolresult==Names[jj]):
                                    flag=True
                                    break
                            if(flag==True):
                                ws.write(mi, appointcol, namecolresult)  # 由谁完成列
                                timeStr = Times[jj].split(' ')[0]
                                timeStr2 = str(int(timeStr.split('-')[0])) + "/" + str(
                                    int(timeStr.split('-')[1])) + "/" + str(int(timeStr.split('-')[2]))
                                ws.write(mi, accappointcol, timeStr2)  # 实际完成时间列
                                ws.write(mi, statecol, "已完成")  # 任务状态列

                                if (("0000" not in endcolresult) & ("0000" not in timeStr2)):  # 截止时间和完成时间不为空
                                    strftime = datetime.datetime.strptime(endcolresult, "%Y/%m/%d")
                                    strftime2 = datetime.datetime.strptime(timeStr2, "%Y/%m/%d")
                                    if (strftime >= strftime2):
                                        # print("已完成")
                                        # df.loc[a[i], '责任人'] = "按时完成"  # 根据索引来改变名字
                                        ws.write(mi, finishTimecol, "按时完成")  # 及时完成性
                                    else:
                                        # print("已完成")
                                        # df.loc[a[i], '责任人'] = "延期完成"  # 根据索引来改变名字
                                        ws.write(mi, finishTimecol, "延期完成")  # 及时完成性
                                for j in range(0, ncols):
                                    if ((j != departcol) & (j != namecol) & (j != leftcol) & (j != alllosecol) & (
                                            j != firstexceptcol) & (j != statecol) & (j != appointcol) & (
                                            j != accappointcol) & (j != finishTimecol)):
                                        ws.write(mi, j, sh.cell_value(i, j))  # 向第i行第j列写入获取到的值

                            else:
                                for j in range(0, ncols):
                                    if ((j != departcol) & (j != namecol) & (j != leftcol) & (j != alllosecol) & (
                                            j != firstexceptcol)):
                                        ws.write(mi, j, sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                        else:
                            for j in range(0, ncols):
                                if ((j != departcol) & (j != namecol) & (j != leftcol) & (j != alllosecol) & (
                                        j != firstexceptcol)):
                                    ws.write(mi, j, sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                else:
                    mi = mi + 1
                    for j in range(0, ncols):
                        ws.write(mi, j, sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        wb.save(xlsxpath)
    return xlsxpath




#把任务和bug和用例版本都结合在一个excel里的不同sheet里，时间格式弄成excel里的时间格式
def  merge_Excel(isHalfYear):
    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    if(int(isHalfYear)==0):
        xlsxpath2 = Folder_Path + "/一月tasks.xls"
        xlsxpath3 = Folder_Path + "/一月解决bugs.xls"
        xlsxpath4 = Folder_Path + "/一月回归bugs.xls"
        xlsxpath5 = Folder_Path + "/一月创建bugs.xls"
        xlsxpath6 = Folder_Path + "/一月关闭bugs.xls"
    else:
        xlsxpath2 = Folder_Path + "/半年tasks.xls"
        xlsxpath3 = Folder_Path + "/半年解决bugs.xls"
        xlsxpath4 = Folder_Path + "/半年回归bugs.xls"
        xlsxpath5 = Folder_Path + "/半年创建bugs.xls"
        xlsxpath6 = Folder_Path + "/半年关闭bugs.xls"

    #
    xlsxpath1 = Folder_Path + "/一周tasks.xls"
    xlsxpath7 = Folder_Path + "/版本测试单.xls"
    xlsxpath8 = Folder_Path + "/用例测试单.xls"

    paths = []
    if (os.path.exists(xlsxpath1) == True):
        paths.append(xlsxpath1)
    if (os.path.exists(xlsxpath2) == True):
        paths.append(xlsxpath2)
    if (os.path.exists(xlsxpath3) == True):
        paths.append(xlsxpath3)
    if (os.path.exists(xlsxpath4) == True):
        paths.append(xlsxpath4)
    if (os.path.exists(xlsxpath5) == True):
        paths.append(xlsxpath5)
    if (os.path.exists(xlsxpath6) == True):
        paths.append(xlsxpath6)
    if (os.path.exists(xlsxpath7) == True):
        paths.append(xlsxpath7)
    if (os.path.exists(xlsxpath8) == True):
        paths.append(xlsxpath8)

    #先创建一个excel文件，并且添加相对应的sheet
    if(int(isHalfYear)==0):
        allpath = Folder_Path + "/all.xls"
    else:
        allpath = Folder_Path + "/all2.xls"  #半年统计的数据

    if (os.path.exists(allpath) == True):
        os.remove(allpath)

    work_book = xlwt.Workbook(encoding='utf-8')
    work_book.add_sheet('一周tasks')#第一个sheet名称
    work_book.save(allpath)

    for m in range(1,len(paths)):#添加sheet-名称
        # print("m:",m)
        rb = xlrd.open_workbook(allpath, formatting_info=True)
        wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
        if(m==1):
            if(int(isHalfYear)==0):
                wb.add_sheet('一月tasks')
            else:
                wb.add_sheet('半年tasks')
            wb.save(allpath)
        elif(m==2):
            if (int(isHalfYear) == 0):
                wb.add_sheet('一月解决bugs')
            else:
                wb.add_sheet('半年解决bugs')

            wb.save(allpath)
        elif (m == 3):
            if (int(isHalfYear) == 0):
                wb.add_sheet('一月回归bugs')
            else:
                wb.add_sheet('半年回归bugs')
            wb.save(allpath)
        elif (m == 4):
            if (int(isHalfYear) == 0):
                wb.add_sheet('一月创建bugs')
            else:
                wb.add_sheet('半年创建bugs')
            wb.save(allpath)
        elif (m == 5):
            if (int(isHalfYear) == 0):
                wb.add_sheet('一月关闭bugs')
            else:
                wb.add_sheet('半年关闭bugs')
            wb.save(allpath)
        elif (m == 6):
            if (int(isHalfYear) == 0):
                wb.add_sheet('一月版本测试单')
            else:
                wb.add_sheet('半年版本测试单')
            wb.save(allpath)
        elif (m == 7):
            if (int(isHalfYear) == 0):
                wb.add_sheet('一月用例测试单')
            else:
                wb.add_sheet('半年用例测试单')
            wb.save(allpath)

    rb = xlrd.open_workbook(allpath, formatting_info=True)
    print("sheets.size:", len(rb.sheets()))  # sheet数量
    for sheet in rb.sheets():
        print(sheet.name)  # sheet名称
    sheetlen = len(rb.sheets())

    today = time.strftime("%Y-%m-%d")  # 今天
    day_num = int(today.split('-')[2])  # 几号
    month_num = int(today.split('-')[1])  # 几月
    year_num = int(today.split('-')[0])  # 几年
    tt = time.strftime(str(year_num) + "-" + str(1) + "-" + str(1))
    flag = is_leap_year(year_num)
    if (flag == True):  # 该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    if(int(isHalfYear)==0):
        # if (month_num == 1):
        #     if (int(day_num) < 27):#那就上一年的27号到前30天
        #         start = time.strftime(str(year_num-1) + "-" + str(11) + "-" + str(days[10]-3))  # 开始日期
        #         end = time.strftime(str(year_num-1) + "-" + str(12) + "-" + str(27))  # 结束日期
        #     else:#那就该月的27号到前30天
        #         start = time.strftime(str(year_num - 1) + "-" + str(12) + "-" + str(days[11] - 3))  # 开始日期
        #         end = time.strftime(str(year_num) + "-" + str(1) + "-" + str(27))  # 结束日期
        # elif (month_num == 2):
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         start = time.strftime(str(year_num-1) + "-" + str(12) + "-" + str(days[11] - 3))  # 开始日期
        #         end = time.strftime(str(year_num) + "-" + str(1) + "-" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         start = time.strftime(str(year_num) + "-" + str(month_num-1) + "-" + str(days[month_num-2] - 3))  # 开始日期
        #         end = time.strftime(str(year_num) + "-" + str(month_num) + "-" + str(27))  # 结束日期
        # else:
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         start = time.strftime(str(year_num) + "-" + str(month_num-2) + "-" + str(days[month_num-3] - 3))  # 开始日期
        #         end = time.strftime(str(year_num) + "-" + str(month_num-1) + "-" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         start = time.strftime(str(year_num) + "-" + str(month_num-1) + "-" + str(days[month_num-2] - 3))  # 开始日期
        #         end = time.strftime(str(year_num) + "-" + str(month_num) + "-" + str(27))  # 结束日期
        if (month_num == 1):  # 就要计算到上一年的12月
            start = time.strftime(str(year_num - 1) + "-" + str(12) + "-" + str(1))  # 截止开始日期
            end = time.strftime(str(year_num - 1) + "-" + str(12) + "-" + str(days[11]))  # 截止结束日期
        else:  # 就要计算到上一月
            start = time.strftime(str(year_num) + "-" + str(month_num - 1) + "-" + str(1))  # 截止开始日期
            end = time.strftime(str(year_num) + "-" + str(month_num - 1) + "-" + str(days[month_num - 2]))  # 截止结束日期

    elif (int(isHalfYear) == 1):
        start = time.strftime(str(year_num) + "-" + str(1) + "-" + str(1))  # 开始日期
        end = time.strftime(str(year_num) + "-" + str(6) + "-" + str(days[5]))  # 结束日期
    elif (int(isHalfYear) == 2):
        start = time.strftime(str(year_num) + "-" + str(7) + "-" + str(1))  # 开始日期
        end = time.strftime(str(year_num) + "-" + str(12) + "-" + str(days[11]))  # 结束日期

    for m in range(0,sheetlen):#把其他excel的值放入到all.xsl
        # print("m:",m)
        if((m!=6)&(m!=7)):
            rb = xlrd.open_workbook(allpath, formatting_info=True)  # 打开all.xls文件，把其他文件的值复制到该文件
            ro = rb.sheets()[m]  # 读取表单m
            wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
            ws = wb.get_sheet(m)  # 获取表单m

            bk = xlrd.open_workbook(paths[m])  # 打开该文件，复制到all.xls文件
            try:
                sh = bk.sheet_by_name("data")
            except:
                print("代码出错")
            nrows = sh.nrows  # 获取行数
            ncols = sh.ncols  # 获取列数
            # print("行，列:",nrows,ncols)
            dateFormat = xlwt.XFStyle()
            dateFormat.num_format_str = 'yyyy/mm/dd'
            for i in range(0, nrows):
                # print("-----正在写入 " + str(i) + " 行")
                for j in range(0, ncols):
                    shvalue = sh.cell_value(i, j)
                    if ((len(shvalue.split('/')) == 3) & (len(shvalue) < 12)):  # 把str转换为excel格式里的时间格式
                        # print((shvalue.split('/')[0]).isdigit(),(shvalue.split('/')[1]).isdigit())
                        if ((sh.cell_value(i, j).split('/')[2]).isdigit() == True & (
                        sh.cell_value(i, j).split('/')[1]).isdigit() == True & (
                        sh.cell_value(i, j).split('/')[0]).isdigit() == True):
                            year = int(sh.cell_value(i, j).split('/')[0])
                            month = int(sh.cell_value(i, j).split('/')[1])
                            day = int(sh.cell_value(i, j).split('/')[2])
                            # print(month, year, day)
                            ws.write(i, j, dt.date(year, month, day), dateFormat)
                    else:
                        ws.write(i, j, sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        else:
            rb = xlrd.open_workbook(allpath, formatting_info=True)  # 打开 上月版本测试完成情况.xls文件，把其他文件的值复制到该文件
            ro = rb.sheets()[m]  # 读取表单m
            wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
            ws = wb.get_sheet(m)  # 获取表单m
            bk = xlrd.open_workbook(paths[m])  # 打开该文件，复制到 上月版本测试完成情况.xls文件
            try:
                sh = bk.sheet_by_name("data")
            except:
                print("代码出错")
            nrows = sh.nrows  # 获取行数
            ncols = sh.ncols  # 获取列数
            # print("行，列:",nrows,ncols)
            dateFormat = xlwt.XFStyle()
            dateFormat.num_format_str = 'yyyy-mm-dd'
            mi = -1  # 新的excel的第几行
            for i in range(0, nrows):
                if (i == 0):
                    # print("-----正在写入 " + str(i) + " 行")
                    mi = mi + 1
                    for j in range(0, ncols):
                        ws.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                else:
                    starttime = sh.cell_value(i, 7)
                    endtime = sh.cell_value(i, 8)
                    state = sh.cell_value(i, 9)
                    strftime = datetime.datetime.strptime(starttime, "%Y-%m-%d")  # 版本开始时间
                    strftime2 = datetime.datetime.strptime(endtime, "%Y-%m-%d")  # 版本结束时间
                    strftime3 = datetime.datetime.strptime(start, "%Y-%m-%d")
                    strftime4 = datetime.datetime.strptime(end, "%Y-%m-%d")
                    strftime5 = datetime.datetime.strptime(tt, "%Y-%m-%d")
                    if (state == "已完成"):
                        if ((strftime2 <= strftime4) & (strftime2 >= strftime3)):
                            mi = mi + 1
                            for j in range(0, ncols):
                                if ((j == 7) | (j == 8)):
                                    # if ((len(shvalue.split('-')) == 3) & (len(shvalue) < 12)):  # 把str转换为excel格式里的时间格式
                                    #     # print((shvalue.split('/')[0]).isdigit(),(shvalue.split('/')[1]).isdigit())
                                    #     if ((sh.cell_value(i, j).split('-')[2]).isdigit() == True & (
                                    #             sh.cell_value(i, j).split('-')[1]).isdigit() == True & (
                                    #             sh.cell_value(i, j).split('-')[0]).isdigit() == True):
                                    year = int(sh.cell_value(i, j).split('-')[0])
                                    month = int(sh.cell_value(i, j).split('-')[1])
                                    day = int(sh.cell_value(i, j).split('-')[2])
                                    # print(month, year, day)
                                    ws.write(mi, j, dt.date(year, month, day), dateFormat)
                                else:
                                    ws.write(mi, j, sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                    else:
                        if ((strftime2 <= strftime4) & (strftime5 <= strftime2)):
                            mi = mi + 1
                            for j in range(0, ncols):
                                if ((j == 7) | (j == 8)):
                                    # if ((len(shvalue.split('-')) == 3) & (len(shvalue) < 12)):  # 把str转换为excel格式里的时间格式
                                    #     # print((shvalue.split('/')[0]).isdigit(),(shvalue.split('/')[1]).isdigit())
                                    #     if ((sh.cell_value(i, j).split('-')[2]).isdigit() == True & (
                                    #             sh.cell_value(i, j).split('-')[1]).isdigit() == True & (
                                    #             sh.cell_value(i, j).split('-')[0]).isdigit() == True):
                                    year = int(sh.cell_value(i, j).split('-')[0])
                                    month = int(sh.cell_value(i, j).split('-')[1])
                                    day = int(sh.cell_value(i, j).split('-')[2])
                                    # print(month, year, day)
                                    ws.write(mi, j, dt.date(year, month, day), dateFormat)
                                else:
                                    ws.write(mi, j, sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        wb.save(allpath)
        change_task_excel(allpath)





#修改excel文件
#已完成和已关闭状态的任务，状态单元格设置为“绿填充色深绿色文本”
#过期未完成任务，状态单元格设置为“浅红填充色深红色文本”
#并且添加冻结窗口和筛选
def  change_task_excel(excelpath):
    styleRedBkg = xlwt.easyxf('pattern: pattern solid, fore_colour red;')  # 红色
    styleGreenBkg = xlwt.easyxf('pattern: pattern solid, fore_colour green;')  # 绿色
    #  ,'font: colour_index red;'
    rb = xlrd.open_workbook(excelpath)  # 打开xls文件
    # print("sheets.size:",len(rb.sheets()))#sheet数量
    # for sheet in rb.sheets():
    #     print(sheet.name) # sheet名称
    sheetlen=len(rb.sheets())
    for i in range(0,sheetlen):
        # print("i:",i)
        if((i==0)|(i==1)):
            rb = xlrd.open_workbook(excelpath,formatting_info=True)  # 打开xls文件
            ro = rb.sheets()[i]  # 读取表单0
            wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
            ws = wb.get_sheet(i)  # 获取表单0
            col = 0  # 指定修改的列
            statecol = 12  # 任务状态
            startcol = 10  # 实际开始
            endcol = 11  # 截止日期
            exceptcol = 13  # 最初预计多少个工时
            today = time.strftime("%Y/%m/%d")
            # print("today:",today)
            # print(ro.nrows)
            # week=datetime.datetime.strptime("20190407","%Y/%m/%d").weekday()
            # week2=datetime.datetime.now().weekday()
            # print(week,week2)
            strftime = datetime.datetime.strptime(today, "%Y/%m/%d")
            # print("2017-11-02大于2017-01-04：", strftime > strftime2)
            # print("行:",ro.nrows)
            # print("列：",ro.ncols)#取总列数
            for i in range(1, ro.nrows):  # 循环所有的行
                stateresult = ro.cell(i, statecol).value
                endresult = ro.cell(i, endcol).value
                if ((stateresult == "已完成") | (stateresult == "已关闭")):  # 判断状态是否等于已完成，已关闭
                    ws.write(i, statecol, ro.cell(i, statecol).value, styleGreenBkg)
                elif ((stateresult == "进行中") | (stateresult == "未开始")):  # 判断状态是否等于进行中或未开始，然后截止日期判断是否大于现在的日期
                    # print(i)
                    # endstr = int(endresult.split('-')[0] + endresult.split('-')[1] + endresult.split('-')[2])
                    # print(endstr)
                    # print(1,endresult)
                    if ("0000" not in str(endresult)):
                        strftime = datetime.datetime.strptime(today, "%Y/%m/%d")
                        if (len(str(endresult).split('/')) == 3):
                            strftime2 = datetime.datetime.strptime((endresult), "%Y/%m/%d")
                        else:
                            # 数字转化为元祖--》(2019, 3, 19, 0, 0, 0)
                            arra = xlrd.xldate_as_tuple(endresult, 0)  # 转化为元组形式
                            # print("arra:",arra)
                            otherStyleTime = str(arra[0]) + "/" + str(arra[1]) + "/" + str(arra[2])
                            strftime2 = datetime.datetime.strptime((otherStyleTime), "%Y/%m/%d")
                            # timeArray = time.localtime(endresult)
                            # otherStyleTime = time.strftime("%Y/%m/%d", timeArray)
                            # strftime2=datetime.datetime.strptime((otherStyleTime), "%Y/%m/%d")
                        # print(2,strftime,strftime2)
                        if (strftime > strftime2):
                            ws.write(i, statecol, ro.cell(i, statecol).value, styleRedBkg)
                    # else:
                    #     ws.write(i, statecol, ro.cell(i, statecol).value, styleRedBkg)
            wb.save(excelpath)

'''
    #修改所有时间格式-修改时间都为2019-9-21（类似这个格式的）
    for i in range(0,sheetlen):
        # print("i:",i)
        rb = xlrd.open_workbook(excelpath,formatting_info=True)  # 打开xls文件
        ro = rb.sheets()[i]  # 读取表单0
        wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
        ws = wb.get_sheet(i)  # 获取表单0
        col = 0  # 指定修改的列
        exceptstartcol=8# 预计开始
        startcol = 9  # 实际开始
        endcol = 10  # 截止日期
        buildcol=18  #创建日期
        appointcol=20 #指派日期
        finishcol=22 #完成时间
        canclecol=24 #取消时间
        closecol=26 #关闭时间
        for i in range(1, ro.nrows):  # 循环所有的行
            exceptstartresult = ro.cell(i, exceptstartcol).value
            startresult = ro.cell(i, startcol).value
            endresult = ro.cell(i, endcol).value
            buildresult = ro.cell(i, buildcol).value
            appointresult = ro.cell(i, appointcol).value
            finishresult = ro.cell(i, finishcol).value
            cancleresult = ro.cell(i, canclecol).value
            closeresult = ro.cell(i, closecol).value
            exceptstartresult0=exceptstartresult.split('-')[0]
            startresult0=startresult.split('-')[0]
            endresult0=endresult.split('-')[0]
            buildresult0=buildresult.split('-')[0]
            appointresult0=appointresult.split('-')[0]
            finishresult0=finishresult.split('-')[0]
            cancleresult0=cancleresult.split('-')[0]
            closeresult0=closeresult.split('-')[0]
            if(int(exceptstartresult0)!=0):
                strvalue=exceptstartresult.split('-')[0]+"-"+str(int(exceptstartresult.split('-')[1]))+""+exceptstartresult.split('-')[2]
                ws.write(i, exceptstartcol, strvalue)
            if (int(startresult0) != 0):
                strvalue = startresult.split('-')[0] + "-" + str(int(startresult.split('-')[1])) + "" + \
                           startresult.split('-')[2]
                ws.write(i, startcol, strvalue)
            if (int(endresult0) != 0):
                strvalue = endresult.split('-')[0] + "-" + str(int(endresult.split('-')[1])) + "" + \
                           endresult.split('-')[2]
                ws.write(i, endcol, strvalue)
            if (int(buildresult0) != 0):
                strvalue = buildresult.split('-')[0] + "-" + str(int(buildresult.split('-')[1])) + "" + \
                           buildresult.split('-')[2]
                ws.write(i, buildcol, strvalue)
            if (int(appointresult0) != 0):
                strvalue = appointresult.split('-')[0] + "-" + str(int(appointresult.split('-')[1])) + "" + \
                           appointresult.split('-')[2]
                ws.write(i, appointcol, strvalue)
            if (int(finishresult0) != 0):
                strvalue = finishresult.split('-')[0] + "-" + str(int(finishresult.split('-')[1])) + "" + \
                           finishresult.split('-')[2]
                ws.write(i, finishcol, strvalue)
            if (int(cancleresult0) != 0):
                strvalue = cancleresult.split('-')[0] + "-" + str(int(cancleresult.split('-')[1])) + "" + \
                           cancleresult.split('-')[2]
                ws.write(i, canclecol, strvalue)
            if (int(closeresult0) != 0):
                strvalue = closeresult.split('-')[0] + "-" + str(int(closeresult.split('-')[1])) + "" + \
                           closeresult.split('-')[2]
                ws.write(i, closecol, strvalue)
'''
'''
    #再一次打开文件,修改冻结窗格为第一行
    wb = load_workbook(excelpath,data_only=True)
    sheet = wb.active
    # sheet.freeze_panes = 'B1'                       # 冻结列A
    # sheet.freeze_panes = 'C1'                       # 冻结列A和B
    # sheet.freeze_panes = 'C2'                       # 冻结行1和列A和列B
    # sheet.freeze_panes = 'A1'                       # 无冻结
    # sheet.freeze_panes = None                       # 无冻结
    sheet.freeze_panes = 'A2'                                     # 设置第一行为冻结
    wb.save(excelpath)
 '''

#修改bug的csv，
def change_bug_csv(session):

    #得到迭代版本下的时间与bug的id
    dateids=_getIteraBugId(session)

    # 得到产品线-产品
    productlines = _getProduct(session)
    products=productlines[1] #产品
    prolines=productlines[0]#产品线

    # 得到部门下的名称和人名字
    departNamelist = []
    departNamelist = _getDepartment(session)
    departs=departNamelist[0]  #部门
    names=departNamelist[1]  #人

    path = os.path.dirname(os.getcwd()) + '/file'
    Folder_Path = new_file(path)  # 得到下载任务文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    # Folder_Path="E:\\PythonWorkspace\\ZentaoTest\\file\\20190412 155807"
    file_list = getfileName(Folder_Path)
    # 1.修改文件第一步：添加数列
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k], encoding='UTF-8')
        print(1, Folder_Path + '\\' + file_list[k])
        df = pd.read_csv(f, header=0, index_col=0)
        if (len(df) != 0):
            df["产品线"] = ""  # 修改文件
            df["责任部门"] = ""  # 修改文件
            df["责任人"] = ""  # 修改文件
            df["要求解决时间"] = ""  # 修改文件
            df["解决及时性"] = ""  # 修改文件
            df["回归截止时间"] = ""  # 修改文件
            df["回归测试结果"] = ""  # 修改文件
            df["回归测试时间"] = ""  # 修改文件
            df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)


    # 有的时间格式不一致，要修改成一样的格式时间，如2019/9/21
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k], encoding='UTF-8')
        print(2, Folder_Path + '\\' + file_list[k])
        df = pd.read_csv(f, header=0, index_col=0)
        if (len(df) != 0):
            a = df.index.tolist()  # 行索引
            col = 0  # 指定修改的列
            endcol = 14  # 截止日期
            buildcol = 19  # 创建日期
            appointcol = 22  # 指派日期
            solvecol = 26  # 解决时间
            closecol = 28  # 关闭时间
            lastupcol = 33  # 最后修改时间
            for i in range(len(df)):
                endresult = df.iloc[i, endcol]
                buildresult = df.iloc[i, buildcol]
                appointresult = df.iloc[i, appointcol]
                solveresult = df.iloc[i, solvecol]
                closeresult = df.iloc[i, closecol]
                lastupresult = df.iloc[i, lastupcol]
                if (("0000" not in endresult) & ("-" in endresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = endresult.split('-')[0] + "/" + str(
                        int(endresult.split('-')[1])) + "/" + endresult.split('-')[2]
                    # print("3", i, strresult,a[i])
                    df.loc[a[i], 'Bug状态'] = strresult  # 根据索引来改变名字
                if (("0000" not in buildresult) & ("-" in buildresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = buildresult.split('-')[0] + "/" + str(
                        int(buildresult.split('-')[1])) + "/" + buildresult.split('-')[2]
                    # print("4", i, strresult,a[i])
                    df.loc[a[i], '由谁创建'] = strresult  # 根据索引来改变名字
                if (("0000" not in appointresult) & ("-" in appointresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = appointresult.split('-')[0] + "/" + str(
                        int(appointresult.split('-')[1])) + "/" + appointresult.split('-')[2]
                    # print("5", i, strresult,a[i])
                    df.loc[a[i], '指派给'] = strresult  # 根据索引来改变名字
                if (("0000" not in solveresult) & ("-" in solveresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = solveresult.split('-')[0] + "/" + str(
                        int(solveresult.split('-')[1])) + "/" + solveresult.split('-')[2]
                    # print("6", i, strresult,a[i])
                    df.loc[a[i], '解决版本'] = strresult  # 根据索引来改变名字

                if (("0000" not in closeresult) & ("-" in closeresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = closeresult.split('-')[0] + "/" + str(
                        int(closeresult.split('-')[1])) + "/" + closeresult.split('-')[2]
                    # print("8", i, strresult,a[i])
                    df.loc[a[i], '由谁关闭'] = strresult  # 根据索引来改变名字
                if (("0000" not in lastupresult) & ("-" in lastupresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = lastupresult.split('-')[0] + "/" + str(
                        int(lastupresult.split('-')[1])) + "/" + lastupresult.split('-')[2]
                    # print("9", i, strresult,a[i])
                    df.loc[a[i], '最后修改者'] = strresult  # 根据索引来改变名字
            df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)

    # 2.修改文件第二步：修改数列--添加产品线、负责人、负责部门
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k], encoding='UTF-8')
        print(3, Folder_Path + '\\' + file_list[k])
        df = pd.read_csv(f, header=0, index_col=0)
        if(len(df)!=0):
            a = df.index.tolist()  # 行索引
            # print("索引:",a)
            for i in range(len(df)):
                # print("第几行：", i)
                # 添加产品线
                productName = df.iloc[i, 0]  # 所属产品名字
                # print("productName:",productName)
                if (productName != ""):
                    productName = productName.split('(')[0]
                    ind = -1
                    for pro in products:
                        ind = ind + 1
                        for j in pro:
                            if (j == productName):
                                m = ind
                                # print(productName, m)
                                productlinename = prolines[m]  # 产品线名称
                                # print("productlinename:", productlinename)
                                df.loc[a[i], '附件'] = productlinename  # 根据索引来改变名字
                                break

                    # print("修改产品线成功")

                # 添加责任人、责任部门
                # print("修改责任人：", i)
                appointName = df.iloc[i, 21]  # 指派人名字
                bugState = df.iloc[i, 13]  # bug状态
                solveName=df.iloc[i, 23]  # 解决人名字
                closeName=df.iloc[i, 27]  # 关闭名字
                respoDepart = ""
                # print(appointName,bugState)
                if (bugState == "激活"):  # 如果bug是激活状态的话
                    # print("责任人名字：", appointName)
                    if (pd.isnull(appointName) == False):  # 指派人不为空
                        df.loc[a[i], '责任部门'] = appointName  # 根据索引来改变名字
                        ind = -1
                        for na in names:
                            ind = ind + 1
                            for name in na:
                                if (appointName in name):
                                    respoDepart = departs[ind]
                                    # print("修改责任部门为：" + respoDepart)
                                    df.loc[a[i], '产品线'] = respoDepart
                elif ((bugState == "已关闭")|(bugState == "已解决")):
                    if (pd.isnull(solveName) == False):  # 解决者不为空
                        # print("责任人名字：", solveName)
                        df.loc[a[i], '责任部门'] = solveName  # 根据索引来改变名字
                        ind = -1
                        for na in names:
                            ind = ind + 1
                            for name in na:
                                if (solveName in name):
                                    respoDepart = departs[ind]
                                    # print("修改责任部门为：" + respoDepart)
                                    df.loc[a[i], '产品线'] = respoDepart
            df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)


    # 3.修改文件第3步：添加要求解决时间，解决及时率
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k], encoding='UTF-8')
        print(4, Folder_Path + '\\' + file_list[k])
        df = pd.read_csv(f, header=0, index_col=0)
        if (len(df) != 0):
            a = df.index.tolist()  # 行索引
            for i in range(len(df)):
                endcol = 14  # 截止日期
                buildcol = 19  # 创建日期
                appointcol = 22  # 指派日期
                solvecol = 26  # 解决时间
                closecol = 28  # 关闭时间

                priority=df.iloc[i, 8]  #优先级
                bugstate=df.iloc[i, 13] #bug状态
                bulidtime = df.iloc[i, 19]  # 创建时间
                endcolresult=df.iloc[i, 14] #截止日期
                solvetime=df.iloc[i,26]  #解决时间
                closetime=df.iloc[i,28] #关闭时间

                buildyear=int(bulidtime.split('/')[0])
                buildmonth=int(bulidtime.split('/')[1])
                buildday=int(bulidtime.split('/')[2])
                flag = is_leap_year(buildyear)
                if (flag == True):  # 该年是闰年
                    days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
                else:
                    days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
                if(pd.isnull(priority) == False):#优先级不为空
                    if (int(priority) == 1):  # bug要立即解决
                        day = 2
                    elif (int(priority) == 2):  # bug要急需解决
                        day = 5
                    elif (int(priority) == 3):  # bug重要不紧急
                        day = 7
                    elif (int(priority) == 4):  # bug正常处理
                        day = 17
                else:# bug默认缺省值
                    day=17

                # print("一开始day：", day,bulidtime)
                # moday=day
                # #判断是否为节假日
                # if (buildmonth==12):
                #     if(buildday!=31):
                #         if (days[buildmonth - 1] - buildday < moday):
                #             for d in range(buildday + 1, days[11]+1):
                #                 if (len(str(d)) == 1):
                #                     datestr = str(buildyear) + "120" + str(d)
                #                 else:
                #                     datestr = str(buildyear) + "12" + str(d)
                #                 isHoliday = getHoliday(datestr)  # 1和3代表节假日,2和1是工作日
                #                 if ((isHoliday == 1) | (isHoliday == 3)):
                #                     day = day + 1
                #             for d in range(1, moday-(days[11]-buildday)+1):
                #                 if (len(str(d)) == 1):
                #                     datestr = str(buildyear + 1) + "010" + str(d)
                #                 else:
                #                     datestr = str(buildyear + 1) + "01" + str(d)
                #                 isHoliday = getHoliday(datestr)  # 1和3代表节假日,2和1是工作日
                #                 if ((isHoliday == 1) | (isHoliday == 3)):
                #                     day = day + 1
                #         else:
                #             for d in range(buildday+1,buildday+moday+1):
                #                 if (len(str(d)) == 1):
                #                     datestr = str(buildyear) + "120" + str(d)
                #                 else:
                #                     datestr = str(buildyear ) + "12" + str(d)
                #                 isHoliday = getHoliday(datestr)  # 1和3代表节假日,2和1是工作日
                #                 if ((isHoliday == 1) | (isHoliday == 3)):
                #                     day = day + 1
                #
                #     else:
                #         for d in range(1, moday):
                #             if(len(str(d))==1):
                #                 datestr = str(buildyear + 1) + "010" + str(d)
                #             else:
                #                 datestr = str(buildyear + 1) + "01" + str(d)
                #             isHoliday=getHoliday(datestr)#1和3代表节假日,2和1是工作日
                #             if((isHoliday==1)|(isHoliday==3)):
                #                 day=day+1
                # else:#不为12月
                #     if (buildday != days[buildmonth-1]):
                #         if (days[buildmonth - 1] - buildday < moday):
                #             for d in range(buildday + 1, days[11] + 1):
                #                 if (len(str(buildmonth)) == 1):
                #                     datestr = str(buildyear) +"0"+str(buildmonth)
                #                 else:
                #                     datestr = str(buildyear) + str(buildmonth)
                #                 if(len(str(d))==1):
                #                     datestr=datestr+"0" + str(d)
                #                 else:
                #                     datestr = datestr + str(d)
                #                 isHoliday = getHoliday(datestr)  # 1和3代表节假日,2和1是工作日
                #                 if ((isHoliday == 1) | (isHoliday == 3)):
                #                     day = day + 1
                #             for d in range(1, moday - (days[buildmonth-1] - buildday) + 1):
                #                 if (len(str(buildmonth)) == 1):
                #                     datestr = str(buildyear) +"0"+str(buildmonth+1)
                #                 else:
                #                     datestr = str(buildyear) + str(buildmonth+1)
                #                 if (len(str(d)) == 1):
                #                     datestr = datestr+"0" + str(d)
                #                 else:
                #                     datestr = datestr+ str(d)
                #                 isHoliday = getHoliday(datestr)  # 1和3代表节假日,2和1是工作日
                #                 if ((isHoliday == 1) | (isHoliday == 3)):
                #                     day = day + 1
                #         else:
                #             for d in range(buildday + 1, buildday + moday + 1):
                #
                #                 if (len(str(buildmonth)) == 1):
                #                     datestr = str(buildyear) +"0"+str(buildmonth)
                #                 else:
                #                     datestr = str(buildyear) + str(buildmonth)
                #                 if (len(str(d)) == 1):
                #                     datestr = datestr+"0" + str(d)
                #                 else:
                #                     datestr = datestr+ str(d)
                #                 isHoliday = getHoliday(datestr)  # 1和3代表节假日,2和1是工作日
                #                 if ((isHoliday == 1) | (isHoliday == 3)):
                #                     day = day + 1
                #     else:#为31号
                #         for d in range(1, moday):
                #             if (len(str(buildmonth)) == 1):
                #                 datestr = str(buildyear) + "0" + str(buildmonth + 1)
                #             else:
                #                 datestr = str(buildyear) + str(buildmonth + 1)
                #             if (len(str(d)) == 1):
                #                 datestr = datestr+"0" + str(d)
                #             else:
                #                 datestr = datestr+ str(d)
                #             isHoliday = getHoliday(datestr)  # 1和3代表节假日,2和1是工作日
                #             if ((isHoliday == 1) | (isHoliday == 3)):
                #                 day = day + 1
                #
                # print("最后day：",day)
                if (days[buildmonth-1] - buildday < day):
                    if(buildmonth==12):#如果该月为12月
                        endyear = buildyear + 1
                        endmonth = 1
                        endday = day - (days[11] - buildday)
                    else:
                        endyear = buildyear
                        endmonth = buildmonth+1
                        endday = day - (days[buildmonth-1] - buildday)
                else:
                    endyear = buildyear
                    endmonth = buildmonth
                    endday = buildday+day


                today = time.strftime("%Y/%m/%d")
                if (pd.isnull(priority) == False):  # 优先级不为空
                    if (int(priority) == 4 and "0000" not in endcolresult):  # 优先级等于4的时候，要求截止解决时间为截止日期
                        endtime=endcolresult
                    else:
                        endtime = str(endyear) + "/" + str(endmonth) + "/" + str(endday)  # 截止解决bug时间（自己得到的）
                else:  # bug默认缺省值
                    endtime = str(endyear) + "/" + str(endmonth) + "/" + str(endday)  # 截止解决bug时间（自己得到的）

                # print("solvetime:",endtime)
                strftime = datetime.datetime.strptime(today, "%Y/%m/%d")
                strftime2 = datetime.datetime.strptime(endtime, "%Y/%m/%d")
                df.loc[a[i], '责任人'] = endtime  # 根据索引来改变名字
                if((bugstate=="已解决") | (bugstate=="已关闭")):
                    if("0000" not in solvetime):#如果解决时间不为空
                        strftime3= datetime.datetime.strptime(solvetime, "%Y/%m/%d")
                        if (strftime2 < strftime3):#要求截止时间小于解决时间
                            df.loc[a[i], '要求解决时间'] = "延时解决"  # 根据索引来改变名字
                        else:
                            df.loc[a[i], '要求解决时间'] = "按时解决"  # 根据索引来改变名字
                    else:#如果解决时间为空
                        strftime4 = datetime.datetime.strptime(closetime, "%Y/%m/%d")
                        if (strftime2 < strftime4):#要求截止时间小于关闭时间
                            df.loc[a[i], '要求解决时间'] = "延时解决"  # 根据索引来改变名字
                        else:
                            df.loc[a[i], '要求解决时间'] = "按时解决"  # 根据索引来改变名字
                elif(bugstate=="激活"):
                    df.loc[a[i], '要求解决时间'] = "未解决"  # 根据索引来改变名字
                    # if (strftime2 < strftime):#要求截止时间小于今天
                    #     df.loc[a[i], '要求解决时间'] = "未解决"  # 根据索引来改变名字
            df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)

    # 4.修改文件第4步：添加回归测试及时性及测试回归时间
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k], encoding='UTF-8')
        print(5, Folder_Path + '\\' + file_list[k])
        df = pd.read_csv(f, header=0, index_col=0)
        if (len(df) != 0):
            a = df.index.tolist()  # 行索引
            for i in range(len(df)):
                endcol = 14  # 截止日期
                buildcol = 19  # 创建日期
                appointcol = 22  # 指派日期
                solvecol = 26  # 解决时间
                closecol = 28  # 关闭时间

                priority = df.iloc[i, 8]  # 优先级
                bulidtime = df.iloc[i, 19]  # 创建时间
                endcolresult = df.iloc[i, 14]  # 截止日期
                solvetime = df.iloc[i, 26]  # 解决时间

                closetime = df.iloc[i, 28]  # 关闭时间
                bugstate = df.iloc[i, 13]  # bug状态
                bugid = a[i]  # bug id
                activetime = df.iloc[i, 15] #激活次数
                date=""
                for dateid in dateids:
                    if(int(dateid.split('+')[1])==int(bugid)):
                        date=dateid.split('+')[0]
                        break
                # print("bugid,date",bugid,"  ",date)
                if(date!=""):
                    year = int(date.split('-')[0])
                    month = int(date.split('-')[1])
                    day = int(date.split('-')[2])
                    flag = is_leap_year(year)
                    if (flag == True):  # 该年是闰年
                        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
                    else:
                        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
                    starttime = date   #回归的截止开始日期
                    #假设一个月31天
                    if (month == 12):  # 就要计算到下一年的1月
                        endtime = time.strftime(
                            str(year + 1) + "/" + str(1) + "/" + str(31 - (days[11] - day)))  # 回归的截止结束日期
                    elif (month == 1):  # 计算到2月或者3月
                        if (days[1] + (days[0] - day)<31):
                            endtime = time.strftime(
                                str(year ) + "/" + str(3) + "/" + str(31 - (days[1] + (days[0] - day))))  # 回归的截止结束日期
                        else:
                            endtime = time.strftime(
                                str(year) + "/" + str(2) + "/" + str(31 - (days[0] - day)))  # 回归的截止结束日期
                    else:  # 就要计算到下一月或者下下一月
                        if (days[month] + (days[month-1] - day)<31):
                            endtime = time.strftime(
                                str(year ) + "/" + str(month+2) + "/" + str(31 - (days[month] + (days[month-1] - day))))  # 回归的截止结束日期
                        else:
                            endtime = time.strftime(
                                str(year) + "/" + str(month + 1) + "/" + str(31 - (days[month-1] - day)))  # 回归的截止结束日期
                    # print(2222,starttime,endtime)
                    #添加回归截止时间
                    df.loc[a[i], '解决及时性'] = endtime  # 根据索引来改变名字
                    #添加回归测试结果
                    strftime2 = datetime.datetime.strptime(endtime, "%Y/%m/%d")  #回归的截止结束日期
                    if (bugstate == "已关闭"):#bug状态为关闭
                        if ("0000" not in closetime):  # 关闭时间不为空
                            strftime = datetime.datetime.strptime(closetime, "%Y/%m/%d")  # 关闭时间
                            if (strftime2 >= strftime):
                                df.loc[a[i], '回归截止时间'] = "准时回归"  # 根据索引来改变名字
                            else:
                                df.loc[a[i], '回归截止时间'] = "延时回归"  # 根据索引来改变名字
                                # 添加回归测试时间
                            # print("测试回归时间:", closetime)
                            df.loc[a[i], '回归测试结果'] = closetime  # 根据索引来改变名字
                    else:#bug状态不关闭
                        if (int(activetime) == 0):  # 激活次数为0
                            df.loc[a[i], '回归截止时间'] = "未回归"  # 根据索引来改变名字
                        else:#激活次数不为0
                            # 得到bug最后一次激活的时间
                            lasttime = _getBugActivityLastTime(session, bugid)
                            # print("测试回归时间:", lasttime)
                            # 添加回归测试时间
                            df.loc[a[i], '回归测试结果'] = lasttime  # 根据索引来改变名字
                            strftime3 = datetime.datetime.strptime(lasttime, "%Y/%m/%d")
                            if (strftime2 >= strftime3):
                                df.loc[a[i], '回归截止时间'] = "准时回归"  # 根据索引来改变名字
                            else:
                                df.loc[a[i], '回归截止时间'] = "延时回归"  # 根据索引来改变名字
            df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)
    #5.有的产品线不对应产品的，所有要根据模块名字来判断产品线
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k],encoding='UTF-8')
        print(6, Folder_Path + '\\' + file_list[k])
        df = pd.read_csv(f, header=0, index_col=0)
        if (len(df) != 0):
            a = df.index.tolist()  # 行索引
            for i in range(len(df)):
                # 添加产品线
                modulename = df.iloc[i, 1]  # 模块名字
                proName=df.iloc[i, 35]  #产品线名称
                # print("第几行：", i," ",iteraName," ",proName)
                if(pd.isnull(proName) == True):#如果产品线为空，就对比产品线与模块名字
                    modul = FIND(modulename)
                    if(len(modul)>1):
                        productlinename = modul.split('/')[1] + "线"
                        df.loc[a[i], '附件'] = productlinename
                    else:
                        df.loc[a[i], '附件'] = "平台部"
            df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)


#修改csv任务文件
def  change_task_csv(session):
    # 得到产品线--产品
    product_iteras = _getProduct(session)
    products = product_iteras[0]  #产品线
    iteras = product_iteras[1] #各个产品线的所有产品

    #得到产品-任务
    productTask=_getTasks(session)

    # 得到部门下的名称和人名字
    departNamelist=_getDepartment(session)


    departs = departNamelist[0]
    # print(1,departs)
    names = departNamelist[1]
    # print(2,names)

    path = os.path.dirname(os.getcwd()) + '/file'
    Folder_Path = new_file(path)  # 得到下载任务文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    file_list = getfileName(Folder_Path)

    #有些csv文件没有相关需求属性列，在指定位置上插入一列
    #并且要子任务名称列
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k],encoding='UTF-8')
        df = pd.read_csv(f, header=0, index_col=0)
        print(0,Folder_Path + '\\' + file_list[k])
        if(len(df)!=0):
            colnames = df.columns.tolist()  # 列索引
            indexs = df.index.tolist()  # 行索引
            flag2 = False
            for colname in colnames:
                if ("相关需求" == colname):
                    flag2 = True
                    break
            # 如果相关需求不存在，就添加该列
            # print("flag2:",flag2)
            if flag2 == False:
                tasks = []
                for jj in range(len(df)):
                    task = df.iloc[jj, 2]
                    tasks.append(task)
                # 在特定列后添加列,并且添加值
                df.insert(3, '相关需求', tasks)
                # colnames.insert(colnames.index('所属模块')+1,'相关需求')
                # df=df.reindex(columns=colnames)
                for jj in range(len(df)):
                    df.loc[indexs[jj], '所属模块'] = ""
            flag3 = False
            for colname in colnames:
                if ("子任务名称" == colname):
                    flag3 = True
                    break
            if (flag3 == False):
                # 添加子任务名称列
                taskDescribes = []
                for jjj in range(len(df)):
                    taskDe = df.iloc[jjj, 4]
                    taskDescribes.append(taskDe)
                df.insert(5, "子任务名称", taskDescribes)
                for jjj in range(len(df)):
                    df.loc[indexs[jjj], '任务名称'] = ""
        df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)


    # 1.修改文件第一步：添加几列，并且删除重复行
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k],encoding='UTF-8')
        df = pd.read_csv(f, header=0, index_col=0)
        print(1,Folder_Path + '\\' + file_list[k])
        df["产品线"] = ""  # 修改文件
        df["责任部门"] = ""  # 修改文件
        df["责任人"] = ""  # 修改文件
        df["完成及时性"] = ""  # 修改文件
        df["备注"] = ""  # 修改文件
        df["最新进展"]="" #修改文件
        # print("行", df.shape[0])
        # print("列", df.shape[1])
        # print("行", len(df))
        # 删除重复行
        indexs = df.index.tolist()  # 行索引
        # print("索引indexs:",indexs)
        df.drop_duplicates(subset=['相关需求','创建日期','指派日期','关闭原因','指派给','由谁创建','进度','最初预计','任务状态','实际开始'], keep='last', inplace=True)
        # print(df)
        df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)

    # 2.修改文件第二步：修改任务名称、子任务名称
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k],encoding='UTF-8')
        df = pd.read_csv(f, header=0, index_col=0)
        print(2,Folder_Path + '\\' + file_list[k])
        if(len(df)!=0):
            df = df[df['任务类型'] > 0]  # 删除csv文件小于某个值的一行数据（删除最后一行)
            # print(1111)
            # print(df['任务类型'])
            # 得到行索引，然后根据索引去删除行
            # print(df['任务类型'])
            a = df[df['任务类型'] > 0].index.tolist()
            # print(a)

            bigstr = ">"
            fatherTaskName = ""  # 父任务名称
            index = -1  # 行索引的值
            flag = False
            deleteArray = []  # 要删除的索引数组
            for i in range(len(df)):
                # print(i)
                taskName = df.iloc[i, 3]
                # print("taskName", taskName)
                if (bigstr in taskName): #是子任务
                    if (fatherTaskName != ""):
                        flag = True
                        # 修改任务名称和子任务名称
                        # name = fatherTaskName + "+" + taskName.split('>')[1]
                        df.loc[a[i], '相关需求'] = fatherTaskName  # 根据索引来改变名字
                        # print("修改任务名字为：" fatherTaskName)
                        df.loc[a[i], '任务名称'] = taskName.split('>')[1]  # 根据索引来改变名字
                        # print("修改子任务名字为：" taskName.split('>')[1])
                else:#是父任务
                    if ((index != -1) & (flag == True)):
                        deleteArray.append(index)
                        # print("父任务：", fatherTaskName)
                        flag = False
                    fatherTaskName = taskName
                    index = a[i]
            if ((index != -1) & (flag == True)):
                deleteArray.append(index)
            # print(deleteArray)
            # 删除父任务
            for i in range(len(deleteArray)):
                df.drop(index=deleteArray[i], inplace=True)  # 删除指定行列(索引)的数据，inplace代表修改文件中的值
            df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)

    # 3.有的时间格式不一致，要修改成一样的格式时间，如2019/9/21
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k],encoding='UTF-8')
        df = pd.read_csv(f, header=0, index_col=0)
        print(3,Folder_Path + '\\' + file_list[k])
        if(len(df)!=0):
            a = df[df['任务类型'] > 0].index.tolist()#索引
            col = 0  # 指定修改的列
            exceptstartcol = 8  # 预计开始
            startcol = 9  # 实际开始
            endcol = 10  # 截止日期
            buildcol = 18  # 创建日期
            appointcol = 20  # 指派日期
            finishcol = 22  # 完成时间
            canclecol = 24  # 取消时间
            closecol = 26  # 关闭时间
            lastupcol=29   #最后修改时间
            for i in range(len(df)):
                exceptstartresult = df.iloc[i, exceptstartcol]
                startresult = df.iloc[i, startcol]
                endresult = df.iloc[i, endcol]
                buildresult = df.iloc[i, buildcol]
                appointresult = df.iloc[i, appointcol]
                finishresult = df.iloc[i, finishcol]
                cancleresult = df.iloc[i, canclecol]
                closeresult = df.iloc[i, closecol]
                lastupresult = df.iloc[i, lastupcol]
                if(("0000" not in  exceptstartresult)&("-" in exceptstartresult)):#如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult=exceptstartresult.split('-')[0]+"/"+str(int(exceptstartresult.split('-')[1]))+"/"+exceptstartresult.split('-')[2]
                    # print("1",i,strresult,a[i])
                    df.loc[a[i], '优先级'] = strresult  # 根据索引来改变名字
                if (("0000" not in startresult) & ("-" in startresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = startresult.split('-')[0] + "/" + str(
                        int(startresult.split('-')[1])) + "/" + startresult.split('-')[2]
                    # print("2", i, strresult,a[i])
                    df.loc[a[i], '预计开始'] = strresult  # 根据索引来改变名字
                if (("0000" not in endresult) & ("-" in endresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = endresult.split('-')[0] + "/" + str(
                        int(endresult.split('-')[1])) + "/" + endresult.split('-')[2]
                    # print("3", i, strresult,a[i])
                    df.loc[a[i], '实际开始'] = strresult  # 根据索引来改变名字
                if (("0000" not in buildresult) & ("-" in buildresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = buildresult.split('-')[0] + "/" + str(
                        int(buildresult.split('-')[1])) + "/" + buildresult.split('-')[2]
                    # print("4", i, strresult,a[i])
                    df.loc[a[i], '由谁创建'] = strresult  # 根据索引来改变名字
                if (("0000" not in appointresult) & ("-" in appointresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = appointresult.split('-')[0] + "/" + str(
                        int(appointresult.split('-')[1])) + "/" + appointresult.split('-')[2]
                    # print("5", i, strresult,a[i])
                    df.loc[a[i], '指派给'] = strresult  # 根据索引来改变名字
                if (("0000" not in finishresult) & ("-" in finishresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = finishresult.split('-')[0] + "/" + str(
                        int(finishresult.split('-')[1])) + "/" + finishresult.split('-')[2]
                    # print("6", i, strresult,a[i])
                    df.loc[a[i], '由谁完成'] = strresult  # 根据索引来改变名字
                if (("0000" not in cancleresult) & ("-" in cancleresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = cancleresult.split('-')[0] + "/" + str(
                        int(cancleresult.split('-')[1])) + "/" + cancleresult.split('-')[2]
                    # print("7", i, strresult,a[i])
                    df.loc[a[i], '由谁取消'] = strresult  # 根据索引来改变名字
                if (("0000" not in closeresult) & ("-" in closeresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = closeresult.split('-')[0] + "/" + str(
                        int(closeresult.split('-')[1])) + "/" + closeresult.split('-')[2]
                    # print("8", i, strresult,a[i])
                    df.loc[a[i], '由谁关闭'] = strresult  # 根据索引来改变名字
                if (("0000" not in lastupresult) & ("-" in lastupresult)):  # 如果时间不为空并且还包含这个“-”格式的话，就修改格式
                    strresult = lastupresult.split('-')[0] + "/" + str(
                        int(lastupresult.split('-')[1])) + "/" + lastupresult.split('-')[2]
                    # print("9", i, strresult,a[i])
                    df.loc[a[i], '最后修改'] = strresult  # 根据索引来改变名字
            df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)


    # 4.修改文件第三步：添加产品线，责任部门，责任人，完成及时性
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k],encoding='UTF-8')
        df = pd.read_csv(f, header=0, index_col=0)
        print(4, Folder_Path + '\\' + file_list[k])
        filename = file_list[k].split('.')[0]
        if (len(df) != 0):
            # 再一次得到行索引，然后根据索引去改变特定行列的值
            # print(df['任务类型'])
            a = df[df['任务类型'] > 0].index.tolist()
            # print(a)
            for i in range(len(df)):
                # print("第几行：", i)
                #任务描述去除&nbsp
                taskDescribe = df.iloc[i, 5]  # 任务描述
                if(pd.isnull(taskDescribe) == False):#任务描述不为空
                    taskDescribe2 = taskDescribe.strip().strip('&nbsp;')
                    taskDescribe2=taskDescribe.replace('&nbsp;','')#都替换为空
                    # print("第几行：", i, taskDescribe,taskDescribe2)
                    df.loc[a[i], '子任务名称'] = taskDescribe2  # 根据索引来改变名字

                # 添加产品线
                taskName = df.iloc[i, 3].strip()  # 父任务名字
                # print("taskName:", taskName)
                if("[多人]" in taskName):
                    taskName=taskName.split(']')[1].strip()
                product = ""  # 产品名字
                # print("productTask:",productTask)
                # print("taskname",taskName)
                for pro in productTask:
                    if (taskName ==pro.split('+')[1].strip()):
                        product = pro.split('+')[0].strip()
                        break
                # print("product:", product)
                if (product != ""):
                    productlinename = ""  # 产品线名称
                    ind = -1
                    for itera in iteras:
                        ind = ind + 1
                        for j in itera:
                            if (j == product):
                                m = ind
                                productlinename = products[m]
                                # print("productlinename:", productlinename)
                                df.loc[a[i], '附件'] = productlinename  # 根据索引来改变名字
                                # print("修改产品线成功")
                                break
                #添加完成及时性
                taskState = df.iloc[i, 11]  # 任务状态
                endtime=df.iloc[i, 10]  #截止日期
                completetime=df.iloc[i, 22]  #完成日期
                today = time.strftime("%Y/%m/%d")
                if ((taskState == "已完成") |(taskState == "已关闭")):
                    # endstr = int(endtime.split('-')[0] + endtime.split('-')[1] + endtime.split('-')[2])
                    # completestr = int(completetime.split('-')[0] + completetime.split('-')[1] + completetime.split('-')[2])
                    # if ((endstr != 0)&(completestr!= 0) ):#截止时间和完成时间不为空
                    if (("0000" not in endtime) & ("0000" not in completetime)):  # 截止时间和完成时间不为空
                        strftime = datetime.datetime.strptime(endtime, "%Y/%m/%d")
                        strftime2 = datetime.datetime.strptime(completetime, "%Y/%m/%d")
                        if (strftime >= strftime2):
                            # print("已完成")
                            df.loc[a[i], '责任人'] = "按时完成"  # 根据索引来改变名字
                        else:
                            # print("已完成")
                            df.loc[a[i], '责任人'] = "延期完成"  # 根据索引来改变名字
                elif ((taskState == "进行中")|(taskState == "未开始")):
                    # endstr = int(endtime.split('-')[0] + endtime.split('-')[1] + endtime.split('-')[2])
                    # print(endstr)
                    if ("0000" not in endtime):#截止时间不为空
                        strftime = datetime.datetime.strptime(today, "%Y/%m/%d")
                        strftime2 = datetime.datetime.strptime(endtime, "%Y/%m/%d")
                        if (strftime > strftime2):
                            # print("未完成")
                            df.loc[a[i], '责任人'] = "未完成"  # 根据索引来改变名字
                # 添加责任人、责任部门
                # print("修改责任人：", i)
                respoName = df.iloc[i, 19]  # 指派人名字--负责人
                taskState = df.iloc[i, 11]  # 任务状态
                respoDepart = ""
                if (taskState == "已完成"):  # 如果任务已经关闭或者已完成
                    respoName2 = df.iloc[i, 21]  # 完成者
                    # print("责任人名字：", respoName2)
                    df.loc[a[i], '责任部门'] = respoName2  # 根据索引来改变名字
                    ind = -1
                    for na in names:
                        ind = ind + 1
                        for name in na:
                            # print(respoName2,name)
                            if (respoName2 in name):
                                respoDepart = departs[ind]
                                # print("修改责任部门为：" + respoDepart)
                                df.loc[a[i], '产品线'] = respoDepart
                elif (taskState == "已关闭"):
                    respoName2 = df.iloc[i, 21]  # 完成者
                    closeName = df.iloc[i, 25]  # 关闭者
                    # print("责任人名字：", respoName2)
                    if (pd.isnull(respoName2) == True):  # 完成者为空
                        respoName2 = closeName
                    df.loc[a[i], '责任部门'] = respoName2  # 根据索引来改变名字
                    ind = -1
                    for na in names:
                        ind = ind + 1
                        for name in na:
                            # print(respoName2,name)
                            if (respoName2 in name):
                                respoDepart = departs[ind]
                                # print("修改责任部门为：" + respoDepart)
                                df.loc[a[i], '产品线'] = respoDepart
                elif (taskState == "已取消"):
                    respoName2 = df.iloc[i, 23]  # 取消者
                    # print("责任人名字：", respoName2)
                    df.loc[a[i], '责任部门'] = respoName2  # 根据索引来改变名字
                    ind = -1
                    for na in names:
                        ind = ind + 1
                        for name in na:
                            if (respoName2 in name):
                                respoDepart = departs[ind]
                                # print("修改责任部门为：" + respoDepart)
                                df.loc[a[i], '产品线'] = respoDepart
                else:
                    # print("责任人名字：", respoName)
                    if (pd.isnull(respoName) == False):  # 指派人不为空
                        df.loc[a[i], '责任部门'] = respoName  # 根据索引来改变名字
                        # print("修改责任人为：" + respoName)
                        ind = -1
                        for na in names:
                            ind = ind + 1
                            for name in na:
                                if (respoName in name):
                                    respoDepart = departs[ind]
                                    # print("修改责任部门为：" + respoDepart)
                                    df.loc[a[i], '产品线'] = respoDepart
            # print("最终")
            # print(df['产品线'])
            # print(df['责任部门'])
            df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)

    #5.有的产品线不对应模块的，所有要根据迭代名字来判断产品线
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k],encoding='UTF-8')
        df = pd.read_csv(f, header=0, index_col=0)
        print(5, Folder_Path + '\\' + file_list[k])
        if (len(df) != 0):
            a = df[df['任务类型'] > 0].index.tolist()
            for i in range(len(df)):
                # 添加产品线
                iteraName = df.iloc[i, 0]  # 迭代名字
                proName=df.iloc[i, 31]  #产品线名称
                # print("第几行：", i," ",iteraName," ",proName)
                if(pd.isnull(proName) == True):#如果产品线为空，就对比产品线与迭代名字
                    if("ZEDNE Sprint" in iteraName):
                        df.loc[a[i], '附件'] = "平台部"
                    else:
                        ite = iteraName[0:2]
                        # print("ite:",ite)
                        for na in products:
                            if (ite in na):
                                df.loc[a[i], '附件'] = na
                                break
            df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)

    #6.添加文件第三步：添加任务最后一条备注
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k],encoding='UTF-8')
        df = pd.read_csv(f, header=0, index_col=0)
        print(6, Folder_Path + '\\' + file_list[k])
        if (len(df) != 0):
            indexs = df[df['任务类型'] > 0].index.tolist()
            for i in  range(0,len(indexs)):
                index=indexs[i]
                # print("indexindex:",index)
                rs = session.get(url + "/zentao/task-view-"+str(index)+".html")
                rs.encoding = 'utf-8'
                soup = BeautifulSoup(rs.text, "html.parser")
                taskTimeDescribe="" #任务的最后一条描述和时间
                for ol in soup.find_all('ol', class_='histories-list'):
                    # print("111:",ol)
                    for li in ol.findAll("li"):
                        # print("li:",li)
                        if("article-content comment" in str(li)):
                            taskDescribetime=str(li).strip().split(">")[1].split(",")[0].strip()
                            taskDescribeUser=str(li).strip().split(">")[2].split("<")[0].strip()  #谁添加的备注
                            for tag in li.find_all('div', class_='article-content comment'):
                                # print("tag:",i,tag)
                                ss=str(tag).strip().split(">")
                                comment = ""
                                for j in range(0,len(ss)):
                                    # print(j, ss[j])
                                    if(j!=0):
                                        comment0=ss[j].split("<")[0].strip()
                                        # print("jjjj:",j,comment0)
                                        if(comment0!="" and comment0!=None):
                                            comment = comment + " " + comment0
                            taskTimeDescribe =taskDescribeUser+" "+taskDescribetime+" "+comment
                df.loc[indexs[i], '备注'] = taskTimeDescribe
            df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)


#修改csv需求文件
def  change_request_csv(session):
    # 得到产品线--产品
    product_iteras = _getProduct(session)
    products = product_iteras[0]  #产品线
    iteras = product_iteras[1] #各个产品线的所有产品

    # 得到部门下的名称和人名字
    departNamelist=_getDepartment(session)

    departs = departNamelist[0]
    print(1,departs)
    names = departNamelist[1]
    print(2,names)

    path = os.path.dirname(os.getcwd()) + '/file'
    Folder_Path = new_file(path)  # 得到下载任务文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    file_list = getfileName(Folder_Path)

    # 1.修改文件第一步：添加几列，并且删除重复行
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k],encoding='UTF-8')
        df = pd.read_csv(f, header=0, index_col=0)
        print(1,Folder_Path + '\\' + file_list[k])
        df["提需求部门"] = ""  # 修改文件
        df["需求所属产品线"] = ""  # 修改文件
        df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)

    # 2.修改文件第二步：添加提需求部门、需求所属产品线
    for k in range(0, len(file_list)):
        f = open(Folder_Path + '\\' + file_list[k],encoding='UTF-8')
        df = pd.read_csv(f, header=0, index_col=0)
        column_headers = list(df.columns.values)
        # print(2222,column_headers)
        print(2, Folder_Path + '\\' + file_list[k])
        filename = file_list[k].split('.')[0]
        if (len(df) != 0):
            print(2222, Folder_Path + '\\' + file_list[k])
            # 再一次得到行索引，然后根据索引去改变特定行列的值
            a = df[df['关键词'] > 0].index.tolist()
            for i in range(len(df)):
                # 添加需求所属产品线
                productName = df.iloc[i, 0].strip()  # 所属产品名称
                if("(" in productName):
                    productName = productName.split("(")[0].strip()
                if (productName != ""):
                    productlinename = ""  # 产品线名称
                    ind = -1
                    for itera in iteras:
                        ind = ind + 1
                        for j in itera:
                            if (j == productName):
                                m = ind
                                productlinename = products[m]
                                df.loc[a[i], '提需求部门'] = productlinename  # 根据索引来改变名字
                                break
                # 添加提需求部门
                requestCreateName = df.iloc[i, 16]  # 需求创建人
                # print("需求创建人:" + requestCreateName+" 所属产品名称："+productName)
                requestDapart = ""
                if(requestCreateName!=""):
                    ind = -1
                    for na in names:
                        ind = ind + 1
                        for name in na:
                            if (requestCreateName in name):
                                requestDapart = departs[ind]
                                df.loc[a[i], '附件'] = requestDapart
            df.to_csv(Folder_Path + '\\' + file_list[k], encoding="utf_8_sig", index_label=False)


#得到测试单-测试人、版本号
def getExcuteVersionTestExcel(session):
    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹

    if (os.path.exists(Folder_Path + "/用例测试单.xls") == True):
        os.remove(Folder_Path + "/用例测试单.xls")

    departNamelist = _getDepartment(session)
    departs = departNamelist[0]  # 部门
    names = departNamelist[1]  # 人

    # 得到产品线--产品
    product_iteras = _getProduct(session)
    products = product_iteras[0]  # 产品线
    iteras = product_iteras[1]  # 各个产品线的所有产品


    today = time.strftime("%Y-%m-%d")  # 今天
    day_num = int(today.split('-')[2])  # 几号
    month_num = int(today.split('-')[1])  # 几月
    year_num = int(today.split('-')[0])  # 几年

    #版本名称和链接
    all=_getTestCases(session)
    versions=all[0]
    versionurls=all[1]

    book = Workbook(encoding='utf-8')
    sheet = book.add_sheet('data')  # 创建一个sheet
    sheet.write(0, 0, "用例编号")
    sheet.write(0, 1, "版本")
    sheet.write(0, 2, "用例标题")
    sheet.write(0, 3, "用例类型")
    sheet.write(0, 4, "指派人")
    sheet.write(0, 5, "执行人")
    sheet.write(0, 6, "部门")
    sheet.write(0, 7, "开始日期")
    sheet.write(0, 8, "结束日期")
    sheet.write(0, 9, "版本状态")
    sheet.write(0, 10, "用例状态")
    sheet.write(0, 11, "执行时间")
    sheet.write(0, 12, "执行结果")
    sheet.write(0, 13, "所属产品")
    sheet.write(0, 14, "产品线")
    sheet.write(0, 15, "执行及时性")

    mi = 0  # excel的第几行
    for jj in range(0,len(versionurls)):
        rs = session.get(versionurls[jj])
        rs.encoding = 'utf-8'
        # print(rs.text)
        # print("url:",versionurls[jj],versions[jj])
        if ("data-rec-total" in str(rs.text)):
            soup = BeautifulSoup(rs.text, "html.parser")
            for tbody in soup.find_all('tbody'):
                for tr in tbody.find_all('tr'):
                    mi = mi + 1
                    id = ""
                    testname = ""
                    testtype = ""
                    appointname = ""  # 指派人
                    excutename = ""  # 执行人
                    excutetime = ""  # 执行时间
                    result = ""
                    state = ""
                    for td in tr.find_all('td'):
                        if (id == ""):
                            for input in td.find_all('input'):
                                id = input.attrs.get('value')
                                break
                        elif (testname == ""):
                            for a in td.find_all('a'):
                                testname = a.string
                                break
                        elif (testtype == ""):
                            testtype = td.string
                        elif (appointname == ""):
                            appointname = td.string
                        elif (excutename == ""):
                            excutename = td.string
                        elif (excutetime == ""):
                            excutetime = td.string
                        elif (result == ""):
                            for span in td.find_all('span'):
                                result = span.string
                                break
                        elif (state == ""):
                            state = td.string
                        else:
                            break
                    # print("2:", id, versions[jj], testname, testtype, appointname, excutename, excutetime,
                    #       result, state)
                    # 得到部门并且添加到excel中
                    ind = -1
                    # print("excutename,appointname:",excutename,appointname)
                    responame=""#责任人
                    if ((excutename != None)):  # 不为空且不是全字母
                        res1 = True
                        for w in str(excutename):
                            if not '\u4e00' <= w <= '\u9fff':
                                res1 = False
                        if (res1 == True):
                            responame=excutename
                        # print(11,res1)
                    if ((appointname != None)&(responame=="")):  # 不为空且不是全字母
                        res2 = True
                        for w in str(appointname):
                            if not '\u4e00' <= w <= '\u9fff':
                                res2 = False
                        if (res2 == True):
                            responame = appointname
                        # print(22,res2)
                    # print("责任人：",responame)
                    respoDepart = ""
                    for na in names:
                        ind = ind + 1
                        for name in na:
                            if (responame == name):
                                # print("11111")
                                respoDepart = departs[ind]  # 部门
                    sheet.write(mi, 0, id)
                    sheet.write(mi, 1, versions[jj][0])
                    sheet.write(mi, 2, testname)
                    sheet.write(mi, 3, testtype)
                    sheet.write(mi, 4, appointname)
                    sheet.write(mi, 5, excutename)
                    sheet.write(mi, 6, respoDepart)
                    sheet.write(mi, 7, versions[jj][1])
                    sheet.write(mi, 8, versions[jj][2])
                    sheet.write(mi, 9, versions[jj][3])
                    sheet.write(mi, 10, state)
                    sheet.write(mi, 11, excutetime)
                    sheet.write(mi, 12, result)
                    sheet.write(mi, 13, versions[jj][4])
                    #得到产品线
                    productlinename = ""  # 产品线名称
                    if (versions[jj][4] != ""):
                        ind = -1
                        for itera in iteras:
                            ind = ind + 1
                            for j in itera:
                                if (j == versions[jj][4]):
                                    m = ind
                                    productlinename = products[m]
                                    # print("productlinename:", productlinename)
                                    break
                    if(productlinename!=""):
                        sheet.write(mi, 14, productlinename)
                    else:#产品线为空
                        if(respoDepart!=""):
                            ros = FIND(respoDepart)
                            productlinename = ros + "线"
                            sheet.write(mi, 14, productlinename)
                    if(excutetime==None):
                        sheet.write(mi, 15, "未执行")
                    else:
                        ex=str(year_num)+"-"+str(excutetime.split(' ')[0].split('/')[0])+"-"+str(excutetime.split(' ')[0].split('/')[1])
                        strftime = datetime.datetime.strptime(ex, "%Y-%m-%d")  # 执行时间
                        strftime2 = datetime.datetime.strptime(versions[jj][1], "%Y-%m-%d")  # 开始时间
                        strftime3 = datetime.datetime.strptime(versions[jj][2], "%Y-%m-%d")  # 结束时间
                        if((strftime<=strftime3)&(strftime>=strftime2)):
                            sheet.write(mi, 15, "准时执行")
                        else:
                            sheet.write(mi, 15, "延时执行")
    book.save(Folder_Path + "/用例测试单.xls")




#得到测试单-负责人、版本号
def getResponVersionTestExcel(session):

    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹

    # 得到产品线--产品
    product_iteras = _getProduct(session)
    products = product_iteras[0]  # 产品线
    iteras = product_iteras[1]  # 各个产品线的所有产品

    if (os.path.exists(Folder_Path+"/版本测试单.xls") == True):
        os.remove(Folder_Path+"/版本测试单.xls")

    departNamelist = _getDepartment(session)
    departs = departNamelist[0]  # 部门
    names = departNamelist[1]  # 人

    testhrefs=_getVersions(session)

    book = Workbook(encoding='utf-8')
    sheet = book.add_sheet('data')  # 创建一个sheet
    sheet.write(0, 0, "版本编号")
    sheet.write(0, 1, "名称")
    sheet.write(0, 2, "所属产品")
    sheet.write(0, 3, "所属迭代")
    sheet.write(0, 4, "版本")
    sheet.write(0, 5, "负责人")
    sheet.write(0, 6, "部门")
    sheet.write(0, 7, "开始日期")
    sheet.write(0, 8, "结束日期")
    sheet.write(0, 9, "状态")
    sheet.write(0, 10, "产品线")
    mi = 0  # excel的第几行
    for testhref in testhrefs:
        rs = session.get(testhref)
        rs.encoding = 'utf-8'
        # print("testhref:",testhref)
        soup = BeautifulSoup(rs.text, "html.parser")
        for tbody in soup.find_all('tbody'):
            for tr in tbody.find_all('tr'):
                mi=mi+1
                id =""
                versionname="" #版本名称
                productname=""
                iteraname=""
                versionnum=""  #版本号
                people=""
                start=""
                end=""
                state=""
                for td in tr.find_all('td'):
                    if(id==""):
                        for a in td.find_all('a'):
                            id=a.string
                            break
                    elif(versionname==""):
                        for a in td.find_all('a'):
                            versionname=a.string
                            break
                    elif (productname == ""):
                        productname = td.string
                    elif (iteraname == ""):
                        iteraname = td.string
                    elif (versionnum == ""):
                        if("/zentao/build-view" in str(td)):
                            for a in td.find_all('a'):
                                versionnum = a.string
                                break
                        else:
                            versionnum=td.string
                    elif (people == ""):
                        people = td.string
                    elif (start == ""):
                        start = td.string
                    elif (end == ""):
                        end = td.string
                    elif (state == ""):
                        state = td.attrs.get('title')
                    else:
                        break
                # print("1:", id, versionname, productname, iteraname, versionnum, people, start, end, state)
                #得到部门并且添加到excel中
                ind = -1
                respoDepart=""
                for na in names:
                    ind = ind + 1
                    for name in na:
                        if(people!=None and people!=""):
                            if (people in name):
                                respoDepart = departs[ind]  # 部门

                #得到产品线，并且添加到excel中
                productlinename = ""  # 产品线名称
                if (productname != ""):
                    ind = -1
                    for itera in iteras:
                        ind = ind + 1
                        for j in itera:
                            if (j == productname):
                                m = ind
                                productlinename = products[m]
                                # print("productlinename:", productlinename)
                                break
                sheet.write(mi, 0, id)
                sheet.write(mi, 1, versionname)
                sheet.write(mi, 2, productname)
                sheet.write(mi, 3, iteraname)
                sheet.write(mi, 4, versionnum)
                sheet.write(mi, 5, people)
                sheet.write(mi, 6, respoDepart)
                sheet.write(mi, 7, start)
                sheet.write(mi, 8, end)
                sheet.write(mi, 9, state)
                if(productlinename!=""):
                    sheet.write(mi, 10, productlinename)
                else:
                    if(respoDepart!=""): #部门不为空
                        ros=FIND(respoDepart)
                        productlinename=ros+"线"
                        sheet.write(mi, 10, productlinename)

    book.save(Folder_Path + "/版本测试单.xls")


#得到一月的任务的excel，并且修改
def getMonthTasksExcel(isHalfYear):

    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    xlsxpath=Folder_Path+"/tasks.xls"
    # print(xlsxpath)
    # xlsxpath = "E:\\PythonWorkspace\\ZentaoTest\\result\\20190410 160724\\20190410160724tasks.xls"
    if (int(isHalfYear) == 0):
        allpath = Folder_Path + "/一月tasks.xls"
    else:
        allpath = Folder_Path + "/半年tasks.xls"
    if (os.path.exists(allpath) == True):
        os.remove(allpath)

    bk = xlrd.open_workbook(xlsxpath)
    shxrange = range(bk.nsheets)
    try:
        sh = bk.sheet_by_name("data")
    except:
        print("代码出错")
    #35列
    nrows = sh.nrows  # 获取行数
    ncols=sh.ncols# 获取列数
    book = Workbook(encoding='utf-8')
    sheet = book.add_sheet('data')  # 创建一个sheet

    today = time.strftime("%Y/%m/%d")#今天
    day_num=int(today.split('/')[2])  #几号
    month_num=int(today.split('/')[1])  #几月
    year_num=int(today.split('/')[0]) #几年
    flag=is_leap_year(year_num)
    if(flag==True):#该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    if(int(isHalfYear)==0):
        # if (month_num == 1):
        #     if (int(day_num) < 27):#那就上一年的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(11) + "/" + str(days[10]-3))  # 开始日期
        #         endtime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(27))  # 结束日期
        #     else:#那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        # elif (month_num == 2):
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        # else:
        #     # print("day_num:",day_num)
        #     if (int(day_num) <27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-2) + "/" + str(days[month_num-3] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        if (month_num == 1):  # 就要计算到上一年的12月
            starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(1))  # 开始日期
            endtime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11]))  # 结束日期
        else:  # 就要计算到上一月
            starttime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(1))  # 开始日期
            endtime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(days[month_num - 2]))  # 结束日期
        print("一月任务starttime,endtime:", starttime, endtime)
    elif(int(isHalfYear)==1):
        starttime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num ) + "/" + str(6) + "/" + str(days[5]))  # 结束日期
        print("前半年任务starttime,endtime:", starttime, endtime)
    elif (int(isHalfYear) == 2):
        starttime = time.strftime(str(year_num ) + "/" + str(7) + "/" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num ) + "/" + str(12) + "/" + str(days[11]))  # 结束日期
        print("后半年任务starttime,endtime:", starttime, endtime)

    col = 0  # 指定修改的列
    statecol = 12  # 任务状态
    startcol = 10  # 实际开始
    endcol = 11  # 截止日期
    completecol=23 #完成时间
    exceptcol = 13  # 最初预计多少个工时
    losecol=14 #总消耗工时
    leftcol=15 #总剩余工时
    closereasoncol= 28  #关闭原因--已完成，已取消，空白

    mi = -1#新的excel的第几行
    for i in range(0,nrows):
        if (i == 0):
            # print("-----正在写入 " + str(i) + " 行")
            mi=mi+1
            for j in range(0,ncols):
                sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        else:
            # 获取第i行第3列数据
            # sh.cell_value(i,3)
            # print("i:",i)
            stateresult = sh.cell_value(i, statecol)
            endresult = sh.cell_value(i, endcol)
            closereasonresult = sh.cell_value(i, closereasoncol)
            completeresult = sh.cell_value(i, completecol)
            exceptcolresult= sh.cell_value(i, exceptcol)
            losecolresult=sh.cell_value(i,losecol)
            leftcolresult=sh.cell_value(i,leftcol)
            strftime = datetime.datetime.strptime(starttime, "%Y/%m/%d")
            strftime2 = datetime.datetime.strptime(endtime, "%Y/%m/%d")
            # print("状态:" + stateresult)
            # ---------写出文件到excel--------
            if ((stateresult == "已完成")):  # 判断状态是否等于已完成
                strftime4 = datetime.datetime.strptime(completeresult, "%Y/%m/%d")  # 完成时间
                if ((strftime4 > strftime) & (strftime4 < strftime2)):
                    # print("-----正在写入 " + str(i) + " 行")
                    # print("状态:"+stateresult)
                    mi = mi + 1
                    for j in range(0, ncols):
                        sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
            elif (stateresult == "已关闭"):  # 判断状态是否等于已关闭
                if (closereasonresult == "已完成"):  # 如果关闭原因是已完成
                    # completestr = int(completeresult.split('-')[0] + completeresult.split('-')[1] + completeresult.split('-')[2])
                    if("0000" not in completeresult):#完成时间不为空
                        strftime4 = datetime.datetime.strptime(completeresult, "%Y/%m/%d")  # 完成时间
                        if ((strftime4 >= strftime) & (strftime4 <= strftime2)):
                            # print("-----正在写入 " + str(i) + " 行")
                            # print("状态:" + stateresult)
                            mi = mi + 1
                            for j in range(0, ncols):
                                sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
            elif ((stateresult == "进行中")|(stateresult == "未开始")):  #进行中、未开始状态都要加入进去
                # endstr = int(endresult.split('-')[0] + endresult.split('-')[1] + endresult.split('-')[2])
                if ("0000" not in endresult):  # 截止时间不为空
                    strftime3 = datetime.datetime.strptime(endresult, "%Y/%m/%d")  # 截止时间
                    if (strftime3 <= strftime2):
                        # print("-----正在写入 " + str(i) + " 行")
                        # print("状态:" + stateresult)
                        mi = mi + 1
                        for j in range(0, ncols):
                            sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值

    book.save(allpath)#保存
    change_task_excel(allpath)#修改


def getMonthTasksConsume(isHalfYear):

    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹

    print("Folder_path:"+os.path.basename(Folder_Path))
    a1=os.path.basename(Folder_Path)
    a=a1[:9]
    print("a"+a)
    if a[4:6]=='01':
        lastnum=str(int(a[:4])-1)+str(12)+a[6:]
    else:
        d=str(int(a[4:6])-1)
        if len(d)==1:
            d='0'+d
        lastnum = a[:4] + d + "01"
    print(lastnum)
    Folder_Path2=get_last_file(path,lastnum)
    logWriteToTxt("Folder_Path2"+Folder_Path2)

    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    xlsxpath=Folder_Path+"/tasks.xls"

    # xlsxpath=r'F:\project\at-other\at_other_scripts\zentaoTest\result\20200608 120123\tasks.xls'
    xlsxpath2=Folder_Path2+"/tasks.xls"
    logWriteToTxt(xlsxpath2)
    # print(xlsxpath)
    # xlsxpath = "E:\\PythonWorkspace\\ZentaoTest\\result\\20190410 160724\\20190410160724tasks.xls"
    if (int(isHalfYear) == 0):
        allpath = Folder_Path + "/工时消耗.xls"
    # else:
    #     allpath = Folder_Path + "/半年tasks.xls"
    if (os.path.exists(allpath) == True):
        os.remove(allpath)

    bk = xlrd.open_workbook(xlsxpath)
    shxrange = range(bk.nsheets)
    try:
        sh = bk.sheet_by_name("data")
    except:
        logWriteToTxt("代码出错")
    #35列
    nrows = sh.nrows  # 获取行数
    ncols=sh.ncols# 获取列数

    bk2=xlrd.open_workbook(xlsxpath2)

    shxrange2 = range(bk2.nsheets)
    try:
        sh2 = bk2.sheet_by_name("data")
    except:
        logWriteToTxt("代码出错")
    # 35列
    nrows2 = sh2.nrows  # 获取行数
    ncols2 = sh2.ncols  # 获取列数

    book = Workbook(encoding='utf-8')
    sheet = book.add_sheet('data')  # 创建一个sheet
    # sheet=book.active
    # sheet.title='data'
    today = time.strftime("%Y/%m/%d")#今天
    day_num=int(today.split('/')[2])  #几号
    month_num=int(today.split('/')[1])  #几月
    year_num=int(today.split('/')[0]) #几年
    flag=is_leap_year(year_num)
    if(flag==True):#该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    if(int(isHalfYear)==0):
        # if (month_num == 1):
        #     if (int(day_num) < 27):#那就上一年的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(11) + "/" + str(days[10]-3))  # 开始日期
        #         endtime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(27))  # 结束日期
        #     else:#那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        # elif (month_num == 2):
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        # else:
        #     # print("day_num:",day_num)
        #     if (int(day_num) <27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-2) + "/" + str(days[month_num-3] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        if (month_num == 1):  # 就要计算到上一年的12月
            starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(1))  # 开始日期
            endtime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11]))  # 结束日期

            #此出时间为上上月30号的日期
            glastTime=time.strftime(str(year_num - 1) + "/" + str(11) + "/" + str(days[10]))
        else:  # 就要计算到上一月
            starttime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(1))  # 开始日期
            endtime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(days[month_num - 2]))  # 结束日期
            # 此出时间为上上月30号的日期
            glastTime=time.strftime(str(year_num) + "/" + str(month_num - 2) + "/" + str(days[month_num - 3]))

    # elif(int(isHalfYear)==1):
    #     starttime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(1))  # 开始日期
    #     endtime = time.strftime(str(year_num ) + "/" + str(6) + "/" + str(days[5]))  # 结束日期
    #     print("前半年任务starttime,endtime:", starttime, endtime)
    # elif (int(isHalfYear) == 2):
    #     starttime = time.strftime(str(year_num ) + "/" + str(7) + "/" + str(1))  # 开始日期
    #     endtime = time.strftime(str(year_num ) + "/" + str(12) + "/" + str(days[11]))  # 结束日期
    #     print("后半年任务starttime,endtime:", starttime, endtime)

    col = 0  # 指定修改的列
    statecol = 12  # 任务状态
    startcol = 10  # 实际开始
    endcol = 11  # 截止日期
    completecol=23 #完成时间
    exceptcol = 13  # 最初预计多少个工时
    losecol=14 #总消耗工时
    leftcol=15 #总剩余工时
    closereasoncol= 28  #关闭原因--已完成，已取消，空白
    dutyman=34 #责任人
    d1 = sh.col_values(col)
    d2 = sh.col_values(losecol)
    d3 = sh.col_values(dutyman)
    d1 = d1[1:]
    d2 = d2[1:]
    d3= d3[1:]
    list1=list_dic(d1,d3)
    dic = list_dic2(list1, d2)
    logWriteToTxt("新版本数据："+str(dic))
    print("新版本数据：" + str(dic))
    d1 = sh2.col_values(col)
    d2 = sh2.col_values(losecol)
    d3 = sh2.col_values(dutyman)
    d1 = d1[1:]
    d2 = d2[1:]
    d3= d3[1:]
    list2 = list_dic(d1, d3)
    dic2 = list_dic2(list2, d2)
    logWriteToTxt("老版本数据："+str(dic2))
    print("老版本数据："+str(dic2))
    news = {}

    for key, value in dic.items():

        if key in dic2.keys():
            # print(key)
            new = float(value) - float(dic2[key])
        else:
            new = value
        news[key] = str(new)
    print("news:" + str(news))
    logWriteToTxt("news数据：" + str(news))
    mi = -1#新的excel的第几行
    for i in range(0,nrows):
        if (i == 0):
            # print("-----正在写入 " + str(i) + " 行")
            mi=mi+1
            for j in range(0,ncols):
                sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值

            sheet.write(mi,ncols,label='当月消耗')
        else:
            # 获取第i行第3列数据
            # sh.cell_value(i,3)
            # print("i:",i)
            stateresult = sh.cell_value(i, statecol)
            endresult = sh.cell_value(i, endcol)
            closereasonresult = sh.cell_value(i, closereasoncol)
            completeresult = sh.cell_value(i, completecol)
            exceptcolresult= sh.cell_value(i, exceptcol)
            losecolresult=sh.cell_value(i,losecol)
            leftcolresult=sh.cell_value(i,leftcol)
            strftime = datetime.datetime.strptime(starttime, "%Y/%m/%d")
            strftime2 = datetime.datetime.strptime(endtime, "%Y/%m/%d")
            glastTime2=datetime.datetime.strptime(glastTime,"%Y/%m/%d")
            # print("状态:" + stateresult)
            # ---------写出文件到excel--------

            if ((stateresult == "已完成")):  # 判断状态是否等于已完成
                strftime4 = datetime.datetime.strptime(completeresult, "%Y/%m/%d")  # 完成时间
                # print("strftime4:"+str(strftime4))

                # if ((strftime4 > strftime) & (strftime4 < strftime2)):
                if strftime4>glastTime2:

                # else:
                    # print("-----正在写入 " + str(i) + " 行")
                    # print("状态:"+stateresult)
                    mi = mi + 1
                    for j in range(0, ncols):
                        sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                    # print("cell0"+sh.cell_value(i,col))
                    # keyId=sh.cell_value(i,col)
                    keyId1 = sh.cell_value(i, col)
                    keyId2 = sh.cell_value(i, dutyman)
                    keyId = str([keyId1, keyId2])
                    sheet.write(mi,ncols,label=news[keyId])
            elif (stateresult == "已关闭"):  # 判断状态是否等于已关闭
                if (closereasonresult == "已完成"):  # 如果关闭原因是已完成
                    # completestr = int(completeresult.split('-')[0] + completeresult.split('-')[1] + completeresult.split('-')[2])
                    if("0000" not in completeresult):#完成时间不为空
                        strftime4 = datetime.datetime.strptime(completeresult, "%Y/%m/%d")  # 完成时间
                        # print("strftime4:" + str(strftime4))
                        # if ((strftime4 >= strftime) & (strftime4 <= strftime2)):
                        if strftime4 > glastTime2:
                            # print("-----正在写入 " + str(i) + " 行")
                            # print("状态:" + stateresult)
                            mi = mi + 1
                            for j in range(0, ncols):
                                sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                            # print("cell0" + sh.cell_value(i, col))
                            # keyId = sh.cell_value(i, col)
                            keyId1 = sh.cell_value(i, col)
                            keyId2 = sh.cell_value(i, dutyman)
                            keyId = str([keyId1, keyId2])
                            sheet.write(mi, ncols, label=news[keyId])
            elif ((stateresult == "进行中")|(stateresult == "未开始")):  #进行中、未开始状态都要加入进去
                # endstr = int(endresult.split('-')[0] + endresult.split('-')[1] + endresult.split('-')[2])
                if ("0000" not in endresult):  # 截止时间不为空
                    strftime3 = datetime.datetime.strptime(endresult, "%Y/%m/%d")  # 截止时间
                    # if (strftime3 <= strftime2):
                        # print("-----正在写入 " + str(i) + " 行")
                        # print("状态:" + stateresult)
                    mi = mi + 1
                    for j in range(0, ncols):
                        sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                    # print("cell0" + sh.cell_value(i, col))
                    # keyId = sh.cell_value(i, col)
                    keyId1 = sh.cell_value(i, col)
                    keyId2 = sh.cell_value(i, dutyman)
                    keyId = str([keyId1, keyId2])
                    sheet.write(mi, ncols, label=news[keyId])

    book.save(allpath)#保存
    logWriteToTxt("保存好文件"+allpath)
    change_task_excel(allpath)#修改


def getMonthTasksConsume2(isHalfYear):

    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹

    print("Folder_path:"+os.path.basename(Folder_Path))
    a1=os.path.basename(Folder_Path)
    a=a1[:9]
    print("a:"+a)
    lastnum=a[:6]+"01"
    print(lastnum)
    Folder_Path2=get_last_file(path,lastnum)
    print("Folder_Path2"+Folder_Path2)

    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    xlsxpath=Folder_Path+"/tasks.xls"

    # xlsxpath=r'F:\project\at-other\at_other_scripts\zentaoTest\result\20200608 120123\tasks.xls'
    xlsxpath2=Folder_Path2+"/tasks.xls"
    print(xlsxpath2)
    # print(xlsxpath)
    # xlsxpath = "E:\\PythonWorkspace\\ZentaoTest\\result\\20190410 160724\\20190410160724tasks.xls"
    if (int(isHalfYear) == 0):
        allpath = Folder_Path + "/工时消耗.xls"
    # else:
    #     allpath = Folder_Path + "/半年tasks.xls"
    if (os.path.exists(allpath) == True):
        os.remove(allpath)

    bk = xlrd.open_workbook(xlsxpath)
    shxrange = range(bk.nsheets)
    try:
        sh = bk.sheet_by_name("data")
    except:
        print("代码出错")
    #35列
    nrows = sh.nrows  # 获取行数
    ncols=sh.ncols# 获取列数

    bk2=xlrd.open_workbook(xlsxpath2)

    shxrange2 = range(bk2.nsheets)
    try:
        sh2 = bk2.sheet_by_name("data")
    except:
        print("代码出错")
    # 35列
    nrows2 = sh2.nrows  # 获取行数
    ncols2 = sh2.ncols  # 获取列数

    book = Workbook(encoding='utf-8')
    sheet = book.add_sheet('data')  # 创建一个sheet
    # sheet=book.active
    # sheet.title='data'
    today = time.strftime("%Y/%m/%d")#今天
    day_num=int(today.split('/')[2])  #几号
    month_num=int(today.split('/')[1])  #几月
    year_num=int(today.split('/')[0]) #几年
    flag=is_leap_year(year_num)
    if(flag==True):#该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    if(int(isHalfYear)==0):
        # if (month_num == 1):
        #     if (int(day_num) < 27):#那就上一年的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(11) + "/" + str(days[10]-3))  # 开始日期
        #         endtime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(27))  # 结束日期
        #     else:#那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        # elif (month_num == 2):
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        # else:
        #     # print("day_num:",day_num)
        #     if (int(day_num) <27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-2) + "/" + str(days[month_num-3] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        # if (month_num == 1):  # 就要计算到上一年的12月
            # starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(1))  # 开始日期
            # endtime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11]))  # 结束日期

            #此出时间为上上月30号的日期
        glastTime=time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(1))
        # else:  # 就要计算到上一月
            # starttime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(1))  # 开始日期
            # endtime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(days[month_num - 2]))  # 结束日期
            # 此出时间为上上月30号的日期
            # glastTime=time.strftime(str(year_num) + "/" + str(month_num - 2) + "/" + str(days[month_num - 3]))
        print("一月任务,glastTime:",glastTime)
    # elif(int(isHalfYear)==1):
    #     starttime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(1))  # 开始日期
    #     endtime = time.strftime(str(year_num ) + "/" + str(6) + "/" + str(days[5]))  # 结束日期
    #     print("前半年任务starttime,endtime:", starttime, endtime)
    # elif (int(isHalfYear) == 2):
    #     starttime = time.strftime(str(year_num ) + "/" + str(7) + "/" + str(1))  # 开始日期
    #     endtime = time.strftime(str(year_num ) + "/" + str(12) + "/" + str(days[11]))  # 结束日期
    #     print("后半年任务starttime,endtime:", starttime, endtime)

    col = 0  # 指定修改的列
    statecol = 12  # 任务状态
    startcol = 10  # 实际开始
    endcol = 11  # 截止日期
    completecol=23 #完成时间
    exceptcol = 13  # 最初预计多少个工时
    losecol=14 #总消耗工时
    leftcol=15 #总剩余工时
    closereasoncol= 28  #关闭原因--已完成，已取消，空白
    dutyman=34 #责任人

    d1 = sh.col_values(col)
    d2 = sh.col_values(losecol)
    d3 = sh.col_values(dutyman)
    d1 = d1[1:]
    d2 = d2[1:]
    d3=  d3[1:]
    list1 = list_dic(d1, d3)
    dic=list_dic2(list1,d2)
    logWriteToTxt("新版本数据："+str(dic))
    print("新版本数据："+str(dic))
    d1 = sh2.col_values(col)
    d2 = sh2.col_values(losecol)
    d3 = sh.col_values(dutyman)
    d1 = d1[1:]
    d2 = d2[1:]
    d3= d3[1:]

    list2 = list_dic(d1, d3)
    dic2=list_dic2(list2,d2)
    logWriteToTxt("老版本数据："+str(dic2))
    print("老版本数据：" + str(dic2))
    news = {}
    for key, value in dic.items():

        if key in dic2.keys():
            # print(key)
            new = float(value) - float(dic2[key])
        else:
            new = value
        news[key] = str(new)
    print("news:"+str(news))
    logWriteToTxt("news数据：" + str(news))
    mi = -1#新的excel的第几行
    for i in range(0,nrows):
        if (i == 0):
            # print("-----正在写入 " + str(i) + " 行")
            mi=mi+1
            for j in range(0,ncols):
                sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值

            sheet.write(mi,ncols,label='当月消耗')
        else:
            # 获取第i行第3列数据
            # sh.cell_value(i,3)
            # print("i:",i)
            stateresult = sh.cell_value(i, statecol)
            endresult = sh.cell_value(i, endcol)
            closereasonresult = sh.cell_value(i, closereasoncol)
            completeresult = sh.cell_value(i, completecol)
            exceptcolresult= sh.cell_value(i, exceptcol)
            losecolresult=sh.cell_value(i,losecol)
            leftcolresult=sh.cell_value(i,leftcol)
            # dutymanresult=sh.col_value(i,dutyman)
            # strftime = datetime.datetime.strptime(starttime, "%Y/%m/%d")
            # strftime2 = datetime.datetime.strptime(endtime, "%Y/%m/%d")
            glastTime2=datetime.datetime.strptime(glastTime,"%Y/%m/%d")
            # print("状态:" + stateresult)
            # ---------写出文件到excel--------

            if ((stateresult == "已完成")):  # 判断状态是否等于已完成
                strftime4 = datetime.datetime.strptime(completeresult, "%Y/%m/%d")  # 完成时间
                # print("strftime4:"+str(strftime4))

                # if ((strftime4 > strftime) & (strftime4 < strftime2)):
                if strftime4>=glastTime2:

                # else:
                    # print("-----正在写入 " + str(i) + " 行")
                    # print("状态:"+stateresult)
                    mi = mi + 1
                    for j in range(0, ncols):
                        sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                    # print("cell0"+sh.cell_value(i,col))
                    keyId1=sh.cell_value(i,col)
                    keyId2=sh.cell_value(i,dutyman)
                    keyId=str([keyId1,keyId2])
                    # print("keyId:"+keyId)
                    sheet.write(mi,ncols,label=news[keyId])
            elif (stateresult == "已关闭"):  # 判断状态是否等于已关闭
                if (closereasonresult == "已完成"):  # 如果关闭原因是已完成
                    # completestr = int(completeresult.split('-')[0] + completeresult.split('-')[1] + completeresult.split('-')[2])
                    if("0000" not in completeresult):#完成时间不为空
                        strftime4 = datetime.datetime.strptime(completeresult, "%Y/%m/%d")  # 完成时间
                        # print("strftime4:" + str(strftime4))
                        # if ((strftime4 >= strftime) & (strftime4 <= strftime2)):
                        if strftime4 >= glastTime2:
                            # print("-----正在写入 " + str(i) + " 行")
                            # print("状态:" + stateresult)
                            mi = mi + 1
                            for j in range(0, ncols):
                                sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                            # print("cell0" + sh.cell_value(i, col))
                            # keyId = sh.cell_value(i, col)
                            keyId1 = sh.cell_value(i, col)
                            keyId2 = sh.cell_value(i, dutyman)
                            keyId = str([keyId1, keyId2])
                            # print("keyId:" + keyId)
                            sheet.write(mi, ncols, label=news[keyId])
            elif ((stateresult == "进行中")|(stateresult == "未开始")):  #进行中、未开始状态都要加入进去
                # endstr = int(endresult.split('-')[0] + endresult.split('-')[1] + endresult.split('-')[2])
                if ("0000" not in endresult):  # 截止时间不为空
                    strftime3 = datetime.datetime.strptime(endresult, "%Y/%m/%d")  # 截止时间
                    # if (strftime3 <= strftime2):
                        # print("-----正在写入 " + str(i) + " 行")
                        # print("状态:" + stateresult)
                    mi = mi + 1
                    for j in range(0, ncols):
                        sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                    # print("cell0" + sh.cell_value(i, col))
                    # keyId = sh.cell_value(i, col)
                    keyId1 = sh.cell_value(i, col)
                    keyId2 = sh.cell_value(i, dutyman)
                    keyId = str([keyId1, keyId2])
                    # print("keyId:" + keyId)
                    sheet.write(mi, ncols, label=news[keyId])

    book.save(allpath)#保存
    logWriteToTxt("保存好文件" + allpath)
    change_task_excel(allpath)#修改
















#得到上一月的关闭bug的excel
def getMonthBugsCloseExcel(session,isHalfYear):
    departNamelist = _getDepartment(session)
    departs = departNamelist[0]  # 部门
    names = departNamelist[1]  # 人

    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    xlsxpath=Folder_Path+"/bugs.xls"
    # print(xlsxpath)
    # xlsxpath = "E:\\PythonWorkspace\\ZentaoTest\\result\\20190410 160724\\20190410160724tasks.xls"
    # 先创建一个excel文件，并且添加相对应的sheet
    if (int(isHalfYear) == 0):
        allpath = Folder_Path + "/一月关闭bugs.xls"
    else:
        allpath = Folder_Path + "/半年关闭bugs.xls"

    if (os.path.exists(allpath) == True):
        os.remove(allpath)

    bk = xlrd.open_workbook(xlsxpath)
    shxrange = range(bk.nsheets)
    try:
        sh = bk.sheet_by_name("data")
    except:
        print("代码出错")
    nrows = sh.nrows  # 获取行数
    ncols = sh.ncols  # 获取列数
    book = Workbook(encoding='utf-8')
    sheet = book.add_sheet('data')  # 创建一个sheet

    today = time.strftime("%Y/%m/%d")  # 今天
    day_num = int(today.split('/')[2])  # 几号
    month_num = int(today.split('/')[1])  # 几月
    year_num = int(today.split('/')[0])  # 几年
    flag = is_leap_year(year_num)
    if (flag == True):  # 该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    if(int(isHalfYear)==0):
        # if (month_num == 1):
        #     if(int(day_num)<27):#那就上一年的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(11) + "/" + str(days[10]-3))  # 开始日期
        #         endtime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(27))  # 结束日期
        #     else:#那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        # elif (month_num == 2):
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        # else:
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-2) + "/" + str(days[month_num-3] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        if (month_num == 1):  # 就要计算到上一年的12月
            starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(1))  # 开始日期
            endtime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11]))  # 结束日期
        else:  # 就要计算到上一月
            starttime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(1))  # 开始日期
            endtime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(days[month_num - 2]))  # 结束日期
        print("一月关闭bugstarttime,endtime:", starttime, endtime)
    elif (int(isHalfYear) == 1):
        starttime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num) + "/" + str(6) + "/" + str(days[5]))  # 结束日期
        print("前半年关闭bugstarttime,endtime:", starttime, endtime)
    elif (int(isHalfYear) == 2):
        starttime = time.strftime(str(year_num) + "/" + str(7) + "/" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num) + "/" + str(12) + "/" + str(days[11]))  # 结束日期
        print("后半年关闭bugstarttime,endtime:", starttime, endtime)

    closetimecol=29  #关闭日期
    closecol = 28 # 由谁关闭
    responcol = 38  # 负责人
    respondecol = 37  # 负责部门

    mi = -1  # 新的excel的第几行
    for i in range(0, nrows):
        if (i == 0):
            # print("-----正在写入 " + str(i) + " 行")
            mi = mi + 1
            for j in range(0, ncols):
                if((j!=43)&(j!=41)&(j!=42)&(j!=40)&(j!=39)):
                    sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        else:
            # 获取第i行第3列数据
            # sh.cell_value(i,3)
            # print("i:",i)
            closetimecolresult = sh.cell_value(i, closetimecol)
            strftime = datetime.datetime.strptime(starttime, "%Y/%m/%d")
            strftime2 = datetime.datetime.strptime(endtime, "%Y/%m/%d")
            if ("0000" not in closetimecolresult):  # 关闭时间不为空
                strftime3 = datetime.datetime.strptime(closetimecolresult, "%Y/%m/%d")  # 关闭时间
                if ((strftime3 >= strftime) & (strftime3 <= strftime2)):
                    # print("-----正在写入 " + str(i) + " 行")
                    # print("状态:" + stateresult)
                    mi = mi + 1
                    for j in range(0, ncols):
                        if ((j != 43) & (j != 41) & (j != 42)& (j != 40)& (j != 39)):
                            sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
    book.save(allpath)

    rb = xlrd.open_workbook(allpath, formatting_info=True)  # 打开xls文件
    ro = rb.sheets()[0]  # 读取表单0
    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws = wb.get_sheet(0)  # 获取表单0
    for i in range(1, ro.nrows):  # 循环所有的行
        closename = ro.cell(i, closecol).value
        ind = -1
        for na in names:
            ind = ind + 1
            for name in na:
                if (closename != None and closename != ""):
                    if (closename in name):
                        respoDepart = departs[ind]
                        ws.write(i, respondecol, respoDepart)
        ws.write(i, responcol, closename)
    wb.save(allpath)





#得到上一月的创建bug的excel
def getMonthBugsCreateExcel(session,isHalfYear):
    departNamelist = _getDepartment(session)
    departs = departNamelist[0]  # 部门
    names = departNamelist[1]  # 人

    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    xlsxpath=Folder_Path+"/bugs.xls"
    # print(xlsxpath)
    # xlsxpath = "E:\\PythonWorkspace\\ZentaoTest\\result\\20190410 160724\\20190410160724tasks.xls"
    # 先创建一个excel文件，并且添加相对应的sheet
    if (int(isHalfYear) == 0):
        allpath = Folder_Path + "/一月创建bugs.xls"
    else:
        allpath = Folder_Path + "/半年创建bugs.xls"

    if (os.path.exists(allpath) == True):
        os.remove(allpath)

    bk = xlrd.open_workbook(xlsxpath)
    shxrange = range(bk.nsheets)
    try:
        sh = bk.sheet_by_name("data")
    except:
        print("代码出错")
    nrows = sh.nrows  # 获取行数
    ncols = sh.ncols  # 获取列数
    book = Workbook(encoding='utf-8')
    sheet = book.add_sheet('data')  # 创建一个sheet
    today = time.strftime("%Y/%m/%d")  # 今天
    day_num = int(today.split('/')[2])  # 几号
    month_num = int(today.split('/')[1])  # 几月
    year_num = int(today.split('/')[0])  # 几年
    flag = is_leap_year(year_num)
    if (flag == True):  # 该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    if(int(isHalfYear)==0):
        # if (month_num == 1):
        #     if(int(day_num)<27):#那就上一年的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(11) + "/" + str(days[10]-3))  # 开始日期
        #         endtime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(27))  # 结束日期
        #     else:#那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        # elif (month_num == 2):
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        # else:
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-2) + "/" + str(days[month_num-3] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        if (month_num == 1):  # 就要计算到上一年的12月
            starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(1))  # 开始日期
            endtime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11]))  # 结束日期
        else:  # 就要计算到上一月
            starttime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(1))  # 开始日期
            endtime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(days[month_num - 2]))  # 结束日期
        print("一月创建bugstarttime,endtime:", starttime, endtime)
    elif (int(isHalfYear) == 1):
        starttime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num) + "/" + str(6) + "/" + str(days[5]))  # 结束日期
        print("前半年创建bugstarttime,endtime:", starttime, endtime)
    elif (int(isHalfYear) == 2):
        starttime = time.strftime(str(year_num) + "/" + str(7) + "/" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num) + "/" + str(12) + "/" + str(days[11]))  # 结束日期
        print("后半年创建bugstarttime,endtime:", starttime, endtime)

    createtimecol=20  #创建时间
    createcol=19  #由谁创建
    responcol = 38  # 负责人
    respondecol = 37  # 负责部门

    mi = -1  # 新的excel的第几行
    for i in range(0, nrows):
        if (i == 0):
            # print("-----正在写入 " + str(i) + " 行")
            mi = mi + 1
            for j in range(0, ncols):
                if((j!=43)&(j!=41)&(j!=42)&(j!=40)&(j!=39)):
                    sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        else:
            # 获取第i行第3列数据
            # sh.cell_value(i,3)
            # print("i:",i)
            createtimeresult = sh.cell_value(i, createtimecol)
            strftime = datetime.datetime.strptime(starttime, "%Y/%m/%d")
            strftime2 = datetime.datetime.strptime(endtime, "%Y/%m/%d")
            if ("0000" not in createtimeresult):  # 创建时间不为空
                strftime3 = datetime.datetime.strptime(createtimeresult, "%Y/%m/%d")  # 创建时间
                if ((strftime3 >= strftime) & (strftime3 <= strftime2)):
                    # print("-----正在写入 " + str(i) + " 行")
                    # print("状态:" + stateresult)
                    mi = mi + 1
                    for j in range(0, ncols):
                        if ((j != 43) & (j != 41) & (j != 42)& (j != 40)& (j != 39)):
                            sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
    book.save(allpath)

    rb = xlrd.open_workbook(allpath, formatting_info=True)  # 打开xls文件
    ro = rb.sheets()[0]  # 读取表单0
    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws = wb.get_sheet(0)  # 获取表单0
    for i in range(1, ro.nrows):  # 循环所有的行
        createname=ro.cell(i, createcol).value
        ind = -1
        for na in names:
            ind = ind + 1
            for name in na:
                if(createname!=None and createname!=""):
                    if (createname in name):
                        respoDepart = departs[ind]
                        ws.write(i, respondecol, respoDepart)
        ws.write(i, responcol, createname)
    wb.save(allpath)








#得到上一月的及时解决性bug的excel
def getMonthBugsSolveExcel(isHalfYear):
    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    xlsxpath=Folder_Path+"/bugs.xls"
    # print(xlsxpath)
    # xlsxpath = "E:\\PythonWorkspace\\ZentaoTest\\result\\20190410 160724\\20190410160724tasks.xls"

    if (int(isHalfYear) == 0):
        allpath = Folder_Path + "/一月解决bugs.xls"
    else:
        allpath = Folder_Path + "/半年解决bugs.xls"
    if (os.path.exists(allpath) == True):
        os.remove(allpath)
    bk = xlrd.open_workbook(xlsxpath)
    shxrange = range(bk.nsheets)
    try:
        sh = bk.sheet_by_name("data")
    except:
        print("代码出错")
    nrows = sh.nrows  # 获取行数
    ncols = sh.ncols  # 获取列数
    book = Workbook(encoding='utf-8')
    sheet = book.add_sheet('data')  # 创建一个sheet

    today = time.strftime("%Y/%m/%d")  # 今天
    day_num = int(today.split('/')[2])  # 几号
    month_num = int(today.split('/')[1])  # 几月
    year_num = int(today.split('/')[0])  # 几年
    flag = is_leap_year(year_num)
    if (flag == True):  # 该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    if(int(isHalfYear)==0):
        # if (month_num == 1):
        #     if (int(day_num) < 27):#那就上一年的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(11) + "/" + str(days[10]-3))  # 开始日期
        #         endtime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(27))  # 结束日期
        #     else:#那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        # elif (month_num == 2):
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        # else:
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-2) + "/" + str(days[month_num-3] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期

        if (month_num == 1):  # 就要计算到上一年的12月
            starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(1))  # 开始日期
            endtime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11]))  # 结束日期
        else:  # 就要计算到上一月
            starttime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(1))  # 开始日期
            endtime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(days[month_num - 2]))  # 结束日期
    elif (int(isHalfYear) == 1):
        starttime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num) + "/" + str(6) + "/" + str(days[5]))  # 结束日期
    elif (int(isHalfYear) == 2):

        starttime = time.strftime(str(year_num) + "/" + str(7) + "/" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num) + "/" + str(12) + "/" + str(days[11]))  # 结束日期


    print("一月解决bugstarttime,endtime:", starttime, endtime)
    statecol = 14  # Bug状态
    solvecol=27  #解决时间
    resolvecol=39  #要求解决时间
    closecol=29 #关闭时间
    mi = -1  # 新的excel的第几行
    for i in range(0, nrows):
        if (i == 0):
            # print("-----正在写入 " + str(i) + " 行")
            mi = mi + 1
            for j in range(0, ncols):
                if((j!=43)&(j!=41)&(j!=42)):
                    sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        else:
            # 获取第i行第3列数据
            # sh.cell_value(i,3)
            # print("i:",i)
            stateresult = sh.cell_value(i, statecol)
            solveresult = sh.cell_value(i, solvecol)
            resolveresult= sh.cell_value(i, resolvecol)
            closeresult=sh.cell_value(i, closecol)
            strftime = datetime.datetime.strptime(starttime, "%Y/%m/%d")
            strftime2 = datetime.datetime.strptime(endtime, "%Y/%m/%d")
            if ((stateresult == "已关闭")):
                if("0000" not in solveresult):#解决时间不为空
                    strftime3 = datetime.datetime.strptime(solveresult, "%Y/%m/%d")  # 解决时间
                else:
                    strftime3 = datetime.datetime.strptime(closeresult, "%Y/%m/%d")  # 关闭时间
                if ((strftime3 >= strftime) & (strftime3 <= strftime2)):
                    # print("-----正在写入 " + str(i) + " 行")
                    # print("状态:" + stateresult)
                    mi = mi + 1
                    for j in range(0, ncols):
                        if ((j != 43) & (j != 41) & (j != 42)):
                            sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
            elif((stateresult == "已解决")):
                if ("0000" not in solveresult):  # 解决时间不为空
                    strftime3 = datetime.datetime.strptime(solveresult, "%Y/%m/%d")  # 解决时间
                    if ((strftime3 >= strftime) & (strftime3 <= strftime2)):
                        # print("-----正在写入 " + str(i) + " 行")
                        # print("状态:" + stateresult)
                        mi = mi + 1
                        for j in range(0, ncols):
                            if ((j != 43) & (j != 41) & (j != 42)):
                                sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
            elif(stateresult=="激活"):
                strftime3 = datetime.datetime.strptime(resolveresult, "%Y/%m/%d")  # 要求解决时间
                # if ((strftime3 >= strftime) & (strftime3 <= strftime2)):
                if (strftime3<=strftime2):
                    # print("-----正在写入 " + str(i) + " 行")
                    # print("状态:" + stateresult)
                    mi = mi + 1
                    for j in range(0, ncols):
                        if ((j != 43) & (j != 41) &(j != 42)):
                            sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
    book.save(allpath)



#上个月已完成及截止上月底今年要求完成而未完成的测试单
def getMonthTestExcel(isHalfYear):
    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    testpath1=Folder_Path+"/版本测试单.xls"
    testpath2 = Folder_Path + "/用例测试单.xls"
    paths = []
    if (os.path.exists(testpath1) == True):
        paths.append(testpath1)
    if (os.path.exists(testpath2) == True):
        paths.append(testpath2)

    today = time.strftime("%Y-%m-%d")#今天
    day_num=int(today.split('-')[2])  #几号
    month_num=int(today.split('-')[1])  #几月
    year_num=int(today.split('-')[0]) #几年
    tt=time.strftime(str(year_num) + "-" + str(1) + "-" + str(1))
    flag=is_leap_year(year_num)
    start=''
    end=''
    if(flag==True):#该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    if(int(isHalfYear)==0):
        # if (month_num == 1):
        #     if (int(day_num) < 27):#那就上一年的27号到前30天
        #         start = time.strftime(str(year_num-1) + "-" + str(11) + "-" + str(days[10]-3))  # 开始日期
        #         end = time.strftime(str(year_num-1) + "-" + str(12) + "-" + str(27))  # 结束日期
        #     else:#那就该月的27号到前30天
        #         start = time.strftime(str(year_num - 1) + "-" + str(12) + "-" + str(days[11] - 3))  # 开始日期
        #         end = time.strftime(str(year_num) + "-" + str(1) + "-" + str(27))  # 结束日期
        # elif (month_num == 2):
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         start = time.strftime(str(year_num-1) + "-" + str(12) + "-" + str(days[11] - 3))  # 开始日期
        #         end = time.strftime(str(year_num) + "-" + str(1) + "-" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         start = time.strftime(str(year_num) + "-" + str(month_num-1) + "-" + str(days[month_num-2] - 3))  # 开始日期
        #         end = time.strftime(str(year_num) + "-" + str(month_num) + "-" + str(27))  # 结束日期
        # else:
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         start = time.strftime(str(year_num) + "-" + str(month_num-2) + "-" + str(days[month_num-3] - 3))  # 开始日期
        #         end = time.strftime(str(year_num) + "-" + str(month_num-1) + "-" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         start = time.strftime(str(year_num) + "-" + str(month_num-1) + "-" + str(days[month_num-2] - 3))  # 开始日期
        #         end = time.strftime(str(year_num) + "-" + str(month_num) + "-" + str(27))  # 结束日期
        if (month_num == 1):  # 就要计算到上一年的12月
            start = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(1))  # 开始日期
            end = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11]))  # 结束日期
        else:  # 就要计算到上一月
            start = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(1))  # 开始日期
            end = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(days[month_num - 2]))  # 结束日期
        print("一月版本starttime,endtime:", start, end)
    elif (int(isHalfYear) == 1):
        start = time.strftime(str(year_num) + "-" + str(1) + "-" + str(1))  # 开始日期
        end = time.strftime(str(year_num) + "-" + str(6) + "-" + str(days[5]))  # 结束日期
    elif (int(isHalfYear) == 2):
        start = time.strftime(str(year_num) + "-" + str(7) + "-" + str(1))  # 开始日期
        end = time.strftime(str(year_num) + "-" + str(12) + "-" + str(days[11]))  # 结束日期
    # 先创建一个excel文件，并且添加相对应的sheet
    if(int(isHalfYear)==0):
        allpath = Folder_Path + "/上月版本测试完成情况.xls"
    else:
        allpath = Folder_Path + "/半年版本测试完成情况.xls"

    if (os.path.exists(allpath) == True):
        os.remove(allpath)

    work_book = xlwt.Workbook(encoding='utf-8')
    work_book.add_sheet('版本测试单')  # 第一个sheet名称
    work_book.save(allpath)

    for m in range(1, len(paths)):  # 添加sheet-名称
        # print("m:",m)
        rb = xlrd.open_workbook(allpath, formatting_info=True)
        wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
        if (m == 1):
            wb.add_sheet('用例测试单')
            wb.save(allpath)
    rb = xlrd.open_workbook(allpath, formatting_info=True)
    sheetlen = len(rb.sheets())
    for m in range(0, sheetlen):

        # print("m:",m)
        rb = xlrd.open_workbook(allpath, formatting_info=True)  # 打开 上月版本测试完成情况.xls文件，把其他文件的值复制到该文件
        ro = rb.sheets()[m]  # 读取表单m
        wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
        ws = wb.get_sheet(m)  # 获取表单m

        bk = xlrd.open_workbook(paths[m])  # 打开该文件，复制到 上月版本测试完成情况.xls文件
        try:
            sh = bk.sheet_by_name("data")
        except:
            print("代码出错")
        # 35列
        nrows = sh.nrows  # 获取行数
        ncols = sh.ncols  # 获取列数
        # print("行，列:",nrows,ncols)
        dateFormat = xlwt.XFStyle()
        dateFormat.num_format_str = 'yyyy-mm-dd'

        mi = -1  # 新的excel的第几行
        for i in range(0, nrows):
            if (i == 0):
                # print("-----正在写入 " + str(i) + " 行")
                mi = mi + 1
                for j in range(0, ncols):
                    ws.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
            else:
                starttime=sh.cell_value(i,7)
                endtime=sh.cell_value(i, 8)
                state=sh.cell_value(i,9)
                strftime = datetime.datetime.strptime(starttime, "%Y-%m-%d")#版本开始时间
                strftime2 = datetime.datetime.strptime(endtime, "%Y-%m-%d") #版本结束时间
                strftime3 = datetime.datetime.strptime(start, "%Y/%m/%d")
                strftime4 = datetime.datetime.strptime(end, "%Y/%m/%d")
                strftime5 = datetime.datetime.strptime(tt, "%Y-%m-%d")
                if (state == "已完成"):
                    if ((strftime2 <= strftime4) & (strftime2 >= strftime3)):
                        mi = mi + 1
                        for j in range(0, ncols):
                            if ((j == 7) | (j == 8)):
                                # if ((len(shvalue.split('-')) == 3) & (len(shvalue) < 12)):  # 把str转换为excel格式里的时间格式
                                #     # print((shvalue.split('/')[0]).isdigit(),(shvalue.split('/')[1]).isdigit())
                                #     if ((sh.cell_value(i, j).split('-')[2]).isdigit() == True & (
                                #             sh.cell_value(i, j).split('-')[1]).isdigit() == True & (
                                #             sh.cell_value(i, j).split('-')[0]).isdigit() == True):
                                year = int(sh.cell_value(i, j).split('-')[0])
                                month = int(sh.cell_value(i, j).split('-')[1])
                                day = int(sh.cell_value(i, j).split('-')[2])
                                # print(month, year, day)
                                ws.write(mi, j, dt.date(year, month, day), dateFormat)
                            else:
                                ws.write(mi, j, sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                else:
                    if ((strftime2 <= strftime4) & (strftime5 <= strftime2)):
                        mi = mi + 1
                        for j in range(0, ncols):
                            if ((j == 7) | (j == 8)):
                                # if ((len(shvalue.split('-')) == 3) & (len(shvalue) < 12)):  # 把str转换为excel格式里的时间格式
                                #     # print((shvalue.split('/')[0]).isdigit(),(shvalue.split('/')[1]).isdigit())
                                #     if ((sh.cell_value(i, j).split('-')[2]).isdigit() == True & (
                                #             sh.cell_value(i, j).split('-')[1]).isdigit() == True & (
                                #             sh.cell_value(i, j).split('-')[0]).isdigit() == True):
                                year = int(sh.cell_value(i, j).split('-')[0])
                                month = int(sh.cell_value(i, j).split('-')[1])
                                day = int(sh.cell_value(i, j).split('-')[2])
                                # print(month, year, day)
                                ws.write(mi, j, dt.date(year, month, day), dateFormat)
                            else:
                                ws.write(mi, j, sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                # if(m==0):#第一个excel文件
                # else:#第二个excel文件
                #     if (state == "已完成"):
                #         if((strftime2<=strftime4)&(strftime2>=strftime3)):
                #             mi=mi+1
                #             for j in range(0, ncols):
                #                 if ((j == 7) | (j == 8)):
                #                 # if ((len(shvalue.split('-')) == 3) & (len(shvalue) < 12)):  # 把str转换为excel格式里的时间格式
                #                 #     # print((shvalue.split('/')[0]).isdigit(),(shvalue.split('/')[1]).isdigit())
                #                 #     if ((sh.cell_value(i, j).split('-')[2]).isdigit() == True & (
                #                 #             sh.cell_value(i, j).split('-')[1]).isdigit() == True & (
                #                 #             sh.cell_value(i, j).split('-')[0]).isdigit() == True):
                #                         year = int(sh.cell_value(i, j).split('-')[0])
                #                         month = int(sh.cell_value(i, j).split('-')[1])
                #                         day = int(sh.cell_value(i, j).split('-')[2])
                #                         # print(month, year, day)
                #                         ws.write(mi, j, dt.date(year, month, day), dateFormat)
                #                 else:
                #                     ws.write(mi, j, sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
                #     elif((state=="被阻塞") |(state=="未开始")):
                #         if ((strftime2 <= strftime4)&(strftime5<=strftime2)):
                #             mi = mi + 1
                #             for j in range(0, ncols):
                #                 if ((j == 7) | (j == 8)):
                #                     # if ((len(shvalue.split('-')) == 3) & (len(shvalue) < 12)):  # 把str转换为excel格式里的时间格式
                #                     #     # print((shvalue.split('/')[0]).isdigit(),(shvalue.split('/')[1]).isdigit())
                #                     #     if ((sh.cell_value(i, j).split('-')[2]).isdigit() == True & (
                #                     #             sh.cell_value(i, j).split('-')[1]).isdigit() == True & (
                #                     #             sh.cell_value(i, j).split('-')[0]).isdigit() == True):
                #                     year = int(sh.cell_value(i, j).split('-')[0])
                #                     month = int(sh.cell_value(i, j).split('-')[1])
                #                     day = int(sh.cell_value(i, j).split('-')[2])
                #                     # print(month, year, day)
                #                     ws.write(mi, j, dt.date(year, month, day), dateFormat)
                #                 else:
                #                     ws.write(mi, j, sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        wb.save(allpath)

#得到上一月的回归测试的bug的excel
def getMonthBugsReturnExcel(session,isHalfYear):

    departNamelist = _getDepartment(session)
    departs = departNamelist[0]  # 部门
    names = departNamelist[1]  # 人

    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    xlsxpath=Folder_Path+"/bugs.xls"
    # print(xlsxpath)
    # xlsxpath = "E:\\PythonWorkspace\\ZentaoTest\\result\\20190410 160724\\20190410160724tasks.xls"
    if (int(isHalfYear) == 0):
        allpath = Folder_Path + "/一月回归bugs.xls"
    else:
        allpath = Folder_Path + "/半年回归bugs.xls"
    if (os.path.exists(allpath) == True):
        os.remove(allpath)

    bk = xlrd.open_workbook(xlsxpath)
    shxrange = range(bk.nsheets)
    try:
        sh = bk.sheet_by_name("data")
    except:
        print("代码出错")
    nrows = sh.nrows  # 获取行数
    ncols = sh.ncols  # 获取列数
    book = Workbook(encoding='utf-8')
    sheet = book.add_sheet('data')  # 创建一个sheet

    today = time.strftime("%Y/%m/%d")  # 今天
    day_num = int(today.split('/')[2])  # 几号
    month_num = int(today.split('/')[1])  # 几月
    year_num = int(today.split('/')[0])  # 几年
    flag = is_leap_year(year_num)
    if (flag == True):  # 该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    if(int(isHalfYear)==0):
        # if (month_num == 1):
        #     if (int(day_num) < 27):#那就上一年的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(11) + "/" + str(days[10]-3))  # 开始日期
        #         endtime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(27))  # 结束日期
        #     else:#那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        # elif (month_num == 2):
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(days[11] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        # else:
        #     if (int(day_num) < 27):  #那就上月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-2) + "/" + str(days[month_num-3] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(27))  # 结束日期
        #     else:  # 那就该月的27号到前30天
        #         starttime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2] - 3))  # 开始日期
        #         endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(27))  # 结束日期
        if (month_num == 1):  # 就要计算到上一年的12月
            starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(1))  # 开始日期
            endtime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11]))  # 结束日期
        else:  # 就要计算到上一月
            starttime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(1))  # 开始日期
            endtime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(days[month_num - 2]))  # 结束日期
        print("一月回归bugstarttime,endtime:", starttime, endtime)
    elif (int(isHalfYear) == 1):

        starttime = time.strftime(str(year_num) + "/" + str(1) + "/" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num) + "/" + str(6) + "/" + str(days[5]))  # 结束日期
        print("前半年回归bugstarttime,endtime:", starttime, endtime)
    elif (int(isHalfYear) == 2):

        starttime = time.strftime(str(year_num) + "/" + str(7) + "/" + str(1))  # 开始日期
        endtime = time.strftime(str(year_num) + "/" + str(12) + "/" + str(days[11]))  # 结束日期
        print("后半年回归bugstarttime,endtime:", starttime, endtime)

    createcol=19  #由谁创建
    statecol = 14  # Bug状态
    responcol = 38  # 负责人
    respondecol = 37  # 负责部门
    returncol=41  #回归截止时间
    returntestcol = 43  # 回归测试时间
    closecol=29 #关闭时间

    mi = -1  # 新的excel的第几行
    for i in range(0, nrows):
        mj=-1 #新的excel的第几列
        if (i == 0):
            # print("-----正在写入 " + str(i) + " 行")
            mi = mi + 1
            for j in range(0, ncols):
                if ((j != 39) & (j != 40)):
                    mj=mj+1
                    sheet.write(mi, mj, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        else:
            # 获取第i行第3列数据
            # sh.cell_value(i,3)
            # print("i:",i)
            createresult=sh.cell_value(i,createcol)
            stateresult = sh.cell_value(i, statecol)
            returnresult = sh.cell_value(i, returncol)
            returntestresult= sh.cell_value(i, returntestcol)
            closeresult=sh.cell_value(i, closecol)
            strftime = datetime.datetime.strptime(starttime, "%Y/%m/%d")
            strftime2 = datetime.datetime.strptime(endtime, "%Y/%m/%d")
            if(returntestresult!=""):#回归测试时间不为空
                strftime3 = datetime.datetime.strptime(returntestresult, "%Y/%m/%d")  # 回归测试时间
                if ((strftime3 >= strftime) & (strftime3 <= strftime2)):
                    # print("-----正在写入 " + str(i) + " 行")
                    # print("状态:" + stateresult)
                    mi = mi + 1
                    for j in range(0, ncols):
                        if ((j != 39) & (j != 40)):
                            mj = mj + 1

                            sheet.write(mi, mj, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
            else:#回归测试时间为空，就查看回归截止时间
                if(returnresult!=""):#回归截止时间不为空
                    strftime3 = datetime.datetime.strptime(returnresult, "%Y/%m/%d")  # 回归截止时间
                    if ((strftime3 >= strftime) & (strftime3 <= strftime2)):
                        # print("-----正在写入 " + str(i) + " 行")
                        # print("状态:" + stateresult)
                        mi = mi + 1
                        for j in range(0, ncols):
                            if ((j != 39) & (j != 40)):
                                mj = mj + 1
                                sheet.write(mi, mj, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
    book.save(allpath)


    rb = xlrd.open_workbook(allpath, formatting_info=True)  # 打开xls文件
    ro = rb.sheets()[0]  # 读取表单0
    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws = wb.get_sheet(0)  # 获取表单0
    for i in range(1, ro.nrows):  # 循环所有的行
        createname=ro.cell(i, createcol).value
        ind = -1
        for na in names:
            ind = ind + 1
            for name in na:
                if(createname!=None and createname!=""):
                    if (createname in name):
                        respoDepart = departs[ind]
                        ws.write(i, respondecol, respoDepart)
        ws.write(i, responcol, createname)
    wb.save(allpath)





#得到一周的任务的excel，并且修改
def getWeekTasksExcel():
    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    xlsxpath=Folder_Path+"/tasks.xls"
    # print(xlsxpath)
    # xlsxpath = "E:\\PythonWorkspace\\ZentaoTest\\result\\20190410 160724\\20190410160724tasks.xls"
    if (os.path.exists(Folder_Path+"/一周tasks.xls") == True):
        os.remove(Folder_Path+"/一周tasks.xls")
    bk = xlrd.open_workbook(xlsxpath)
    shxrange = range(bk.nsheets)
    try:
        sh = bk.sheet_by_name("data")
    except:
        print("代码出错")
    #35列
    nrows = sh.nrows  # 获取行数
    ncols = sh.ncols  # 获取列数
    book = Workbook(encoding='utf-8')
    sheet = book.add_sheet('data')  # 创建一个sheet

    today = time.strftime("%Y/%m/%d")#今天
    day_num=int(today.split('/')[2])  #几号
    month_num=int(today.split('/')[1])  #几月
    year_num=int(today.split('/')[0]) #几年
    flag=is_leap_year(year_num)
    if(flag==True):#该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    if(day_num-7<=0):#就要计算到上一个月和该月
        if(month_num==1):#如果这个月是一月份的话，直接计算上一年的12月份、该月
            starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11] - (7-(day_num ))))  # 开始日期
            # if(day_num-1==0):#直接计算12月
            #     starttime = time.strftime(str(year_num - 1) + "/" + str(12) + "/" + str(days[11] - 7))  # 开始日期
            #     # endtime = time.strftime(str(year_num-1) + "/" + str(12) + "/" + str(days[11] - 1))  # 结束日期
            # else:#计算12月和该月
            #     starttime = time.strftime((year_num - 1) + "/" + str(12) + "/" + str(days[11]-(day_num-1)))  # 开始日期
            #     # endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(day_num - 1))  # 结束日期

        else:#就计算该月或者上一月
            starttime = time.strftime(str(year_num) + "/" + str(month_num - 1) + "/" + str(days[month_num - 2] - (7-(day_num))) ) # 开始日
            # if (day_num-1== 0):#计算上一月
            #     starttime = time.strftime(str(year_num)+ "/" + str(month_num-1) + "/" + str(days[month_num-2] - 6))  # 开始日期
            #     # endtime = time.strftime(str(year_num) + "/" + str(month_num-1) + "/" + str(days[month_num-2]))  # 结束日期
            # else:#计算上月和该月
            #     starttime = time.strftime(str(year_num)+ "/" + str(month_num-1) + "/" + str(days[month_num-2]-(day_num-1)))  # 开始日期
            #     # endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(day_num - 1))  # 结束日期
    else:#就计算该月
        starttime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(day_num - 7))  # 开始日期
        # endtime = time.strftime(str(year_num) + "/" + str(month_num) + "/" + str(day_num - 1))  # 结束日期
    endtime=today  #结束日期
    print("一周任务starttime,endtime:",starttime,endtime)
    col = 0  # 指定修改的列
    statecol = 12  # 任务状态
    startcol = 10  # 实际开始
    endcol = 11  # 截止日期
    completecol=23 #完成时间
    exceptcol = 13  # 最初预计多少个工时
    closereasoncol= 28  #关闭原因--已完成，已取消，空白

    mi = -1#新的excel的第几行
    for i in range(0,nrows):
        if (i == 0):
            # print("-----正在写入 " + str(i) + " 行")
            mi=mi+1
            for j in range(0,ncols):
                sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        else:
            # 获取第i行第3列数据
            # sh.cell_value(i,3)
            # print("i:",i)
            stateresult = sh.cell_value(i, statecol)
            endresult = sh.cell_value(i, endcol)
            closereasonresult = sh.cell_value(i, closereasoncol)
            completeresult = sh.cell_value(i, completecol)
            strftime = datetime.datetime.strptime(starttime, "%Y/%m/%d")
            strftime2 = datetime.datetime.strptime(endtime, "%Y/%m/%d")
            # print("状态:" + stateresult)
            # ---------写出文件到excel--------
            if ((stateresult == "已完成")):  # 判断状态是否等于已完成
                strftime4 = datetime.datetime.strptime(completeresult, "%Y/%m/%d")  # 完成时间
                if ((strftime4 > strftime) & (strftime4 < strftime2)):
                    # print("-----正在写入 " + str(i) + " 行")
                    # print("状态:"+stateresult)
                    mi = mi + 1
                    for j in range(0, ncols):
                        sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
            elif (stateresult == "已关闭"):  # 判断状态是否等于已关闭
                if (closereasonresult == "已完成"):  # 如果关闭原因是已完成
                    # completestr = int(completeresult.split('-')[0] + completeresult.split('-')[1] + completeresult.split('-')[2])
                    if("0000" not in completeresult):#完成时间不为空
                        strftime4 = datetime.datetime.strptime(completeresult, "%Y/%m/%d")  # 完成时间
                        if ((strftime4 >= strftime) & (strftime4 <= strftime2)):
                            # print("-----正在写入 " + str(i) + " 行")
                            # print("状态:" + stateresult)
                            mi = mi + 1
                            for j in range(0, ncols):
                                sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
            elif ((stateresult == "进行中")|(stateresult == "未开始")):  #进行中、未开始状态都要加入进去
                # print("-----正在写入 " + str(i) + " 行")
                # print("状态:" + stateresult)
                mi = mi + 1
                for j in range(0, ncols):
                    sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        # 判断一下文件夹下的文件是否存在，存在就删除
    book.save(Folder_Path+"/一周tasks.xls")
    change_task_excel(Folder_Path+"/一周tasks.xls")



#得到统计的数据
def getAllData(session,isHalfYear):
    departNamelist = _getDepartment(session)
    departs = departNamelist[0]  # 部门
    names = departNamelist[1]  # 人

    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    if(int(isHalfYear)==0):
        allpath = Folder_Path + "/all.xls"
    else:
        allpath = Folder_Path + "/all2.xls"

    print("allpath:",allpath)
    rb = xlrd.open_workbook(allpath, formatting_info=True)
    sheets = rb.sheet_names()
    flag=False
    for sheet in sheets:
        if(int(isHalfYear)==0):
            if (sheet == "一月统计数据"):
                flag = True
        else:
            if (sheet == "半年统计数据"):
                flag = True

    if(flag==False):
        wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
        if(int(isHalfYear)==0):
            wb.add_sheet('一月统计数据')
        else:
            wb.add_sheet('半年统计数据')
        wb.save(allpath)
    # range(0, sheetlen)
    rb = xlrd.open_workbook(allpath, formatting_info=True)  # 打开 上月版本测试完成情况.xls文件，把其他文件的值复制到该文件
    if(int(isHalfYear)==0):
        ws1 = rb.sheet_by_name("一月tasks")  # 获取表单1
        nrows1 = ws1.nrows  # 获取行数
        ncols1 = ws1.ncols  # 获取列数
        ws2 = rb.sheet_by_name("一月解决bugs")  # 获取表单2
        nrows2 = ws2.nrows  # 获取行数
        ncols2 = ws2.ncols  # 获取列数
        ws3 = rb.sheet_by_name("一月回归bugs")  # 获取表单3
        nrows3 = ws3.nrows  # 获取行数
        ncols3 = ws3.ncols  # 获取列数
        ws4 = rb.sheet_by_name("一月创建bugs")  # 获取表单4
        nrows4 = ws4.nrows  # 获取行数
        ncols4 = ws4.ncols  # 获取列数
        ws5= rb.sheet_by_name("一月版本测试单")  # 获取表单6
        nrows5 = ws5.nrows  # 获取行数
        ncols5 = ws5.ncols  # 获取列数
        ws6 = rb.sheet_by_name("一月用例测试单")  # 获取表单7
        nrows6 = ws6.nrows  # 获取行数
        ncols6 = ws6.ncols  # 获取列数
        ws10 = rb.sheet_by_name("一月关闭bugs")  # 获取表单5
        nrows10 = ws10.nrows  # 获取行数
        ncols10 = ws10.ncols  # 获取列数

    else:
        ws1 = rb.sheet_by_name("半年tasks")  # 获取表单1
        nrows1 = ws1.nrows  # 获取行数
        ncols1 = ws1.ncols  # 获取列数
        ws2 = rb.sheet_by_name("半年解决bugs")  # 获取表单2
        nrows2 = ws2.nrows  # 获取行数
        ncols2 = ws2.ncols  # 获取列数
        ws3 = rb.sheet_by_name("半年回归bugs")  # 获取表单3
        nrows3 = ws3.nrows  # 获取行数
        ncols3 = ws3.ncols  # 获取列数
        ws4 = rb.sheet_by_name("半年创建bugs")  # 获取表单4
        nrows4 = ws4.nrows  # 获取行数
        ncols4 = ws4.ncols  # 获取列数
        ws5 = rb.sheet_by_name("半年版本测试单")  # 获取表单6
        nrows5 = ws5.nrows  # 获取行数
        ncols5 = ws5.ncols  # 获取列数
        ws6 = rb.sheet_by_name("半年用例测试单")  # 获取表单7
        nrows6 = ws6.nrows  # 获取行数
        ncols6 = ws6.ncols  # 获取列数
        ws10 = rb.sheet_by_name("半年关闭bugs")  # 获取表单5
        nrows10 = ws10.nrows  # 获取行数
        ncols10 = ws10.ncols  # 获取列数

    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws7 = wb.get_sheet(8)  # 获取表单8

    ws1Datas = []  # 存储任务tasks数据
    for i in range(0, nrows1):
        if(i!=0):
            departcol = 33 # 责任部门列
            namecol = 34 # 责任人列
            productlinecol = 32  # 产品线列
            firstexceptcol = 13 # 最初预计列
            alllosecol = 14  # 总消耗列
            finishcol = 35  # 任务完成及时
            departcolresult = ws1.cell_value(i, departcol).strip()
            namecolresult = ws1.cell_value(i, namecol).strip()
            productlinecolresult = ws1.cell_value(i, productlinecol).strip()
            firstexceptcolresult = ws1.cell_value(i, firstexceptcol).strip()
            alllosecolresult = ws1.cell_value(i, alllosecol).strip()
            finishcolresult = ws1.cell_value(i, finishcol).strip()
            # print("1111",i,namecolresult,productlinecolresult,firstexceptcolresult,alllosecolresult)

            flag = False
            for index in range(0, len(ws1Datas)):
                if ((ws1Datas[index][1] == namecolresult) & (ws1Datas[index][2] == productlinecolresult)):
                    flag = True
                    ind = index
                    break

            if (flag == True):  # 如果责任人和产品线一致，就修改最初预计、总消耗、总完成和及时完成的数目
                # print("相同ind:",ind)
                ws1Datas[ind][3] = ws1Datas[ind][3] + int(float(firstexceptcolresult))
                ws1Datas[ind][4] = ws1Datas[ind][4] + int(float(alllosecolresult))
                ws1Datas[ind][5] = ws1Datas[ind][5] + 1

                if (finishcolresult == "按时完成"):
                    ws1Datas[ind][6] = ws1Datas[ind][6] + 1
                    ws1Datas[ind][7] = ws1Datas[ind][7] + 1
                elif (finishcolresult == "延期完成"):
                    ws1Datas[ind][7] = ws1Datas[ind][7] + 1
            else:
                ws1Data = []
                ws1Data.append(departcolresult)
                ws1Data.append(namecolresult)
                ws1Data.append(productlinecolresult)
                ws1Data.append(int(float(firstexceptcolresult)))
                ws1Data.append(int(float(alllosecolresult)))
                ws1Data.append(1)  # 总任务数目,第5个
                if (finishcolresult == "按时完成"):
                    ws1Data.append(1)  # 及时完成任务数目，第6个
                    ws1Data.append(1)  # 已完成任务数目，第7个
                elif (finishcolresult == "延期完成"):
                    ws1Data.append(0)
                    ws1Data.append(1)
                else:
                    ws1Data.append(0)
                    ws1Data.append(0)
                ws1Datas.append(ws1Data)
            # print("ws1Datas:",len(ws1Datas),ws1Datas)
    print("任务ws1Datas:",len(ws1Datas),ws1Datas)


    ws11Datas = []  # 存储任务里的需求数据
    for i in range(0, nrows1):
        if (i != 0):
            departcol = 33  # 责任部门列
            namecol = 34  # 责任人列
            productlinecol = 32  # 产品线列
            requ2 = 3  # 相关需求

            departcolresult = ws1.cell_value(i, departcol).strip()
            namecolresult = ws1.cell_value(i, namecol).strip()
            productlinecolresult = ws1.cell_value(i, productlinecol).strip()
            requ2result = ws1.cell_value(i, requ2).strip()
            ws11Data = []
            flag11 = False
            for index in range(0, len(ws11Datas)):
                if ((ws11Datas[index][1] == namecolresult)&(ws11Datas[index][2] == requ2result)):
                    flag11 = True
                    break
            if (flag11 != True):  # 如果责任人和产品线和需求不一致
                ws11Data.append(departcolresult)
                ws11Data.append(namecolresult)
                # ws11Data.append(productlinecolresult)
                ws11Data.append(requ2result)
                ws11Datas.append(ws11Data)
    print("任务需求ws11Datas:", len(ws11Datas), ws11Datas)

    ws2Datas = []  #Bugs解决数量
    for i in range(0, nrows2):
        if(i!=0):
            departcol = 37  # 责任部门列
            namecol = 38  # 责任人列
            productlinecol = 36  # 产品线列
            levelcol=8 #严重等级
            solvecol=40 #解决及时性列
            departcolresult = ws2.cell_value(i, departcol).strip()
            namecolresult = ws2.cell_value(i, namecol).strip()
            productlinecolresult = ws2.cell_value(i, productlinecol).strip()
            levelcolresult=ws2.cell_value(i, levelcol).strip()
            solvecolresult=ws2.cell_value(i, solvecol).strip()
            ws2Data = []
            flag2 = False
            for index in range(0, len(ws2Datas)):
                if ((ws2Datas[index][1] == namecolresult) & (ws2Datas[index][2] == productlinecolresult)):
                    flag2 = True
                    ind2 = index
                    break
            if (flag2 == True):  # 如果责任人和产品线一致，就修改解决bug的严重等级4列和解决及时性的数目
                ws2Datas[ind2][3] = ws2Datas[ind2][3] + 1
                if (solvecolresult == "按时解决"):
                    ws2Datas[ind2][4] = ws2Datas[ind2][4] + 1
                    ws2Datas[ind2][5] = ws2Datas[ind2][5] + 1
                elif (solvecolresult == "延时解决"):
                    ws2Datas[ind2][5] = ws2Datas[ind2][5] + 1
                if ((solvecolresult == "按时解决") | (solvecolresult == "延时解决")):
                    if (int(levelcolresult) == 1):
                        ws2Datas[ind2][6] = ws2Datas[ind2][6] + 1
                    elif (int(levelcolresult) == 2):
                        ws2Datas[ind2][7] = ws2Datas[ind2][7] + 1
                    elif (int(levelcolresult) == 3):
                        ws2Datas[ind2][8] = ws2Datas[ind2][8] + 1
                    else:
                        ws2Datas[ind2][9] = ws2Datas[ind2][9] + 1
            else:
                ws2Data.append(departcolresult)
                ws2Data.append(namecolresult)
                ws2Data.append(productlinecolresult)
                ws2Data.append(1)  # 总解决bug数目,第3个
                if(solvecolresult=="按时解决"):
                    ws2Data.append(1)  # 及时解决bug数目,第4个
                    ws2Data.append(1)  # 已解决bug数目,第5个
                elif(solvecolresult=="延时解决"):
                    ws2Data.append(0)  # 及时解决bug数目,第4个
                    ws2Data.append(1)  # 已解决bug数目,第5个
                else:
                    ws2Data.append(0)  # 及时解决bug数目,第4个
                    ws2Data.append(0)  # 已解决bug数目,第5个
                if((solvecolresult=="按时解决")|(solvecolresult=="延时解决")):
                    if (int(levelcolresult) == 1):
                        ws2Data.append(1)  # 严重等级为1的数目,第6个
                        ws2Data.append(0)
                        ws2Data.append(0)
                        ws2Data.append(0)
                    elif (int(levelcolresult) == 2):
                        ws2Data.append(0)  # 严重等级为1的数目,第6个
                        ws2Data.append(1)
                        ws2Data.append(0)
                        ws2Data.append(0)
                    elif (int(levelcolresult) == 3):
                        ws2Data.append(0)  # 严重等级为1的数目,第6个
                        ws2Data.append(0)
                        ws2Data.append(1)
                        ws2Data.append(0)
                    else:
                        ws2Data.append(0)  # 严重等级为1的数目,第6个
                        ws2Data.append(0)
                        ws2Data.append(0)
                        ws2Data.append(1)
                else:
                    ws2Data.append(0)  # 严重等级为1的数目,第6个
                    ws2Data.append(0)
                    ws2Data.append(0)
                    ws2Data.append(0)
                ws2Datas.append(ws2Data)
    print("Bugs解决数量ws2Datas:", len(ws2Datas), ws2Datas)


    ws3Datas = []  # Bugs回归数量
    for i in range(0, nrows3):
        if (i != 0):
            departcol = 37  # 责任部门列
            namecol = 38  # 责任人列
            productlinecol = 36  # 产品线列
            returncol = 40  # 回归测试结果列
            departcolresult = ws3.cell_value(i, departcol).strip()
            namecolresult = ws3.cell_value(i, namecol).strip()
            productlinecolresult = ws3.cell_value(i, productlinecol).strip()
            returncolresult = ws3.cell_value(i, returncol).strip()
            ws3Data = []
            flag3 = False
            for index in range(0, len(ws3Datas)):
                if ((ws3Datas[index][1] == namecolresult) & (ws3Datas[index][2] == productlinecolresult)):
                    flag3 = True
                    ind3 = index
                    break
            if (flag3 == True):  # 如果责任人和产品线一致，就修改回归bug的数目
                ws3Datas[ind3][3] = ws3Datas[ind3][3] + 1
                if (returncolresult == "准时回归"):
                    ws3Datas[ind3][4] = ws3Datas[ind3][4] + 1
                    ws3Datas[ind3][5] = ws3Datas[ind3][5] + 1
                elif (returncolresult == "延时回归"):
                    ws3Datas[ind3][5] = ws3Datas[ind3][5] + 1
            else:
                ws3Data.append(departcolresult)
                ws3Data.append(namecolresult)
                ws3Data.append(productlinecolresult)
                ws3Data.append(1)  # 总回归bug数目,第3个
                if (returncolresult == "准时回归"):
                    ws3Data.append(1)  # 准时回归bug数目,第4个
                    ws3Data.append(1)  # 已回归bug数目,第5个
                elif (returncolresult == "延时回归"):
                    ws3Data.append(0)  # 准时回归bug数目,第4个
                    ws3Data.append(1)  # 已回归bug数目,第5个
                else:
                    ws3Data.append(0)  # 准时回归bug数目,第4个
                    ws3Data.append(0)  # 已回归bug数目,第5个
                ws3Datas.append(ws3Data)
    print("bugs回归数量ws3Datas:", len(ws3Datas), ws3Datas)


    ws4Datas = []  # Bugs创建数量
    for i in range(0, nrows4):
        if (i != 0):
            departcol = 37  # 责任部门列
            namecol = 38  # 责任人列
            # namecol=19 #创建人
            productlinecol = 36  # 产品线列
            levelcol = 8  # 严重等级
            departcolresult = ws4.cell_value(i, departcol).strip()
            namecolresult = ws4.cell_value(i, namecol).strip()
            productlinecolresult = ws4.cell_value(i, productlinecol).strip()
            levelcolresult = ws4.cell_value(i, levelcol).strip()
            ws4Data = []
            flag4 = False
            for index in range(0, len(ws4Datas)):
                if ((ws4Datas[index][1] == namecolresult) & (ws4Datas[index][2] == productlinecolresult)):
                    flag4 = True
                    ind4 = index
                    break
            if (flag4 == True):  # 如果责任人和产品线一致，就修改创建bug的严重等级4列数目
                if (int(levelcolresult) == 1):
                    ws4Datas[ind4][3] = ws4Datas[ind4][3] + 1
                elif (int(levelcolresult) == 2):
                    ws4Datas[ind4][4] = ws4Datas[ind4][4] + 1
                elif (int(levelcolresult) == 3):
                    ws4Datas[ind4][5] = ws4Datas[ind4][5] + 1
                else:
                    ws4Datas[ind4][6] = ws4Datas[ind4][6] + 1

            else:
                # respoDepart="" #根据创建人来得到部门
                # ind=-1
                # for na in names:
                #     ind = ind + 1
                #     for name in na:
                #         if(namecolresult!=None and namecolresult!=""):
                #             if (namecolresult in name):
                #                 respoDepart = departs[ind]

                # ws4Data.append(respoDepart)#部门
                ws4Data.append(departcolresult)#部门
                ws4Data.append(namecolresult)
                ws4Data.append(productlinecolresult)
                if (int(levelcolresult) == 1):
                    ws4Data.append(1)  # 严重等级为1的数目,第6个
                    ws4Data.append(0)
                    ws4Data.append(0)
                    ws4Data.append(0)
                elif (int(levelcolresult) == 2):
                    ws4Data.append(0)  # 严重等级为1的数目,第6个
                    ws4Data.append(1)
                    ws4Data.append(0)
                    ws4Data.append(0)
                elif (int(levelcolresult) == 3):
                    ws4Data.append(0)  # 严重等级为1的数目,第6个
                    ws4Data.append(0)
                    ws4Data.append(1)
                    ws4Data.append(0)
                else:
                    ws4Data.append(0)  # 严重等级为1的数目,第6个
                    ws4Data.append(0)
                    ws4Data.append(0)
                    ws4Data.append(1)
                ws4Datas.append(ws4Data)
    print("创建bugsws4Datas:", len(ws4Datas), ws4Datas)



    ws5Datas = []  # 需求产生的bug数-从创建bugs得到
    for i in range(0, nrows4):
        if (i != 0):
            departcol = 37  # 责任部门列
            namecol = 38  # 责任人列
            productlinecol = 36  # 产品线列
            requirecol = 4 # 相关需求
            levelcol = 8  # 严重等级
            departcolresult = ws4.cell_value(i, departcol).strip()
            namecolresult = ws4.cell_value(i, namecol).strip()
            productlinecolresult = ws4.cell_value(i, productlinecol).strip()
            requirecolresult = ws4.cell_value(i, requirecol).strip()
            levelcolresult = ws4.cell_value(i, levelcol).strip()
            ws5Data = []
            flag5 = False
            for index in range(0, len(ws5Datas)):
                if ((ws5Datas[index][1] == namecolresult) &(ws5Datas[index][2] == requirecolresult)):
                    flag5 = True
                    ind5 = index
                    break
            if (flag5 == True):  # 如果责任人和产品线和需求一致，就修改需求产生的bug数目
                if (int(levelcolresult) == 1):
                    ws5Datas[ind5][3] = ws5Datas[ind5][3] + 1
                elif (int(levelcolresult) == 2):
                    ws5Datas[ind5][4] = ws5Datas[ind5][4] + 1
                elif (int(levelcolresult) == 3):
                    ws5Datas[ind5][5] = ws5Datas[ind5][5] + 1
                else:
                    ws5Datas[ind5][6] = ws5Datas[ind5][6] + 1
            else:
                ws5Data.append(departcolresult)
                ws5Data.append(namecolresult)
                # ws5Data.append(productlinecolresult)
                ws5Data.append(requirecolresult) #需求名称
                if (int(levelcolresult) == 1):
                    ws5Data.append(1)  # 严重等级为1的数目,第5个
                    ws5Data.append(0)
                    ws5Data.append(0)
                    ws5Data.append(0)
                elif (int(levelcolresult) == 2):
                    ws5Data.append(0)  # 严重等级为1的数目,第5个
                    ws5Data.append(1)
                    ws5Data.append(0)
                    ws5Data.append(0)
                elif (int(levelcolresult) == 3):
                    ws5Data.append(0)  # 严重等级为1的数目,第5个
                    ws5Data.append(0)
                    ws5Data.append(1)
                    ws5Data.append(0)
                else:
                    ws5Data.append(0)  # 严重等级为1的数目,第5个
                    ws5Data.append(0)
                    ws5Data.append(0)
                    ws5Data.append(1)
                ws5Datas.append(ws5Data)
    print("需求产生的bug数ws5Datas:", len(ws5Datas), ws5Datas)

    ws6Datas = []  #版本测试单完成的数目
    for i in range(0, nrows5):
        if (i != 0):
            departcol = 6  # 责任部门列
            namecol = 5  # 责任人列
            productlinecol = 10  # 产品线列
            statecol = 9 # 版本状态

            departcolresult = ws5.cell_value(i, departcol).strip()
            namecolresult = ws5.cell_value(i, namecol).strip()
            productlinecolresult = ws5.cell_value(i, productlinecol).strip()
            statecolresult = ws5.cell_value(i, statecol).strip()
            ws6Data = []
            flag6 = False
            for index in range(0, len(ws6Datas)):
                if ((ws6Datas[index][1] == namecolresult) & (ws6Datas[index][2] == productlinecolresult)):
                    flag6 = True
                    ind6 = index
                    break
            if (flag6 == True):  # 如果责任人和产品线和需求一致，就修改版本完成数目
                ws6Datas[ind6][3] = ws6Datas[ind6][3] + 1
                if (statecolresult == "已完成"):
                    ws6Datas[ind6][4] = ws6Datas[ind6][4] + 1
            else:
                ws6Data.append(departcolresult)
                ws6Data.append(namecolresult)
                ws6Data.append(productlinecolresult)
                ws6Data.append(1) #版本数目
                if(statecolresult=="已完成"):
                    ws6Data.append(1)  # 版本完成的数目为1
                else:
                    ws6Data.append(0)  # 版本完成的数目为0
                ws6Datas.append(ws6Data)
    print("版本测试单完成数目ws6Datas:", len(ws6Datas), ws6Datas)


    ws7Datas = []  #用例测试单执行、完成数目
    for i in range(0, nrows6):
        if (i != 0):
            departcol = 6  # 责任部门列
            appointcol = 4  # 指派人列
            excucol= 5#执行人列
            productlinecol = 14  # 产品线列
            excutecol = 15 # 执行及时性列
            departcolresult = ws6.cell_value(i, departcol).strip()
            appointcolresult=ws6.cell_value(i,appointcol).strip()
            excucolresult=ws6.cell_value(i,excucol).strip()
            namecolresult=""
            if(excucolresult!=""):
                res = True
                for w in excucolresult:
                    if not '\u4e00' <= w <= '\u9fff':
                        res = False
                if(res==True):
                    namecolresult=excucolresult
            elif(appointcolresult!=""):
                res = True
                for w in appointcolresult:
                    if not '\u4e00' <= w <= '\u9fff':
                        res = False
                if (res == True):
                    namecolresult = appointcolresult
            productlinecolresult = ws6.cell_value(i, productlinecol)
            excutecolresult = ws6.cell_value(i, excutecol)
            ws7Data = []
            flag7 = False
            for index in range(0, len(ws7Datas)):
                if ((ws7Datas[index][1] == namecolresult) & (ws7Datas[index][2] == productlinecolresult)):
                    flag7 = True
                    ind7 = index
                    break
            if (flag7 == True):  # 如果责任人和产品线一致，就修改用例测试单及时执行的数目、、
                ws7Datas[ind7][3] = ws7Datas[ind7][3] + 1
                if (excutecolresult == "准时执行"):
                    ws7Datas[ind7][4] = ws7Datas[ind7][4] + 1
                    ws7Datas[ind7][5] = ws7Datas[ind7][5] + 1
                elif (excutecolresult == "延时执行"):
                    ws7Datas[ind7][5] = ws7Datas[ind7][5] + 1
            else:
                ws7Data.append(departcolresult)
                ws7Data.append(namecolresult)
                ws7Data.append(productlinecolresult)
                ws7Data.append(1) #总的用例测试单数量为1
                if(excutecolresult=="准时执行"):
                    ws7Data.append(1)  # 及时执行用例测试单数量为1
                    ws7Data.append(1)  # 完成执行用例测试单数量为1
                elif (excutecolresult == "延时执行"):
                    ws7Data.append(0)  # 及时执行用例测试单数量为0
                    ws7Data.append(1)  # 完成执行用例测试单数量为1
                else:
                    ws7Data.append(0)  # 及时执行用例测试单数量为0
                    ws7Data.append(0)  # 完成执行用例测试单数量为0
                ws7Datas.append(ws7Data)
    print("用例测试单执行完成数ws7Datas:", len(ws7Datas), ws7Datas)

    ws10Datas = []  # Bugs关闭数量
    for i in range(0, nrows10):
        if(i!=0):
            departcol = 37  # 责任部门列
            namecol = 38  # 责任人列
            productlinecol = 36  # 产品线列
            departcolresult = ws10.cell_value(i, departcol).strip()
            namecolresult = ws10.cell_value(i, namecol).strip()
            productlinecolresult = ws10.cell_value(i, productlinecol).strip()
            ws10Data = []
            flag10 = False
            for index in range(0, len(ws10Datas)):
                if ((ws10Datas[index][1] == namecolresult) & (ws10Datas[index][2] == productlinecolresult)):
                    flag10 = True
                    ind10 = index
                    break
            if (flag10 == True):  # 如果责任人和产品线一致，就关闭bugs数量增加1
                ws10Datas[ind10][3] = ws10Datas[ind10][3] + 1
            else:
                ws10Data.append(departcolresult)
                ws10Data.append(namecolresult)
                ws10Data.append(productlinecolresult)
                ws10Data.append(1)  # 总关闭bug数目,第3个
                ws10Datas.append(ws10Data)
    print("关闭bugsws10Datas:", len(ws10Datas), ws10Datas)


    #开始往统计数据sheet中写入数据
    #任务
    styleRedBkg = xlwt.easyxf('pattern: pattern solid, fore_colour red;')  # 红色
    m1=0
    ws7.write(m1, 0, "责任部门")
    ws7.write(m1, 1, "责任人")
    ws7.write(m1, 2, "产品线")
    ws7.write(m1, 3, "最初预计")
    ws7.write(m1, 4, "总消耗")
    ws7.write(m1, 5, "任务数")#
    ws7.write(m1, 6, "开发任务完成率")
    ws7.write(m1, 7, "开发任务完成及时率")
    ws7.write(m1, 8, "BUG解决总数") #
    ws7.write(m1, 9, "BUG解决率")
    ws7.write(m1, 10, "BUG解决及时率")
    ws7.write(m1, 11, "BUG回归测试完成率")
    ws7.write(m1, 12, "BUG回归测试及时率")
    ws7.write(m1, 13, "测试单执行完成率")
    ws7.write(m1, 14, "测试单执行及时率")
    ws7.write(m1, 15, "测试单完成情况")
    ws7.write(m1, 16, "BUG创建数量（1类）")
    ws7.write(m1, 17, "BUG创建数量（2类）")
    ws7.write(m1, 18, "BUG创建数量（3类）")
    ws7.write(m1, 19, "BUG创建数量（4类）")
    ws7.write(m1, 20, "BUG关闭数") #
    #任务完成
    for w1 in ws1Datas:
        m1=m1+1
        depatt=w1[0]
        name=w1[1]
        proline=w1[2]
        exceptr=w1[3]
        loser=w1[4]
        tasks=w1[5]
        # print(w1[5],w1[6],w1[7])
        # finis=str(w1[7])+"/"+str(w1[5]) #已完成率
        # onfinis=str(w1[6])+"/"+str(w1[5]) #及时完成率
        finis="%.0f%%" % (int(w1[7]) / int(w1[5]) * 100)
        onfinis="%.0f%%" % (int(w1[6]) / int(w1[5]) * 100)
        ws7.write(m1, 0, depatt)
        ws7.write(m1, 1, name)
        ws7.write(m1, 2, proline)
        ws7.write(m1, 3, exceptr)
        ws7.write(m1, 4, loser)
        ws7.write(m1, 5, tasks) #任务数
        ws7.write(m1, 6, finis)
        ws7.write(m1, 7, onfinis)
    wb.save(allpath)

    #bug解决率
    bk = xlrd.open_workbook(allpath)  # 打开该文件，复制到all.xls文件
    try:
        if(int(isHalfYear)==0):
            sh = bk.sheet_by_name("一月统计数据")
        else:
            sh = bk.sheet_by_name("半年统计数据")
    except:
        print("代码出错")
    rb = xlrd.open_workbook(allpath, formatting_info=True)  #
    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws7 = wb.get_sheet(8)  # 获取表单8
    nrows7 = sh.nrows  # 获取行数
    ncols7 = sh.ncols  # 获取列数
    for i in range(0,nrows7):
        for j in range(0,ncols7):
            ws7.write(i,j,sh.cell_value(i,j))
    m2=m1
    for w2 in ws2Datas:
        depatt=w2[0]
        name=w2[1]
        proline=w2[2]
        bugs=w2[5] #已经解决的bug总数
        # solve=str(w2[5])+"/"+str(w2[3]) #bug已经解决率
        # onsolve=str(w2[4])+"/"+str(w2[3]) #bug及时解决率
        solve = "%.0f%%" % (int(w2[5]) / int(w2[3]) * 100)
        onsolve="%.0f%%" % (int(w2[4]) / int(w2[3]) * 100)
        fflag2=False
        inde2=0
        for ii in range(1,nrows7):
            if((sh.cell_value(ii,0)==depatt)&(sh.cell_value(ii,1)==name)&(sh.cell_value(ii,2)==proline)):
                fflag2=True
                inde2=ii
                break
        if(fflag2==True):
            ws7.write(inde2, 8, bugs)
            ws7.write(inde2, 9, solve)
            ws7.write(inde2, 10, onsolve)
        else:
            m2 = m2 + 1
            ws7.write(m2, 0, depatt)
            ws7.write(m2, 1, name)
            ws7.write(m2, 2, proline)
            ws7.write(m2, 8, bugs)
            ws7.write(m2, 9, solve)
            ws7.write(m2, 10, onsolve)
    wb.save(allpath)

    #bug及时回归率
    bk = xlrd.open_workbook(allpath)  # 打开该文件，复制到all.xls文件
    try:
        if (int(isHalfYear) == 0):
            sh = bk.sheet_by_name("一月统计数据")
        else:
            sh = bk.sheet_by_name("半年统计数据")
    except:
        print("代码出错")
    rb = xlrd.open_workbook(allpath, formatting_info=True)  #
    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws7 = wb.get_sheet(8)  # 获取表单8
    nrows7 = sh.nrows  # 获取行数
    ncols7 = sh.ncols  # 获取列数
    for i in range(0, nrows7):
        for j in range(0, ncols7):
            ws7.write(i, j, sh.cell_value(i, j))
    m3 = m2
    for w3 in ws3Datas:
        depatt = w3[0]
        name = w3[1]
        proline = w3[2]
        # onreturn = str(w3[4]) + "/" + str(w3[3])  # bug及时回归率
        # retur = str(w3[5]) + "/" + str(w3[3])  # bug已经回归率
        onreturn="%.0f%%" % (int(w3[4]) / int(w3[3]) * 100)
        retur="%.0f%%" % (int(w3[5]) / int(w3[3]) * 100)
        fflag3 = False
        inde3 = 0
        for ii in range(1, nrows7):
            if ((sh.cell_value(ii, 0) == depatt) & (sh.cell_value(ii, 1) == name) & (sh.cell_value(ii, 2) == proline)):
                fflag3 = True
                inde3 = ii
                break
        if(fflag3==True):
            ws7.write(inde3, 11, retur)
            ws7.write(inde3, 12, onreturn)
        else:
            m3 = m3 + 1
            ws7.write(m3, 0, depatt)
            ws7.write(m3, 1, name)
            ws7.write(m3, 2, proline)
            ws7.write(m3, 11, retur)
            ws7.write(m3, 12, onreturn)
    wb.save(allpath)

    #版本测试单完成情况
    bk = xlrd.open_workbook(allpath)  # 打开该文件，复制到all.xls文件
    try:
        if (int(isHalfYear) == 0):
            sh = bk.sheet_by_name("一月统计数据")
        else:
            sh = bk.sheet_by_name("半年统计数据")
    except:
        print("代码出错")
    rb = xlrd.open_workbook(allpath, formatting_info=True)  #
    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws7 = wb.get_sheet(8)  # 获取表单8
    nrows7 = sh.nrows  # 获取行数
    ncols7 = sh.ncols  # 获取列数
    for i in range(0, nrows7):
        for j in range(0, ncols7):
            ws7.write(i, j, sh.cell_value(i, j))
    m4 = m3
    for w6 in ws6Datas:
        depatt = w6[0]
        name = w6[1]
        proline = w6[2]
        onVersi = str(w6[4]) + "/" + str(w6[3])  # 版本完成情况
        fflag4 = False
        inde4 = 0
        for ii in range(1, nrows7):
            if ((sh.cell_value(ii, 0) == depatt) & (sh.cell_value(ii, 1) == name) & (sh.cell_value(ii, 2) == proline)):
                fflag4 = True
                inde4 = ii
                break
        if(fflag4==True):
            ws7.write(inde4, 15, onVersi)
        else:
            m4 = m4 + 1
            ws7.write(m4, 0, depatt)
            ws7.write(m4, 1, name)
            ws7.write(m4, 2, proline)
            ws7.write(m4, 15, onVersi)
    wb.save(allpath)

   #用例及时率/完成率
    bk = xlrd.open_workbook(allpath)  # 打开该文件，复制到all.xls文件
    try:
        if (int(isHalfYear) == 0):
            sh = bk.sheet_by_name("一月统计数据")
        else:
            sh = bk.sheet_by_name("半年统计数据")
    except:
        print("代码出错")

    rb = xlrd.open_workbook(allpath, formatting_info=True)  #
    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws7 = wb.get_sheet(8)  # 获取表单8
    nrows7 = sh.nrows  # 获取行数
    ncols7 = sh.ncols  # 获取列数
    for i in range(0, nrows7):
        for j in range(0, ncols7):
            ws7.write(i, j, sh.cell_value(i, j))
    m5 = m4
    for w7 in ws7Datas:
        depatt = w7[0]
        name = w7[1]
        proline = w7[2]
        # ontest = str(w7[5]) + "/" + str(w7[4])  #测试单执行及时率
        # test = str(w7[6]) + "/" + str(w7[4])  # 测试单执行完成率
        ontest="%.0f%%" % (int(w7[4]) / int(w7[3]) * 100)
        test="%.0f%%" % (int(w7[5]) / int(w7[3]) * 100)
        fflag5 = False
        inde5 = 0
        for ii in range(1, nrows7):
            if ((sh.cell_value(ii, 0) == depatt) & (sh.cell_value(ii, 1) == name) & (sh.cell_value(ii, 2) == proline)):
                fflag5 = True
                inde5 = ii
                break
        if(fflag5==True):
            ws7.write(inde5, 13, test)
            ws7.write(inde5, 14, ontest)
        else:
            m5 = m5 + 1
            ws7.write(m5, 0, depatt)
            ws7.write(m5, 1, name)
            ws7.write(m5, 2, proline)
            ws7.write(m5, 13, test)
            ws7.write(m5, 14, ontest)
    wb.save(allpath)

    #bug创建数量
    bk = xlrd.open_workbook(allpath)  # 打开该文件，复制到all.xls文件
    try:
        if (int(isHalfYear) == 0):
            sh = bk.sheet_by_name("一月统计数据")
        else:
            sh = bk.sheet_by_name("半年统计数据")
    except:
        print("代码出错")
    rb = xlrd.open_workbook(allpath, formatting_info=True)  #
    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws7 = wb.get_sheet(8)  # 获取表单8
    nrows7 = sh.nrows  # 获取行数
    ncols7 = sh.ncols  # 获取列数
    for i in range(0, nrows7):
        for j in range(0, ncols7):
            ws7.write(i, j, sh.cell_value(i, j))
    m6 = m5
    for w4 in ws4Datas:
        depatt = w4[0]
        name = w4[1]
        proline = w4[2]
        bug1=w4[3]
        bug2 = w4[4]
        bug3 = w4[5]
        bug4 = w4[6]
        fflag6 = False
        inde6 = 0
        for ii in range(1, nrows7):
            if ((sh.cell_value(ii, 0) == depatt) & (sh.cell_value(ii, 1) == name) & (sh.cell_value(ii, 2) == proline)):
                fflag6 = True
                inde6 = ii
                break
        if(fflag6==True):
            ws7.write(inde6, 16, bug1)
            ws7.write(inde6, 17, bug2)
            ws7.write(inde6, 18, bug3)
            ws7.write(inde6, 19, bug4)
        else:
            m6 = m6 + 1
            ws7.write(m6, 0, depatt)
            ws7.write(m6, 1, name)
            ws7.write(m6, 2, proline)
            ws7.write(m6, 16, bug1)
            ws7.write(m6, 17, bug2)
            ws7.write(m6, 18, bug3)
            ws7.write(m6, 19, bug4)
    wb.save(allpath)

    #bug关闭数量
    bk = xlrd.open_workbook(allpath)  # 打开该文件，复制到all.xls文件
    try:
        if (int(isHalfYear) == 0):
            sh = bk.sheet_by_name("一月统计数据")
        else:
            sh = bk.sheet_by_name("半年统计数据")
    except:
        print("代码出错")
    rb = xlrd.open_workbook(allpath, formatting_info=True)  #
    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws7 = wb.get_sheet(8)  # 获取表单8
    nrows7 = sh.nrows  # 获取行数
    ncols7 = sh.ncols  # 获取列数
    for i in range(0, nrows7):
        for j in range(0, ncols7):
            ws7.write(i, j, sh.cell_value(i, j))
    for w10 in ws10Datas:
        depatt = w10[0]
        name = w10[1]
        proline = w10[2]
        closebug=w10[3]
        fflag10 = False
        inde10 = 0
        for ii in range(1, nrows7):
            if ((sh.cell_value(ii, 0) == depatt) & (sh.cell_value(ii, 1) == name) & (sh.cell_value(ii, 2) == proline)):
                fflag10 = True
                inde10 = ii
                break
        if(fflag10==True):
            ws7.write(inde10, 20, closebug)
        else:
            m6 = m6 + 1
            ws7.write(m6, 0, depatt)
            ws7.write(m6, 1, name)
            ws7.write(m6, 2, proline)
            ws7.write(m6, 20, closebug)
    wb.save(allpath)


    #获取一些总数据的数组
    wsAllDatas=getAllDatas(ws1Datas, ws2Datas, ws3Datas, ws4Datas, ws6Datas, ws7Datas,ws10Datas)

   #这里添加总的统计
    bk = xlrd.open_workbook(allpath)  # 打开该文件，复制到all.xls文件
    try:
        if (int(isHalfYear) == 0):
            sh = bk.sheet_by_name("一月统计数据")
        else:
            sh = bk.sheet_by_name("半年统计数据")
    except:
        print("代码出错")
    nrows7 = sh.nrows  # 获取行数
    ncols7 = sh.ncols  # 获取列数
    wflags=[]
    for i in range(0, nrows7):
        if(i!=0):
            namecol=1 #责任人
            excepcol=3 #最初预计
            solvecol=8 #bug解决
            retucol=11  #bug回归
            testcol=13 #测试用例
            versicol=15#版本
            bugcol=16  #bug创建
            closecol=20 #bug关闭
            namecolresult=sh.cell_value(i,namecol).strip()
            excepcolresult=sh.cell_value(i,excepcol)
            solvecolresult=sh.cell_value(i,solvecol)
            retucolresult=sh.cell_value(i,retucol)
            testcolresult=sh.cell_value(i,testcol)
            versicolresult=sh.cell_value(i,versicol)
            bugcolresult=sh.cell_value(i,bugcol)
            closecolresult=sh.cell_value(i,closecol) #bug关闭数量
            flag=False
            for index in range(0,len(wflags)):
                item=wflags[index]
                if(item[0]==namecolresult):
                    flag=True
                    break
            if(flag==True):
                if (excepcolresult != ""):
                    item[1]=1

                if (solvecolresult != ""):
                    item[2] =1

                if (retucolresult != ""):
                    item[3] = 1

                if (testcolresult != ""):
                    item[4] =1

                if (versicolresult != ""):
                    item[5] = 1

                if (bugcolresult != ""):
                    item[6] = 1

                if (closecolresult != ""):
                    item[7] = 1
            else:
                wflag = []
                wflag.append(namecolresult)
                if(excepcolresult!=""):
                    wflag.append(1)
                else:
                    wflag.append(0)
                if (solvecolresult != ""):
                    wflag.append(1)
                else:
                    wflag.append(0)
                if (retucolresult != ""):
                    wflag.append(1)
                else:
                    wflag.append(0)
                if (testcolresult != ""):
                    wflag.append(1)
                else:
                    wflag.append(0)
                if (versicolresult != ""):
                    wflag.append(1)
                else:
                    wflag.append(0)
                if (bugcolresult != ""):
                    wflag.append(1)
                else:
                    wflag.append(0)
                if (closecolresult != ""):
                    wflag.append(1)
                else:
                    wflag.append(0)
                wflags.append(wflag)
    print("有无标志wflags:",len(wflags),wflags)   #查看相同名字的哪一项有，哪一项没有,有7项

    rb = xlrd.open_workbook(allpath, formatting_info=True)  #
    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws7 = wb.get_sheet(8)  # 获取表单8
    nrows7 = sh.nrows  # 获取行数
    ncols7 = sh.ncols  # 获取列数
    for i in range(0, nrows7):
        for j in range(0, ncols7):
            ws7.write(i, j, sh.cell_value(i, j))
    for wsAll in wsAllDatas:
        m6 = m6 + 1
        depatt = wsAll[0]
        name = wsAll[1]
        proline = wsAll[2]
        exceptr = wsAll[3]
        loser = wsAll[4]
        tasks=wsAll[5]
        for index in range(0,len(wflags)):
            item=wflags[index]
            if(name==item[0]):#相同就跳出循环
                break
        ws7.write(m6, 0, depatt)
        ws7.write(m6, 1, name)
        ws7.write(m6, 2, proline)
        if (item[1] == 1):
            if (int(wsAll[5]) != 0):
                finis = "%.0f%%" % (int(wsAll[7]) / int(wsAll[5]) * 100)
                onfinis = "%.0f%%" % (int(wsAll[6]) / int(wsAll[5]) * 100)
            else:
                finis = "%.0f%%" % 0
                onfinis = "%.0f%%" % 0
            ws7.write(m6, 3, exceptr)
            ws7.write(m6, 4, loser)
            ws7.write(m6, 5, tasks)
            ws7.write(m6, 6, finis)
            ws7.write(m6, 7, onfinis)
        if (item[2] == 1):
            if (int(wsAll[8]) != 0):
                solve = "%.0f%%" % (int(wsAll[10]) / int(wsAll[8]) * 100)
                onsolve = "%.0f%%" % (int(wsAll[9]) / int(wsAll[8]) * 100)
                bugs=wsAll[10]
            else:
                onsolve = "%.0f%%" % 0
                solve = "%.0f%%" % 0
                bugs=0
            ws7.write(m6, 8, bugs)
            ws7.write(m6, 9, solve)
            ws7.write(m6, 10, onsolve)

        if(item[3]==1):
            if (int(wsAll[15]) != 0):
                onreturn = "%.0f%%" % (int(wsAll[16]) / int(wsAll[15]) * 100)
                retur = "%.0f%%" % (int(wsAll[17]) / int(wsAll[15]) * 100)
            else:
                onreturn = "%.0f%%" % 0
                retur = "%.0f%%" % 0
            ws7.write(m6, 11, retur)
            ws7.write(m6, 12, onreturn)

        if(item[5]==1):
            onVersi = str(wsAll[23]) + "/" + str(wsAll[22])  # 版本完成情况
            ws7.write(m6, 15, onVersi)
        if(item[4]==1):
            if (int(wsAll[24]) != 0):
                ontest = "%.0f%%" % (int(wsAll[25]) / int(wsAll[24]) * 100)
                test = "%.0f%%" % (int(wsAll[26]) / int(wsAll[24]) * 100)
            else:
                ontest = "%.0f%%" % 0
                test = "%.0f%%" % 0
            ws7.write(m6, 13, test)
            ws7.write(m6, 14, ontest)
        bug1 = wsAll[18]
        bug2 = wsAll[19]
        bug3 = wsAll[20]
        bug4 = wsAll[21]
        if(item[6]==1):
            ws7.write(m6, 16, bug1)
            ws7.write(m6, 17, bug2)
            ws7.write(m6, 18, bug3)
            ws7.write(m6, 19, bug4)
        closebugs=wsAll[27]
        if(item[7]==1):
            ws7.write(m6,20,closebugs)
    wb.save(allpath)




    #需求产生的bug
    bk = xlrd.open_workbook(allpath)  # 打开该文件，复制到all.xls文件
    try:
        if (int(isHalfYear) == 0):
            sh = bk.sheet_by_name("一月统计数据")
        else:
            sh = bk.sheet_by_name("半年统计数据")
    except:
        print("代码出错")

    rb = xlrd.open_workbook(allpath, formatting_info=True)  #
    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws7 = wb.get_sheet(8)  # 获取表单8
    nrows7 = sh.nrows  # 获取行数
    ncols7 = sh.ncols  # 获取列数
    for i in range(0, nrows7):
        for j in range(0, ncols7):
            ws7.write(i, j, sh.cell_value(i, j))
    m7 = m6+4
    ws7.write(m7, 0, "责任部门")
    ws7.write(m7, 1, "责任人")
    # ws7.write(m7, 2, "产品线")
    ws7.write(m7, 2, "参与开发需求")
    ws7.write(m7, 3, "bug所属需求")
    ws7.write(m7, 4, "需求产生BUG数（1类）")
    ws7.write(m7, 5, "需求产生BUG数（2类）")
    ws7.write(m7, 6, "需求产生BUG数（3类）")
    ws7.write(m7, 7, "需求产生BUG数（4类）")
    sames=[]
    for index2 in range(0,len(ws5Datas)):
        w5=ws5Datas[index2]
        m7=m7+1
        depatt = w5[0]
        name = w5[1]
        # proline = w5[2]
        reque=w5[2]
        bug1 = w5[3]
        bug2 = w5[4]
        bug3 = w5[5]
        bug4 = w5[6]
        flag55=False
        inds=[]
        for index in range(0,len(ws11Datas)):
            item=ws11Datas[index]
            if((item[1]==w5[1])&(item[2]==w5[2])):#负责人与需求名字相同
                flag55 = True
                ind55 = index
                break
        if (flag55 == True):
            # flag = False
            # for ind in inds:
            #     if (ind == ind55):
            #         flag = True
            #         break
            # if (flag != True):
            #     inds.append(ind55)
            reque2 = item[2] #任务需求
            sames.append(ind55)
            ws11Datas.pop(ind55)#把bug与任务需求里相同的给去掉
            ws7.write(m7, 2, reque2)
        else:
            ws7.write(m7, 3, reque) #bug需求
        ws7.write(m7, 0, depatt)
        ws7.write(m7, 1, name)
        # ws7.write(m7, 2, proline)
        ws7.write(m7, 4, bug1)
        ws7.write(m7, 5, bug2)
        ws7.write(m7, 6, bug3)
        ws7.write(m7, 7, bug4)
    # for m in range(0, len(inds)):
    #     ws11Datas.pop(inds[m])#把bug与任务需求里相同的给去掉

    print("相同的需求有几个：",len(sames))
    for w111 in ws11Datas:
        m7=m7+1
        ws7.write(m7, 2, w111[2])
        ws7.write(m7, 0, w111[0])
        ws7.write(m7, 1, w111[1])
        # ws7.write(m7, 2, proline)
        ws7.write(m7, 4, 0)
        ws7.write(m7, 5, 0)
        ws7.write(m7, 6, 0)
        ws7.write(m7, 7, 0)
    wb.save(allpath)

    #bug解决数量
    bk = xlrd.open_workbook(allpath)  # 打开该文件，复制到all.xls文件
    try:
        if (int(isHalfYear) == 0):
            sh = bk.sheet_by_name("一月统计数据")
        else:
            sh = bk.sheet_by_name("半年统计数据")
    except:
        print("代码出错")

    rb = xlrd.open_workbook(allpath, formatting_info=True)
    wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
    ws7 = wb.get_sheet(8)  # 获取表单8
    nrows7 = sh.nrows  # 获取行数
    ncols7 = sh.ncols  # 获取列数
    for i in range(0, nrows7):
        if((i==0)|(i==(m6+4))):
            for j in range(0, ncols7):
                ws7.write(i, j, sh.cell_value(i, j),styleRedBkg)
        else:
            for j in range(0, ncols7):
                ws7.write(i, j, sh.cell_value(i, j))
    m8 = m7+4
    ws7.write(m8, 0, "责任部门",styleRedBkg)
    ws7.write(m8, 1, "责任人",styleRedBkg)
    ws7.write(m8, 2, "产品线",styleRedBkg)
    ws7.write(m8, 3, "BUG解决数量（1）",styleRedBkg)
    ws7.write(m8, 4, "BUG解决数量（2）",styleRedBkg)
    ws7.write(m8, 5, "BUG解决数量（3）",styleRedBkg)
    ws7.write(m8, 6, "BUG解决数量（4）",styleRedBkg)
    for w2 in ws2Datas:
        m8=m8+1
        depatt = w2[0]
        name = w2[1]
        proline = w2[2]
        bug1 = w2[6]
        bug2 = w2[7]
        bug3 = w2[8]
        bug4 = w2[9]
        ws7.write(m8, 0, depatt)
        ws7.write(m8, 1, name)
        ws7.write(m8, 2, proline)
        ws7.write(m8, 3, bug1)
        ws7.write(m8, 4, bug2)
        ws7.write(m8, 5, bug3)
        ws7.write(m8, 6, bug4)
    wb.save(allpath)


#得到Bug状态为激活的Bug
def getActiviBugExcel():
    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    xlsxpath=Folder_Path+"/bugs.xls"
    # print(xlsxpath)
    # xlsxpath = "E:\\PythonWorkspace\\ZentaoTest\\result\\20191108 163817\\bugs.xls"

    allpath = Folder_Path + "/激活bugs.xls"
    if (os.path.exists(allpath) == True):
        os.remove(allpath)
    bk = xlrd.open_workbook(xlsxpath)
    shxrange = range(bk.nsheets)
    try:
        sh = bk.sheet_by_name("data")
    except:
        print("代码出错")
    nrows = sh.nrows  # 获取行数
    ncols = sh.ncols  # 获取列数
    book = Workbook(encoding='utf-8')
    sheet = book.add_sheet('data')  # 创建一个sheet

    today = time.strftime("%Y/%m/%d")  # 今天
    day_num = int(today.split('/')[2])  # 几号
    month_num = int(today.split('/')[1])  # 几月
    year_num = int(today.split('/')[0])  # 几年
    flag = is_leap_year(year_num)
    if (flag == True):  # 该年是闰年
        days = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 闰年
    else:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    statecol = 14  # Bug状态
    mi = -1  # 新的excel的第几行
    for i in range(0, nrows):
        if (i == 0):
            # print("-----正在写入 " + str(i) + " 行")
            mi = mi + 1
            for j in range(0, ncols):
                if((j!=43)&(j!=41)&(j!=42)):
                    sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        else:
            # 获取第i行第3列数据
            # sh.cell_value(i,3)
            # print("i:",i)
            stateresult = sh.cell_value(i, statecol)
            if(stateresult=="激活"):
                mi = mi + 1
                for j in range(0, ncols):
                    if ((j != 43) & (j != 41) & (j != 42)):
                        sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
    book.save(allpath)



#得到任务状态为未开始和进行中的任务
def getNoStartTasksExcel():
    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    xlsxpath=Folder_Path+"/tasks.xls"
    # print(xlsxpath)
    # xlsxpath = "E:\\PythonWorkspace\\ZentaoTest\\result\\20191108 163817\\tasks.xls"
    if (os.path.exists(Folder_Path+"/未开始进行中tasks.xls") == True):
        os.remove(Folder_Path+"/未开始进行中tasks.xls")
    bk = xlrd.open_workbook(xlsxpath)
    shxrange = range(bk.nsheets)
    try:
        sh = bk.sheet_by_name("data")
    except:
        print("代码出错")
    #35列
    nrows = sh.nrows  # 获取行数
    ncols = sh.ncols  # 获取列数
    book = Workbook(encoding='utf-8')
    sheet = book.add_sheet('data')  # 创建一个sheet

    today = time.strftime("%Y/%m/%d")#今天
    day_num=int(today.split('/')[2])  #几号
    month_num=int(today.split('/')[1])  #几月
    year_num=int(today.split('/')[0]) #几年

    statecol = 12  # 任务状态


    mi = -1#新的excel的第几行
    for i in range(0,nrows):
        if (i == 0):
            # print("-----正在写入 " + str(i) + " 行")
            mi=mi+1
            for j in range(0,ncols):
                sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        else:
            stateresult = sh.cell_value(i, statecol)
            # print("状态:" + stateresult)
            # ---------写出文件到excel--------
            if ((stateresult == "未开始") or (stateresult == "进行中")):  # 判断状态是否等于未开始、进行中
                mi = mi + 1
                for j in range(0, ncols):
                    sheet.write(mi, j, label=sh.cell_value(i, j))  # 向第i行第j列写入获取到的值
        # 判断一下文件夹下的文件是否存在，存在就删除
    book.save(Folder_Path+"/未开始进行中tasks.xls")



#得到未开始或进行中的任务数组
def getNoStartTasksData():
    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    allpath = Folder_Path + "/未开始进行中tasks.xls"
    print("allpath:", allpath)
    rb = xlrd.open_workbook(allpath, formatting_info=True)  # 打开 上月版本测试完成情况.xls文件，把其他文件的值复制到该文件
    ws1 = rb.sheet_by_name("data")  # 获取表单
    nrows1 = ws1.nrows  # 获取行数
    ncols1 = ws1.ncols  # 获取列数
    ws1Datas = []  # 存储任务tasks数据
    for i in range(0, nrows1):
        if(i!=0):
            idcol=0  #编号列
            fathertaskcol=4 #任务名称列
            sontaskcol=5  #子任务名称列
            exceptstartcol=9   #预计开始时间列
            actuastartcol=10   #实际开始时间列
            endtimecol=11 #截止时间列
            taskstatecol=12 #任务状态列
            createnamecol=18  #创建人列
            createtimecol=19 #创建时间列
            namecol = 34 # 责任人列
            idcolresult=ws1.cell_value(i, idcol).strip()
            fathertaskcolresult = ws1.cell_value(i, fathertaskcol).strip()
            sontaskcolresult = ws1.cell_value(i, sontaskcol).strip()
            exceptstartcolresult = ws1.cell_value(i, exceptstartcol).strip()
            actuastartcolresult = ws1.cell_value(i, actuastartcol).strip()
            endtimecolresult = ws1.cell_value(i, endtimecol).strip()
            taskstatecolresult = ws1.cell_value(i, taskstatecol).strip()
            createnamecolresult = ws1.cell_value(i, createnamecol).strip()
            createtimecolresult = ws1.cell_value(i, createtimecol).strip()
            namecolresult = ws1.cell_value(i, namecol).strip()
            ws1Data = []
            ws1Data.append(idcolresult)
            if(str(sontaskcolresult)!=""):
                ws1Data.append(str(fathertaskcolresult) + "/"+str(sontaskcolresult))
            else:
                ws1Data.append(str(fathertaskcolresult))
            ws1Data.append(exceptstartcolresult)
            ws1Data.append(actuastartcolresult)
            ws1Data.append(endtimecolresult)
            ws1Data.append(taskstatecolresult)
            ws1Data.append(createnamecolresult)
            ws1Data.append(createtimecolresult)
            ws1Data.append(namecolresult)
            ws1Datas.append(ws1Data)
        else:
            ws1Data=[]
            #9个
            ws1Data.append("任务编号")
            ws1Data.append("任务名称")
            ws1Data.append("预计开始时间")
            ws1Data.append("实际开始时间")
            ws1Data.append("截止日期")
            ws1Data.append("任务状态")
            ws1Data.append("任务创建人")
            ws1Data.append("任务创建时间")
            ws1Data.append("任务责任人")
            ws1Datas.append(ws1Data)

    print("未开始或进行中tasks ws1Datas:",len(ws1Datas),ws1Datas)
    return ws1Datas




#得到激活的bug数组
def getActiviBugsData():
    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    allpath = Folder_Path + "/激活bugs.xls"
    print("allpath:", allpath)
    rb = xlrd.open_workbook(allpath, formatting_info=True)  # 打开 上月版本测试完成情况.xls文件，把其他文件的值复制到该文件
    ws1 = rb.sheet_by_name("data")  # 获取表单
    nrows1 = ws1.nrows  # 获取行数
    ncols1 = ws1.ncols  # 获取列数
    ws1Datas=[] #激活状态Bugs详情
    for i in range(0, nrows1):
        if (i != 0):
            idcol=0  #Bug编号列
            bugnamecol=6  #bug标题列
            # endtimecol=15  #bug截止时间列
            createnamecol=19  #创建人列
            createtimecol=20 #创建日期列
            namecol = 38  # 责任人列
            rquiresolvecol=39  #要求解决时间列
            idcolresult = ws1.cell_value(i, idcol).strip()
            bugnamecolresult = ws1.cell_value(i, bugnamecol).strip()
            # endtimecolresult = ws1.cell_value(i, endtimecol).strip()
            createnamecolresult = ws1.cell_value(i, createnamecol).strip()
            createtimecolresult = ws1.cell_value(i, createtimecol).strip()
            namecolresult = ws1.cell_value(i, namecol).strip()
            rquiresolvecolresult = ws1.cell_value(i, rquiresolvecol).strip()
            ws1Data = []
            ws1Data.append(idcolresult)
            ws1Data.append(bugnamecolresult)
            ws1Data.append(createnamecolresult)
            ws1Data.append(createtimecolresult)
            ws1Data.append(namecolresult)
            ws1Data.append(rquiresolvecolresult)
            ws1Datas.append(ws1Data)
            # print(ws1Data)
        else:
            ws1Data=[]
            #6个
            ws1Data.append("Bug编号")
            ws1Data.append("Bug标题")
            # ws1Data.append("Bug截止解决时间")
            ws1Data.append("Bug创建人")
            ws1Data.append("Bug创建日期")
            ws1Data.append("Bug责任人")
            ws1Data.append("Bug要求解决时间")
            ws1Datas.append(ws1Data)
            # print(ws1Data)
    print("激活状态bug ws1Datas:", len(ws1Datas), ws1Datas)
    return  ws1Datas


# 获取
def getWorkTimeTask():
    path = os.path.dirname(os.getcwd()) + '/result'
    Folder_Path = new_file(path)  # 得到结果文件最新的文件夹
    # os.chdir(Folder_Path)
    # file_list = os.listdir()
    allpath = Folder_Path + "/工时消耗.xls"
    print("allpath:", allpath)
    rb = xlrd.open_workbook(allpath, formatting_info=True)  # 打开 上月版本测试完成情况.xls文件，把其他文件的值复制到该文件
    ws1 = rb.sheet_by_name("data")  # 获取表单
    nrows1 = ws1.nrows  # 获取行数
    ncols1 = ws1.ncols  # 获取列数
    ws1Datas = []  # 存储任务tasks数据
    for i in range(0, nrows1):
        if(i!=0):
            idcol=0  #编号列
            taskstatus=1# 任务状态
            fathertaskcol=4 #任务名称列
            sontaskcol=5  #子任务名称列
            exceptstartcol=9   #预计开始时间列
            actuastartcol=10   #实际开始时间列
            endtimecol=11 #截止时间列
            taskstatecol=12 #任务状态列
            createnamecol=18  #创建人列
            createtimecol=19 #创建时间列
            namecol = 34 # 责任人列
            mothworktime=38 #当月消耗工时
            taskstatusresult=ws1.cell_value(i,taskstatus).strip()
            if taskstatecolresult != "未开始":
                idcolresult=ws1.cell_value(i, idcol).strip()
                fathertaskcolresult = ws1.cell_value(i, fathertaskcol).strip()
                sontaskcolresult = ws1.cell_value(i, sontaskcol).strip()
                exceptstartcolresult = ws1.cell_value(i, exceptstartcol).strip()
                actuastartcolresult = ws1.cell_value(i, actuastartcol).strip()
                endtimecolresult = ws1.cell_value(i, endtimecol).strip()
                taskstatecolresult = ws1.cell_value(i, taskstatecol).strip()
                createnamecolresult = ws1.cell_value(i, createnamecol).strip()
                createtimecolresult = ws1.cell_value(i, createtimecol).strip()
                namecolresult = ws1.cell_value(i, namecol).strip()
                mothworktimeresult=ws1.cell_value(i,mothworktime).strip()
                ws1Data = []
                ws1Data.append(idcolresult)
                if(str(sontaskcolresult)!=""):
                    ws1Data.append(str(fathertaskcolresult) + "/"+str(sontaskcolresult))
                else:
                    ws1Data.append(str(fathertaskcolresult))
                ws1Data.append(exceptstartcolresult)
                ws1Data.append(actuastartcolresult)
                ws1Data.append(endtimecolresult)
                ws1Data.append(taskstatecolresult)
                ws1Data.append(createnamecolresult)
                ws1Data.append(createtimecolresult)
                ws1Data.append(namecolresult)
                ws1Data.append(mothworktimeresult)
                ws1Datas.append(ws1Data)
        else:
            ws1Data=[]
            #9个
            ws1Data.append("任务编号")
            ws1Data.append("任务名称")
            ws1Data.append("预计开始时间")
            ws1Data.append("实际开始时间")
            ws1Data.append("截止日期")
            ws1Data.append("任务状态")
            ws1Data.append("任务创建人")
            ws1Data.append("任务创建时间")
            ws1Data.append("任务责任人")
            ws1Data.append("当月消耗工时")
            ws1Datas.append(ws1Data)

    print("已完成或进行中tasks ws1Datas:",len(ws1Datas),ws1Datas)
    return ws1Datas

if __name__=="__main__":
    # d=getActiviBugsData()
    # get_bug_bySameRespname(d)
    # isHalfYear = getConfig("DEFAULT", "isHalfYear")
    # 1.登录打开网页
    # driver = open_browser_login()
    # # 2.得到迭代的文件--下载到本地（指定目录)
    # # session = getTaskFiles(driver, isHalfYear)
    # # 3.修改csv文件
    # session=_login()
    # getMonthTestExcel(0)
    # getMonthTasksConsume(0)
    # change_task_csv(session)
    # getNoStartTasksExcel()
    # getActiviBugExcel()
    #
    # # 得到未开始或进行中状态的任务数组
    # taskDatas=getNoStartTasksData()
    # # 得到激活状态的bug数组
    # bugDatas=getActiviBugsData()
    #
    # # 得到相同任务负责人下的任务编号和任务名称
    # taskrespNames, taskcreateNames = get_task_bySameRespname(taskDatas)
    # # 得到相同bug负责人下的bug编号和bug标题
    # bugrespNames,bugcreateNames = get_bug_bySameRespname(bugDatas)
    # #
    # get_taskbug_bySameName(taskrespNames,taskcreateNames,bugrespNames,bugcreateNames)

    # getClosedTaskFiles(driver, session, 2)
    #
    # taskDatas=getNoStartTasksData()
    # # 得到激活状态的bug数组
    # bugDatas=getActiviBugsData()
    #
    # # 得到相同任务负责人下的任务编号和任务名称
    # taskrespNames, taskcreateNames=get_task_bySameRespname(taskDatas)
    # # 得到相同bug负责人下的bug编号和bug标题
    # bugrespNames,bugcreateNames=get_bug_bySameRespname(bugDatas)
    # #得到名字相同（负责人、创建人）的任务、bug情况
    # taskbugs=get_taskbug_bySameName(taskrespNames,taskcreateNames,bugrespNames,bugcreateNames)
    # print("{taskbugs:"+str(taskbugs)+"}")
    # print("***********************")
    # for i in taskbugs:
    #     print(i)
    #
    # sqlhost = str(getConfig("mass", "sqlhost"))
    # sqlport = int(getConfig("mass", "sqlport"))
    # sqluser = str(getConfig("mass", "sqluser"))
    # sqlpwd = str(getConfig("mass", "sqlpwd"))
    # sqldb = str(getConfig("mass", "sqldb"))
    # users, userids, useraccounts, usernames, usernamepys = get_user_operate_sql(sqlhost, sqlport, sqluser, sqlpwd,
    #                                                                             sqldb)
    # get_taskbug_image(taskbugs, useraccounts, usernames, usernamepys)
    getMonthTasksConsume2(0)